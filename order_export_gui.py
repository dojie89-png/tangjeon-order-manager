from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from urllib.parse import urlparse, parse_qs, urljoin, quote_plus
import pandas as pd
import requests
import time
import re
import os
import sys
import subprocess
import base64
import html
import ctypes
from ctypes import wintypes
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import traceback


APP_VERSION = "13.66"  # 버전 관리: 소수점 = 기능추가/버그수정, 정수 = 대규모 개편

BASE_URL = os.environ.get("KGINBIO_BASE_URL", "https://www.kginbio.com/admin").rstrip("/")
LOGIN_URL = f"{BASE_URL}/"
ORDER_LIST_BASE_URL = f"{BASE_URL}/order/order_list.asp"
ORDER_DECOCTION_BASE_URL = f"{BASE_URL}/order/order_detail_view.asp"
ORDER_DISPENSE_BASE_URL = f"{BASE_URL}/order/order_detail_view_03.asp"
ORDER_CHANGE_BASE_URL = f"{BASE_URL}/order/order_change.asp"
DELIVERY_INSERT_BASE_URL = f"{BASE_URL}/order/delivery_insert.asp"
SHOP_ORDER_LIST_URL    = f"{BASE_URL}/shop_order/order_list.asp"
SHOP_ORDER_VIEW_URL    = f"{BASE_URL}/shop_order/order_view.asp"
SHOP_ORDER_VIEW22_URL  = f"{BASE_URL}/shop_order/order_view22.asp"
SHOP_ORDER_DELIVERY_INSERT_URL = f"{BASE_URL}/shop_order/delivery_insert.asp"
SHOP_GOODS_LIST_URL    = f"{BASE_URL}/goods/goods_list.asp"
SHOP_GOODS_VIEW_URL    = f"{BASE_URL}/goods/goods_view.asp"
CAFE_GOODS_LIST_URL    = f"{BASE_URL}/cafe_goods/goods_list.asp"
CAFE_GOODS_WRITE_URL   = f"{BASE_URL}/cafe_goods/goods_write.asp"

ADMIN_ID = os.environ.get("KGINBIO_ADMIN_ID", "")
ADMIN_PW = os.environ.get("KGINBIO_ADMIN_PW", "")
WINDOWS_CREDENTIAL_TARGET = "KGINBIO_TANGJEON_ADMIN_LOGIN"

PRINT_ONLY_PREFIXED_OR_ALL = "all"
PRINT_WAIT_SEC = 4.0
MAX_PAGE_SAFETY_LIMIT = 50
DEFAULT_ALL_PERIOD_MAX_PAGE = 7

# ---------- 자동 업데이트 ----------
GITHUB_REPO                    = "dojie89-png/tangjeon-order-manager"
GITHUB_RELEASE_EXE_NAME        = "order_export_gui.exe"
GITHUB_CACHE_FILE_NAME         = "goods_tax_cache.json"
GITHUB_TOKEN_CREDENTIAL_TARGET = "KGINBIO_TANGJEON_GITHUB_TOKEN"


def _startup_error_log_path() -> str:
    try:
        if getattr(sys, "frozen", False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base = os.getcwd()
    return os.path.join(base, "startup_error.log")


def _write_startup_error(exc: BaseException):
    try:
        with open(_startup_error_log_path(), "w", encoding="utf-8") as f:
            f.write(f"APP_VERSION={APP_VERSION}\n")
            f.write(f"Python={sys.version}\n")
            f.write(f"Executable={sys.executable}\n")
            f.write(f"Frozen={getattr(sys, 'frozen', False)}\n\n")
            f.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
    except Exception:
        pass

ALL_STATUSES = [
    "접수대기",
    "입금대기",
    "조제중",
    "탕전중",
    "발송",
    "예약발송",
    "완료",
    "환불취소",
]

# 사이트 order_ings 파라미터 값 매핑 (서버사이드 필터링용)
STATUS_VALUE_MAP = {
    "접수대기": "1",
    "입금대기": "2",
    "조제중": "3",
    "탕전중": "4",
    "발송": "5",
    "예약발송": "8",
    "완료": "6",
    "환불취소": "7",
}
STATUS_TEXT_BY_VALUE = {value: key for key, value in STATUS_VALUE_MAP.items()}

# 한의원 필터 콤보 기본 목록 (직접입력도 가능)
HOSPITAL_PRESETS = [
    "",  # (전체)
    "고래한방_관저점",
    "고래한방_판암점",
    "고래한방_세종점",
    "고래한방_오창점",
    "청주필한방병원",
    "필한방병원",
    "약손한의원",
    "본가한의원",
    "대전굿니스한의원",
    "보강한방병원",
    "개금365한의원",
    "태화당한의원",
]

# 프리셋 표시명 → 실제 검색 키워드 (all-match, 모두 포함돼야 일치)
# 목록에 없는 값은 그대로 단일 키워드로 사용
HOSPITAL_SEARCH_MAP = {
    "고래한방_관저점": ["고래", "관저"],
    "고래한방_판암점": ["고래", "판암"],
    "고래한방_세종점": ["고래", "세종"],
    "고래한방_오창점": ["고래", "오창"],
    "청주필한방병원": ["청주필"],
    "필한방병원": ["필한방병원"],
    "약손한의원": ["약손"],
    "본가한의원": ["본가"],
    "대전굿니스한의원": ["굿니스"],
    "보강한방병원": ["보강"],
    "개금365한의원": ["개금"],
    "태화당한의원": ["태화당"],
}


# ---------- 앱 설정 저장/불러오기 ----------
def _get_config_path() -> str:
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "app_config.json")


def load_app_config() -> dict:
    try:
        with open(_get_config_path(), "r", encoding="utf-8") as f:
            import json
            return json.load(f)
    except Exception:
        return {}


def save_app_config(config: dict):
    try:
        import json
        with open(_get_config_path(), "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ---------- 자동 업데이트 함수 ----------
def _fetch_latest_version() -> str:
    """GitHub Releases에서 최신 버전 태그를 가져옴. 실패 시 빈 문자열 반환."""
    try:
        resp = requests.get(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            timeout=5,
            headers={"Accept": "application/vnd.github+json"},
        )
        if resp.status_code != 200:
            return ""
        tag = resp.json().get("tag_name", "")
        return tag.lstrip("v")
    except Exception:
        return ""


def _version_tuple(version: str) -> tuple:
    parts = re.findall(r"\d+", clean_text(version))
    return tuple(int(p) for p in parts) if parts else (0,)


def is_newer_version(latest: str, current: str) -> bool:
    """13.60 > 13.59 처럼 숫자 기준으로만 업데이트 여부를 판단한다."""
    latest_tuple = _version_tuple(latest)
    current_tuple = _version_tuple(current)
    max_len = max(len(latest_tuple), len(current_tuple))
    latest_tuple += (0,) * (max_len - len(latest_tuple))
    current_tuple += (0,) * (max_len - len(current_tuple))
    return latest_tuple > current_tuple


def _do_update(parent):
    """최신 exe를 다운로드하고 현재 exe를 교체한 뒤 재시작."""
    if not getattr(sys, "frozen", False):
        messagebox.showinfo("업데이트", "개발 환경에서는 자동 업데이트가 지원되지 않아요.\n새 버전은 직접 빌드해주세요.", parent=parent)
        return
    try:
        download_url = (
            f"https://github.com/{GITHUB_REPO}/releases/latest/download/{GITHUB_RELEASE_EXE_NAME}"
        )
        current_exe = sys.executable
        new_exe = current_exe + ".new"

        messagebox.showinfo("업데이트", "다운로드를 시작할게요. 잠시 기다려주세요.", parent=parent)
        resp = requests.get(download_url, stream=True, timeout=60)
        resp.raise_for_status()
        with open(new_exe, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)

        # 현재 exe 교체 후 재시작하는 배치 스크립트
        bat_lines = [
            "@echo off",
            "ping 127.0.0.1 -n 3 > nul",
            f'move /y "{new_exe}" "{current_exe}"',
            f'start "" "{current_exe}"',
            'del "%~f0"',
        ]
        bat_path = current_exe + "_updater.bat"
        with open(bat_path, "w", encoding="cp949") as f:
            f.write("\r\n".join(bat_lines) + "\r\n")

        subprocess.Popen(
            ["cmd", "/c", bat_path],
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )
        sys.exit(0)
    except Exception as e:
        messagebox.showerror("업데이트 실패", f"업데이트 중 오류가 발생했어요:\n{e}", parent=parent)


def check_and_prompt_update(parent):
    """시작 시 업데이트 여부 확인 후 안내 팝업."""
    latest = _fetch_latest_version()
    if not latest or not is_newer_version(latest, APP_VERSION):
        return
    answer = messagebox.askyesno(
        "업데이트 알림",
        f"새 버전이 있어요!\n\n현재: v{APP_VERSION}  →  최신: v{latest}\n\n지금 업데이트할까요?",
        parent=parent,
    )
    if answer:
        _do_update(parent)


# ---------- 공통 유틸 ----------
def clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


class NoWorkFound(Exception):
    """정상 실행됐지만 처리할 주문/항목이 없는 경우."""


def is_user_cancel_message(message: str) -> bool:
    return any(token in clean_text(message) for token in ("취소됨", "작업이 취소", "사용자 취소"))


def sanitize_filename(name: str) -> str:
    name = clean_text(name)
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    return name[:180]


def ensure_parent_dir(file_path: str):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)


def ensure_dir(dir_path: str):
    os.makedirs(dir_path, exist_ok=True)


def has_any_file(dir_path: str) -> bool:
    if not dir_path or not os.path.isdir(dir_path):
        return False
    for _, _, files in os.walk(dir_path):
        if files:
            return True
    return False


def remove_empty_dir_tree(dir_path: str):
    if dir_path and os.path.isdir(dir_path) and not has_any_file(dir_path):
        import shutil
        shutil.rmtree(dir_path, ignore_errors=True)


def normalize_date_input(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ""
    # "2026-04-01 14:30" 또는 "2026-04-01 14:30:00" 형식
    m = re.match(r"^(\d{4}[^\d]\d{2}[^\d]\d{2})\s+(\d{2}:\d{2}(?::\d{2})?)$", value)
    if m:
        d = re.sub(r"[^0-9]", "", m.group(1))
        return f"{d[:4]}-{d[4:6]}-{d[6:8]} {m.group(2)}"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return value
    digits = re.sub(r"[^0-9]", "", value)
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
    if len(digits) == 12:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]} {digits[8:10]}:{digits[10:12]}"
    if len(digits) >= 14:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]} {digits[8:10]}:{digits[10:12]}:{digits[12:14]}"
    raise ValueError("날짜는 20260401, 2026-04-01, 20260401 1430, 또는 2026-04-01 14:30 형식으로 입력해줘.")


def td_text_without_form_controls(td) -> str:
    td_copy = BeautifulSoup(str(td), "html.parser")
    for tag in td_copy.find_all(["input", "select", "option", "script", "style", "textarea"]):
        tag.decompose()
    text = td_copy.get_text(" ", strip=True)
    text = text.replace("[변경]", "").replace("[입력]", "")
    return clean_text(text)


def get_selected_option_text(soup, name: str) -> str:
    sel = soup.find("select", attrs={"name": name})
    if not sel:
        return ""
    selected = sel.find("option", selected=True)
    if selected:
        return clean_text(selected.get_text())
    return ""


def get_selected_option_value(soup, name: str) -> str:
    sel = soup.find("select", attrs={"name": name})
    if not sel:
        return ""
    selected = sel.find("option", selected=True)
    if selected:
        return clean_text(selected.get("value", ""))
    return ""


def get_input_value(soup, name: str) -> str:
    tag = soup.find("input", attrs={"name": name})
    if not tag:
        return ""
    return clean_text(tag.get("value", ""))


def get_textarea_value(soup, name: str) -> str:
    tag = soup.find("textarea", attrs={"name": name})
    if not tag:
        return ""
    return clean_text(tag.get_text())


def split_gender_age(value: str):
    value = clean_text(value)
    if "/" in value:
        a, b = value.split("/", 1)
        return clean_text(a), clean_text(b)
    return value, ""


def split_phone_mobile(value: str):
    value = clean_text(value)
    if not value:
        return "", ""
    if "/" in value:
        a, b = value.split("/", 1)
        return clean_text(a), clean_text(b)
    if value.startswith(("010", "011", "016", "017", "018", "019")):
        return "", value
    return value, ""


def split_zip_address(value: str):
    value = clean_text(value)
    if not value:
        return "", ""
    m = re.match(r"^(\d{5})\s+(.*)$", value)
    if m:
        return m.group(1), clean_text(m.group(2))
    return "", value


def tr_to_pairs(tr):
    pairs = []
    tds = tr.find_all("td")
    if len(tds) == 4:
        pairs.append((clean_text(tds[0].get_text(" ", strip=True)), td_text_without_form_controls(tds[1])))
        pairs.append((clean_text(tds[2].get_text(" ", strip=True)), td_text_without_form_controls(tds[3])))
    elif len(tds) == 2:
        pairs.append((clean_text(tds[0].get_text(" ", strip=True)), td_text_without_form_controls(tds[1])))
    return pairs


def extract_card_info_from_text(body_text: str):
    text = clean_text(body_text)
    trade_no = ""
    card_order_no = ""
    approval_no = ""
    m_section = re.search(r"카드결제 정보.*?거래번호(.*?)(?:약재비|진행상태|배송|보내시는분|받으시는분)", text)
    section = m_section.group(1) if m_section else text
    m1 = re.search(r"거래번호\s*(.*?)\s*주문번호", section)
    if m1:
        trade_no = clean_text(m1.group(1))
    m2 = re.search(r"주문번호\s*(.*?)\s*승인번호", section)
    if m2:
        card_order_no = clean_text(m2.group(1))
    m3 = re.search(r"승인번호\s*(.*)", section)
    if m3:
        approval_no = clean_text(m3.group(1))
    for label in ["약재비", "진행상태", "배송", "보내시는분", "받으시는분"]:
        if label in approval_no:
            approval_no = approval_no.split(label)[0].strip()
    return trade_no, card_order_no, approval_no


def to_number(value: str):
    value = clean_text(value)
    if not value:
        return None
    value = value.replace("원", "").replace(",", "").strip()
    m = re.search(r"-?\d+(\.\d+)?", value)
    if not m:
        return None
    num_str = m.group()
    try:
        return float(num_str) if "." in num_str else int(num_str)
    except Exception:
        return None


def normalize_datetime(value: str):
    value = clean_text(value)
    if not value:
        return ""
    value = value.replace("오전", "AM").replace("오후", "PM")
    return value


def format_order_datetime_for_filename(value: str) -> str:
    value = clean_text(value)
    if not value:
        return "날짜없음"
    # parse_order_datetime_obj 는 오전/오후·AM/PM 을 로케일 독립적으로 파싱 (strptime %p 한국 로케일 오작동 방지)
    # ※ parse_order_datetime_obj 는 이 함수보다 아래에 정의되지만 Python 은 호출 시점에 이름을 해석하므로 정상
    dt = parse_order_datetime_obj(value)
    if dt:
        return dt.strftime("%Y-%m-%d_%H-%M-%S")
    # fallback: 숫자만 추출
    nums = re.sub(r"[^0-9]", "", value)
    if len(nums) >= 14:
        return f"{nums[0:4]}-{nums[4:6]}-{nums[6:8]}_{nums[8:10]}-{nums[10:12]}-{nums[12:14]}"
    if len(nums) >= 8:
        return f"{nums[0:4]}-{nums[4:6]}-{nums[6:8]}_00-00-00"
    return sanitize_filename(value)


def parse_filter_datetime(datetime_str: str):
    """날짜+시간 문자열을 datetime 객체로 변환. 시간 없으면 None 반환."""
    if not datetime_str or " " not in datetime_str:
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
        try:
            return datetime.strptime(datetime_str, fmt)
        except Exception:
            pass
    return None


def parse_order_datetime_obj(value: str):
    """주문날짜 문자열을 datetime 객체로 변환. locale 독립적 AM/PM 처리."""
    value = clean_text(value)
    if not value:
        return None
    # %p 는 시스템 locale 의존적이므로 직접 파싱
    # 형식: "2026-04-14 오전 11:14", "2026-04-14 AM 11:14" 등
    m = re.match(
        r'^(\d{4}-\d{2}-\d{2})\s+(오전|오후|AM|PM)\s+(\d{1,2}):(\d{2})(?::(\d{2}))?$',
        value,
    )
    if m:
        date_str, ampm, h_s, mn_s, sec_s = m.groups()
        h, mn, s = int(h_s), int(mn_s), int(sec_s) if sec_s else 0
        is_pm = ampm in ('오후', 'PM')
        if is_pm and h != 12:
            h += 12
        elif not is_pm and h == 12:
            h = 0  # 12 AM = 자정
        try:
            base = datetime.strptime(date_str, '%Y-%m-%d')
            return base.replace(hour=h, minute=mn, second=s)
        except Exception:
            pass
    # 24시간 형식 fallback
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except Exception:
            pass
    return None


def extract_datetime_text(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ""
    patterns = [
        r"\d{4}[-/.]\d{2}[-/.]\d{2}\s+(?:오전|오후|AM|PM)\s+\d{1,2}:\d{2}(?::\d{2})?",
        r"\d{4}[-/.]\d{2}[-/.]\d{2}\s+\d{1,2}:\d{2}(?::\d{2})?",
        r"\d{4}[-/.]\d{2}[-/.]\d{2}",
    ]
    for pattern in patterns:
        match = re.search(pattern, value, flags=re.IGNORECASE)
        if match:
            return match.group(0).replace("/", "-").replace(".", "-")
    return ""


def normalize_tracking_no(value: str) -> str:
    return re.sub(r"[-\s]", "", clean_text(value))


def build_order_list_url(page_no: int, start_date: str = "", end_date: str = "", order_ings: str = "") -> str:
    s_date = quote_plus(start_date[:10] if start_date else "")
    e_date = quote_plus(end_date[:10] if end_date else "")
    return (
        f"{ORDER_LIST_BASE_URL}?"
        f"s_date={s_date}&e_date={e_date}&order_ings={order_ings}&part=&page={page_no}&search=&s_string="
    )


def split_list_prescription_display(cell_text: str, link_text: str):
    cell_text = clean_text(cell_text)
    link_text = clean_text(link_text)
    if not cell_text:
        return link_text, ""
    if link_text and cell_text.startswith(link_text):
        extra = clean_text(cell_text[len(link_text):])
        return link_text, extra
    if link_text and link_text in cell_text:
        extra = clean_text(cell_text.replace(link_text, "", 1))
        return link_text, extra
    return cell_text, ""


def extract_seqno_from_detail_url(detail_url: str) -> str:
    try:
        parsed = urlparse(detail_url)
        qs = parse_qs(parsed.query)
        seqno = qs.get("seqno", [""])[0]
        return clean_text(seqno)
    except Exception:
        m = re.search(r"seqno=([^&]+)", detail_url)
        return m.group(1) if m else ""


def build_decoction_url_from_detail_url(detail_url: str) -> str:
    seqno = extract_seqno_from_detail_url(detail_url)
    if not seqno:
        return ""
    return f"{ORDER_DECOCTION_BASE_URL}?seqno={seqno}"


def build_dispense_url_from_detail_url(detail_url: str) -> str:
    seqno = extract_seqno_from_detail_url(detail_url)
    if not seqno:
        return ""
    return f"{ORDER_DISPENSE_BASE_URL}?seqno={seqno}"


def get_direct_rows(table):
    tbody = table.find("tbody")
    if tbody:
        return tbody.find_all("tr", recursive=False)
    return table.find_all("tr", recursive=False)


def get_direct_cells(tr):
    return tr.find_all(["td", "th"], recursive=False)


def cell_text(cell):
    return clean_text(cell.get_text(" ", strip=True))


def find_main_decoction_table(soup):
    tables = soup.find_all("table")
    for table in tables:
        text = clean_text(table.get_text(" ", strip=True))
        if "회원명(처방자)" in text and "복용자" in text and "주문상품" in text:
            return table
    return None


def find_total_cost_table(main_table):
    nested_tables = main_table.find_all("table")
    for table in nested_tables:
        text = clean_text(table.get_text(" ", strip=True))
        if "약재비" in text and "탕전비" in text and "총결제" in text and "배송비" in text:
            rows = get_direct_rows(table)
            if len(rows) >= 2:
                headers = [cell_text(td) for td in get_direct_cells(rows[0])]
                if headers[:3] == ["약재비", "탕전비", "주수상반"]:
                    return table
    return None


def extract_all_anchor_info_from_cell(cell, base_url=""):
    if cell is None:
        return "", ""
    a = cell.find("a", href=True)
    if not a:
        return "", ""
    href = clean_text(a.get("href", ""))
    text = clean_text(a.get_text(" ", strip=True))
    if not href:
        return "", text
    if href.lower().startswith("javascript"):
        return "", text
    if href.startswith("http://") or href.startswith("https://"):
        return href, text
    return urljoin(base_url, href), text


def save_text_as_pdf(driver, title: str, body_text: str, save_path: str):
    safe_title = html.escape(title or "")
    safe_body = html.escape(body_text or "").replace("\n", "<br>")
    html_doc = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{
                font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif;
                padding: 32px;
                line-height: 1.7;
                font-size: 14px;
                white-space: normal;
                word-break: break-word;
            }}
            h1 {{
                font-size: 20px;
                margin-bottom: 20px;
            }}
            .content {{
                white-space: pre-wrap;
            }}
        </style>
    </head>
    <body>
        <h1>{safe_title}</h1>
        <div class="content">{safe_body}</div>
    </body>
    </html>
    """
    encoded = base64.b64encode(html_doc.encode("utf-8")).decode("ascii")
    driver.get(f"data:text/html;base64,{encoded}")
    time.sleep(0.7)
    pdf_data = driver.execute_cdp_cmd(
        "Page.printToPDF",
        {
            "printBackground": True,
            "paperWidth": 8.27,
            "paperHeight": 11.69,
            "marginTop": 0.3,
            "marginBottom": 0.3,
            "marginLeft": 0.3,
            "marginRight": 0.3,
            "preferCSSPageSize": True,
        },
    )
    ensure_parent_dir(save_path)
    with open(save_path, "wb") as f:
        f.write(base64.b64decode(pdf_data["data"]))


# ---------- 병원 폴더명 ----------
# 고래한방병원 지점별 담당 회원명(주문자명) 매핑
GORAE_MEMBER_BRANCH_MAP: dict = {
    "정재희": "관저",
    "최혁준": "판암",
    "박중현": "세종",
    "이형배": "오창",
}


def extract_branch_keyword(hospital_name: str, address: str, member_name: str = "") -> str:
    """주소에서 지점 키워드를 추출. 주소에서 못 찾으면 회원명으로 fallback."""
    hospital_name = clean_text(hospital_name)
    address = clean_text(address)
    member_name = clean_text(member_name)
    special_branch_rules = {
        "고래한방병원": ["판암", "세종", "관저", "오창"],
    }
    if hospital_name in special_branch_rules:
        # 1차: 주소에서 키워드 검색
        for keyword in special_branch_rules[hospital_name]:
            if keyword in address:
                return keyword
        # 2차: 회원명(담당 의사)으로 지점 매핑
        if member_name and member_name in GORAE_MEMBER_BRANCH_MAP:
            return GORAE_MEMBER_BRANCH_MAP[member_name]
        return "기타"
    return ""


def build_hospital_folder_name(hospital_name: str, address: str, member_name: str = "") -> str:
    hospital_name_clean = clean_text(hospital_name or "한의원명없음")
    branch = extract_branch_keyword(hospital_name_clean, address, member_name)
    if branch:
        return sanitize_filename(f"{hospital_name_clean}_{branch}")
    return sanitize_filename(hospital_name_clean)


# ---------- 입원 판정 ----------
def is_inpatient_dispense(hospital_folder_name: str, dispensing_note: str) -> bool:
    hospital_folder_name = clean_text(hospital_folder_name)
    dispensing_note = clean_text(dispensing_note)
    return hospital_folder_name.startswith("고래한방병원_") and "입원" in dispensing_note


# ---------- 벌크 판정 ----------
BULK_MAX_PACKS = 30  # 팩수 미만이면 벌크 (30개 이상은 별도 박스 포장)

def get_box_capacity(clinic: str, box_type: str) -> int | None:
    """박스 유형·한의원 기준 박스당 최대 포수. None이면 분할 안 함.
    고래한방은 항상 전용박스 용량 적용 (오창 포함, 박스포장 무시).
    본가한의원은 박스포장 무관, 항상 30포 단위로 분할.
    """
    c = clean_text(str(clinic or ""))
    b = clean_text(str(box_type or ""))
    if "고래" in c:
        if "관저" in c or "판암" in c:
            return 40
        return 30  # 세종, 오창 등
    if "본가" in c:
        return 30  # 본가한의원: 박스포장 무관 30포 단위
    if "고급박스2" in b or "고급 박스 2" in b:
        return 70
    if "고급박스" in b:  # 고급박스1 또는 "고급박스"만 적힌 경우
        return 30
    return None  # 박스포장 없거나 인식 불가 → 분할 안 함


def is_bulk_delivery(delivery_type: str, pack_count,
                     hospital_name: str = "", combined_address: str = "") -> bool:
    """고래한방 + 한의원으로 택배 + 팩수<30 → 벌크 대상"""
    if "한의원으로 택배" not in clean_text(delivery_type):
        return False
    if "고래" not in clean_text(hospital_name):
        return False
    try:
        packs = int(str(pack_count).strip())
    except (ValueError, TypeError):
        return False
    if packs <= 0:
        return False
    if packs < BULK_MAX_PACKS:
        return True
    return False


# ---------- 프린트 대상 판정 ----------
def should_print_dispense(is_inpatient: bool) -> bool:
    if PRINT_ONLY_PREFIXED_OR_ALL == "all":
        return True
    if PRINT_ONLY_PREFIXED_OR_ALL == "prefixed_only":
        return is_inpatient
    return True


# ---------- 로그인 ----------
if sys.platform == "win32":
    class _CREDENTIALW(ctypes.Structure):
        _fields_ = [
            ("Flags", wintypes.DWORD),
            ("Type", wintypes.DWORD),
            ("TargetName", wintypes.LPWSTR),
            ("Comment", wintypes.LPWSTR),
            ("LastWritten", wintypes.FILETIME),
            ("CredentialBlobSize", wintypes.DWORD),
            ("CredentialBlob", ctypes.c_void_p),
            ("Persist", wintypes.DWORD),
            ("AttributeCount", wintypes.DWORD),
            ("Attributes", ctypes.c_void_p),
            ("TargetAlias", wintypes.LPWSTR),
            ("UserName", wintypes.LPWSTR),
        ]

    _PCREDENTIALW = ctypes.POINTER(_CREDENTIALW)
    _CRED_TYPE_GENERIC = 1
    _CRED_PERSIST_LOCAL_MACHINE = 2


def _read_windows_admin_credential():
    if sys.platform != "win32":
        return "", ""
    advapi32 = ctypes.windll.advapi32
    advapi32.CredReadW.argtypes = [
        wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD, ctypes.POINTER(_PCREDENTIALW)
    ]
    advapi32.CredReadW.restype = wintypes.BOOL
    advapi32.CredFree.argtypes = [ctypes.c_void_p]
    cred_ptr = _PCREDENTIALW()
    if not advapi32.CredReadW(WINDOWS_CREDENTIAL_TARGET, _CRED_TYPE_GENERIC, 0, ctypes.byref(cred_ptr)):
        return "", ""
    try:
        cred = cred_ptr.contents
        admin_id = cred.UserName or ""
        blob = ctypes.string_at(cred.CredentialBlob, cred.CredentialBlobSize)
        admin_pw = blob.decode("utf-16-le") if blob else ""
        return clean_text(admin_id), admin_pw
    finally:
        advapi32.CredFree(cred_ptr)


def _write_windows_admin_credential(admin_id: str, admin_pw: str):
    if sys.platform != "win32":
        return False
    advapi32 = ctypes.windll.advapi32
    advapi32.CredWriteW.argtypes = [ctypes.POINTER(_CREDENTIALW), wintypes.DWORD]
    advapi32.CredWriteW.restype = wintypes.BOOL
    blob = admin_pw.encode("utf-16-le")
    cred = _CREDENTIALW()
    cred.Type = _CRED_TYPE_GENERIC
    cred.TargetName = WINDOWS_CREDENTIAL_TARGET
    cred.CredentialBlobSize = len(blob)
    blob_buf = ctypes.create_string_buffer(blob)
    cred.CredentialBlob = ctypes.cast(blob_buf, ctypes.c_void_p)
    cred.Persist = _CRED_PERSIST_LOCAL_MACHINE
    cred.UserName = admin_id
    return bool(advapi32.CredWriteW(ctypes.byref(cred), 0))


def _delete_windows_admin_credential():
    if sys.platform != "win32":
        return
    try:
        advapi32 = ctypes.windll.advapi32
        advapi32.CredDeleteW.argtypes = [wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD]
        advapi32.CredDeleteW.restype = wintypes.BOOL
        advapi32.CredDeleteW(WINDOWS_CREDENTIAL_TARGET, _CRED_TYPE_GENERIC, 0)
    except Exception:
        pass


def _read_github_token() -> str:
    """Windows 자격증명에서 GitHub PAT 읽기."""
    if sys.platform != "win32":
        return ""
    try:
        advapi32 = ctypes.windll.advapi32
        advapi32.CredReadW.argtypes = [
            wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD, ctypes.POINTER(_PCREDENTIALW)
        ]
        advapi32.CredReadW.restype = wintypes.BOOL
        advapi32.CredFree.argtypes = [ctypes.c_void_p]
        cred_ptr = _PCREDENTIALW()
        if not advapi32.CredReadW(GITHUB_TOKEN_CREDENTIAL_TARGET, _CRED_TYPE_GENERIC, 0, ctypes.byref(cred_ptr)):
            return ""
        try:
            cred = cred_ptr.contents
            blob = ctypes.string_at(cred.CredentialBlob, cred.CredentialBlobSize)
            return blob.decode("utf-16-le") if blob else ""
        finally:
            advapi32.CredFree(cred_ptr)
    except Exception:
        return ""


def _write_github_token(token: str) -> bool:
    """Windows 자격증명에 GitHub PAT 저장."""
    if sys.platform != "win32":
        return False
    try:
        advapi32 = ctypes.windll.advapi32
        advapi32.CredWriteW.argtypes = [ctypes.POINTER(_CREDENTIALW), wintypes.DWORD]
        advapi32.CredWriteW.restype = wintypes.BOOL
        blob = token.encode("utf-16-le")
        cred = _CREDENTIALW()
        cred.Type = _CRED_TYPE_GENERIC
        cred.TargetName = GITHUB_TOKEN_CREDENTIAL_TARGET
        cred.CredentialBlobSize = len(blob)
        blob_buf = ctypes.create_string_buffer(blob)
        cred.CredentialBlob = ctypes.cast(blob_buf, ctypes.c_void_p)
        cred.Persist = _CRED_PERSIST_LOCAL_MACHINE
        cred.UserName = "github"
        return bool(advapi32.CredWriteW(ctypes.byref(cred), 0))
    except Exception:
        return False


def forget_admin_credentials():
    global ADMIN_ID, ADMIN_PW
    ADMIN_ID = ""
    ADMIN_PW = ""
    _delete_windows_admin_credential()


def _prompt_admin_credentials(parent):
    dialog = tk.Toplevel(parent)
    dialog.title("케이진 관리자 로그인")
    dialog.resizable(False, False)
    dialog.transient(parent)
    dialog.grab_set()

    frame = ttk.Frame(dialog, padding=14)
    frame.grid(row=0, column=0, sticky="nsew")
    ttk.Label(frame, text="관리자 로그인 정보를 입력해주세요.").grid(
        row=0, column=0, columnspan=2, sticky="w", pady=(0, 10)
    )
    ttk.Label(frame, text="아이디").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(0, 6))
    id_var = tk.StringVar(value=ADMIN_ID or os.environ.get("KGINBIO_ADMIN_ID", ""))
    id_entry = ttk.Entry(frame, width=28, textvariable=id_var)
    id_entry.grid(row=1, column=1, sticky="ew", pady=(0, 6))
    ttk.Label(frame, text="비밀번호").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=(0, 6))
    pw_var = tk.StringVar(value="")
    pw_entry = ttk.Entry(frame, width=28, textvariable=pw_var, show="*")
    pw_entry.grid(row=2, column=1, sticky="ew", pady=(0, 6))
    save_var = tk.BooleanVar(value=(sys.platform == "win32"))
    save_cb = ttk.Checkbutton(frame, text="이 PC에 로그인 정보 저장", variable=save_var)
    save_cb.grid(row=3, column=1, sticky="w", pady=(0, 10))
    if sys.platform != "win32":
        save_cb.config(state="disabled")
    status_var = tk.StringVar(value="")
    ttk.Label(frame, textvariable=status_var, foreground="red").grid(
        row=4, column=0, columnspan=2, sticky="w", pady=(0, 8)
    )
    result = {"ok": False, "id": "", "pw": "", "save": False}

    def submit():
        admin_id = clean_text(id_var.get())
        admin_pw = pw_var.get()
        if not admin_id or not admin_pw:
            status_var.set("아이디와 비밀번호를 모두 입력해주세요.")
            return
        result.update({"ok": True, "id": admin_id, "pw": admin_pw, "save": save_var.get()})
        dialog.destroy()

    def cancel():
        dialog.destroy()

    btns = ttk.Frame(frame)
    btns.grid(row=5, column=0, columnspan=2, sticky="ew")
    ttk.Button(btns, text="확인", command=submit).pack(side="right")
    ttk.Button(btns, text="취소", command=cancel).pack(side="right", padx=(0, 6))
    dialog.bind("<Return>", lambda _e: submit())
    dialog.bind("<Escape>", lambda _e: cancel())
    id_entry.focus_set()
    parent.wait_window(dialog)
    return result


def ensure_admin_credentials(parent) -> bool:
    global ADMIN_ID, ADMIN_PW
    if ADMIN_ID and ADMIN_PW:
        return True
    env_id = os.environ.get("KGINBIO_ADMIN_ID", "")
    env_pw = os.environ.get("KGINBIO_ADMIN_PW", "")
    if env_id and env_pw:
        ADMIN_ID, ADMIN_PW = clean_text(env_id), env_pw
        return True
    saved_id, saved_pw = _read_windows_admin_credential()
    if saved_id and saved_pw:
        ADMIN_ID, ADMIN_PW = saved_id, saved_pw
        return True
    result = _prompt_admin_credentials(parent)
    if not result["ok"]:
        return False
    ADMIN_ID, ADMIN_PW = result["id"], result["pw"]
    if result["save"]:
        try:
            if not _write_windows_admin_credential(ADMIN_ID, ADMIN_PW):
                messagebox.showwarning("저장 실패", "Windows 자격 증명 관리자에 로그인 정보를 저장하지 못했어요.")
        except Exception as e:
            messagebox.showwarning("저장 실패", f"로그인 정보 저장 중 오류가 발생했어요:\n{e}")
    return True


def is_login_page(driver) -> bool:
    try:
        id_inputs = driver.find_elements(By.NAME, "id")
        pw_inputs = driver.find_elements(By.NAME, "pwd")
        return bool(id_inputs and pw_inputs)
    except Exception:
        return False


def login_driver(driver, admin_id: str, admin_pw: str):
    wait_local = WebDriverWait(driver, 10)
    driver.get(LOGIN_URL)
    id_input = wait_local.until(EC.presence_of_element_located((By.NAME, "id")))
    pw_input = driver.find_element(By.NAME, "pwd")
    login_btn = driver.find_element(By.CSS_SELECTOR, "input[type='image']")
    id_input.clear()
    id_input.send_keys(admin_id)
    pw_input.clear()
    pw_input.send_keys(admin_pw)
    login_btn.click()
    time.sleep(2)
    if is_login_page(driver):
        forget_admin_credentials()
        raise Exception("로그인 실패: 아이디 또는 비밀번호를 확인해주세요.")


# ---------- 주문마스터 ----------
def parse_detail_html(html: str, detail_url: str = "", collect_index: int = 0, page_no: int = 1) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    trs = soup.find_all("tr")
    basic = {}
    sender = {}
    receiver = {}
    section = "basic"
    for tr in trs:
        row_text = clean_text(tr.get_text(" ", strip=True))
        if row_text == "보내시는분":
            section = "sender"
            continue
        elif row_text == "받으시는분":
            section = "receiver"
            continue
        pairs = tr_to_pairs(tr)
        if not pairs:
            continue
        if section == "basic":
            for k, v in pairs:
                if k:
                    basic[k] = v
        elif section == "sender":
            for k, v in pairs:
                if k:
                    sender[k] = v
        elif section == "receiver":
            for k, v in pairs:
                if k:
                    receiver[k] = v

    gender, age = split_gender_age(basic.get("성별/나이", ""))
    rep_tel, rep_mobile = split_phone_mobile(basic.get("연락처", ""))
    sender_tel, sender_mobile = split_phone_mobile(sender.get("연락처", ""))
    receiver_tel, receiver_mobile = split_phone_mobile(receiver.get("연락처", ""))
    sender_zip, sender_addr = split_zip_address(sender.get("주소", ""))
    receiver_zip, receiver_addr = split_zip_address(receiver.get("주소", ""))
    ordercode = get_input_value(soup, "ordercode")
    receiver_handphone_hidden = get_input_value(soup, "handphone")
    deposit_status = get_selected_option_text(soup, "payment")
    tax_invoice_status = get_selected_option_text(soup, "cash_bill")
    progress_status = get_selected_option_text(soup, "order_ing")
    courier_name = get_selected_option_text(soup, "tak_sel")
    delivery_no = get_input_value(soup, "delivery_no")
    realprice_memo = get_textarea_value(soup, "realprice_memo")
    body_text = soup.get_text(" ", strip=True)
    card_trade_no, card_order_no, card_approval_no = extract_card_info_from_text(body_text)

    return {
        "수집순번": collect_index,
        "페이지": page_no,
        "주문코드": ordercode,
        "회원명": basic.get("회원명", ""),
        "한의원명": basic.get("한의원명", ""),
        "환자명": basic.get("환자명", ""),
        "성별": gender,
        "나이": age,
        "대표연락처_전화": rep_tel,
        "대표연락처_휴대폰": rep_mobile,
        "이메일": basic.get("이메일", ""),
        "처방명": basic.get("처방명", "").replace("[탕전주문내역서]", "").strip(),
        "처방명_목록": "",
        "처방명_목록추가표시": "",
        "주문날짜": normalize_datetime(basic.get("주문날짜", "")),
        "처방번호": basic.get("처방번호", "").replace("[조제지시서]", "").strip(),
        "결제방법": basic.get("결제방법", ""),
        "입금상태": deposit_status,
        "영수증신청": basic.get("영수증신청", ""),
        "계산서/영수증상태": tax_invoice_status,
        "결제금액": to_number(basic.get("결제금액", "")),
        "약재비": to_number(basic.get("약재비", "")),
        "탕전비": to_number(basic.get("탕전비", "")),
        "배송비": to_number(basic.get("배송비", "")),
        "총비용": to_number(basic.get("총비용", "")),
        "할인금액": to_number(basic.get("할인금액", "")),
        "총결제비용": to_number(basic.get("총결제비용", "")),
        "진행상태": progress_status,
        "배송구분": basic.get("배송", ""),
        "택배사": courier_name,
        "송장번호": delivery_no,
        "카드거래번호": card_trade_no,
        "카드주문번호": card_order_no,
        "카드승인번호": card_approval_no,
        "보내는분": sender.get("이름", ""),
        "보내는분_전화": sender_tel,
        "보내는분_휴대폰": sender_mobile,
        "보내는분_우편번호": sender_zip,
        "보내는분_주소": sender_addr,
        "받는분": receiver.get("이름", ""),
        "받는분_전화": receiver_tel,
        "받는분_휴대폰": receiver_mobile or receiver_handphone_hidden,
        "받는분_우편번호": receiver_zip,
        "받는분_주소": receiver_addr,
        "배송시메모": receiver.get("배송시메모", ""),
        "주문시요청사항": receiver.get("주문시요청사항", ""),
        "관리자메모": realprice_memo,
        "상세URL": detail_url,
    }


# ---------- 주문내역서 ----------
def parse_decoction_html(html_text: str, source_row: dict, decoction_url: str) -> dict:
    soup = BeautifulSoup(html_text, "html.parser")
    body_text = clean_text(soup.get_text(" ", strip=True))
    main_table = find_main_decoction_table(soup)
    if not main_table:
        return {
            "수집순번": source_row.get("수집순번", 0),
            "주문코드": source_row.get("주문코드", ""),
            "회원명": source_row.get("회원명", ""),
            "환자명": source_row.get("환자명", ""),
            "처방명": source_row.get("처방명", ""),
            "주문날짜": source_row.get("주문날짜", ""),
            "탕전주문내역서_원문텍스트": body_text,
            "탕전주문내역서URL": decoction_url,
        }

    rows = get_direct_rows(main_table)

    def row_cells(idx):
        if idx >= len(rows):
            return []
        return get_direct_cells(rows[idx])

    def row_cells_text(idx):
        return [cell_text(td) for td in row_cells(idx)]

    r1 = row_cells_text(0)
    r2 = row_cells_text(1)
    r3 = row_cells_text(2)
    r4 = row_cells_text(3)
    r5 = row_cells_text(4)
    r6 = row_cells_text(5)
    r7 = row_cells_text(6)
    r8 = row_cells_text(7)
    r9 = row_cells_text(8)

    practitioner_name = r1[1] if len(r1) > 1 else ""
    clinic_name = r1[3] if len(r1) > 3 else ""
    taker_name = r2[1] if len(r2) > 1 else ""
    gender_age_raw = r2[3] if len(r2) > 3 else ""
    gender, age = split_gender_age(gender_age_raw)
    if not age:
        age = source_row.get("나이", "")
    order_no = r3[1] if len(r3) > 1 else ""
    order_date_doc = normalize_datetime(r3[3] if len(r3) > 3 else "")
    order_product = r4[1] if len(r4) > 1 else ""
    order_product = re.sub(r"\[[^\]]*\]", "", order_product).strip()

    r5_cells = row_cells(4)
    dose_input_map = {"c1": "", "c2": "", "c3": "", "c4": ""}
    for cell in r5_cells:
        for tag in cell.find_all("input"):
            name = clean_text(tag.get("name", ""))
            value = clean_text(tag.get("value", ""))
            if name in dose_input_map:
                dose_input_map[name] = value

    c1 = dose_input_map["c1"]
    c3 = dose_input_map["c3"]
    dose_count_text = r5[1] if len(r5) > 1 else ""
    dose_count = c1 or c3 or dose_count_text
    pack_count = r6[1] if len(r6) > 1 else ""
    pouch_volume = r5[3] if len(r5) > 3 else ""
    decoction_method = r7[1] if len(r7) > 1 else ""
    bundle_delivery = r6[3] if len(r6) > 3 else ""
    shipping_fee_type = r7[3] if len(r7) > 3 else ""
    jusu = r8[1] if len(r8) > 1 else ""
    styrofoam = r8[3] if len(r8) > 3 else ""
    box_pack = r9[1] if len(r9) > 1 else ""
    pouch_pack = r9[3] if len(r9) > 3 else ""

    total_cost_table = find_total_cost_table(main_table)
    herb_cost = decoction_cost = jusu_cost = distilled_cost = shipping_cost = discount_cost = total_paid = None

    if total_cost_table:
        cost_rows = get_direct_rows(total_cost_table)
        if len(cost_rows) >= 2:
            headers = [cell_text(td) for td in get_direct_cells(cost_rows[0])]
            values = [cell_text(td) for td in get_direct_cells(cost_rows[1])]
            cost_map = {}
            for i, h in enumerate(headers):
                v = values[i] if i < len(values) else ""
                cost_map[h] = to_number(v)
            herb_cost = cost_map.get("약재비")
            decoction_cost = cost_map.get("탕전비")
            jusu_cost = cost_map.get("주수상반")
            distilled_cost = cost_map.get("증류탕전")
            shipping_cost = cost_map.get("배송비")
            discount_cost = cost_map.get("할인금액")
            total_paid = cost_map.get("총결제")

    dispensing_note = dispensing_file = dispensing_file_url = ""
    dosage = dosage_text_only = dosage_file = dosage_file_url = dosage_file_name = ""
    shipping_from = shipping_to = delivery_memo = order_request = admin_memo = ""

    for tr in rows:
        cells = get_direct_cells(tr)
        texts = [cell_text(td) for td in cells]
        if len(texts) < 2:
            continue
        label = texts[0]
        value = texts[1] if len(texts) > 1 else ""
        value_cell = cells[1] if len(cells) > 1 else None
        if label == "조제지시사항":
            dispensing_note = value
        elif label == "조제첨부파일":
            dispensing_file = value
            dispensing_file_url, _ = extract_all_anchor_info_from_cell(value_cell, decoction_url)
        elif label == "복용법":
            dosage = value
            dosage_text_only = clean_text(value.replace("[프린트]", ""))
        elif label == "복용첨부파일":
            dosage_file = value
            dosage_file_url, dosage_file_name = extract_all_anchor_info_from_cell(value_cell, decoction_url)
        elif label == "발송정보":
            shipping_from = value
        elif label == "배송정보":
            shipping_to = value
        elif label == "배송시메모":
            delivery_memo = value
        elif label == "주문시요청사항":
            order_request = value
        elif label == "관리자메모":
            admin_memo = value

    return {
        "수집순번": source_row.get("수집순번", 0),
        "주문코드": source_row.get("주문코드", ""),
        "회원명": source_row.get("회원명", ""),
        "환자명": source_row.get("환자명", ""),
        "처방명": source_row.get("처방명", ""),
        "주문날짜": source_row.get("주문날짜", ""),
        "회원명(처방자)": practitioner_name,
        "한의원명": clinic_name,
        "복용자": taker_name,
        "성별": gender,
        "나이": age,
        "주문번호": order_no,
        "주문날짜_내역서": order_date_doc,
        "주문상품": order_product,
        "첩수": dose_count,
        "팩수": pack_count,
        "파우치용량": pouch_volume,
        "탕전방식": decoction_method,
        "묶음배송": bundle_delivery,
        "배송비구분": shipping_fee_type,
        "주수상반": jusu,
        "스티로폼포장": styrofoam,
        "박스포장": box_pack,
        "파우치포장": pouch_pack,
        "약재비": herb_cost,
        "탕전비": decoction_cost,
        "주수상반비용": jusu_cost,
        "증류탕전비용": distilled_cost,
        "배송비": shipping_cost,
        "할인금액": discount_cost,
        "총결제": total_paid,
        "조제지시사항": dispensing_note,
        "조제첨부파일": dispensing_file,
        "조제첨부파일URL": dispensing_file_url,
        "복용법": dosage,
        "복용법_본문": dosage_text_only,
        "복용첨부파일": dosage_file,
        "복용첨부파일URL": dosage_file_url,
        "복용첨부파일명": dosage_file_name,
        "발송정보": shipping_from,
        "배송정보": shipping_to,
        "배송시메모": delivery_memo,
        "주문시요청사항": order_request,
        "관리자메모": admin_memo,
        "탕전주문내역서_원문텍스트": body_text,
        "탕전주문내역서URL": decoction_url,
    }


# ---------- 페이지 목록 수집 ----------
def collect_detail_links_on_current_page(driver, wait):
    """목록 페이지에서 주문 상세 링크 수집. row_date도 함께 추출해 시간 필터 사전 적용 가능."""
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    soup = BeautifulSoup(driver.page_source, "html.parser")
    detail_rows = []
    seen_hrefs = set()
    for row in soup.find_all("tr"):
        try:
            target_link = None
            for a in row.find_all("a", href=True):
                if "order_view.asp" in a["href"] and a.get_text(strip=True):
                    target_link = a
                    break
            if not target_link:
                continue
            href = target_link["href"]
            if not href.startswith("http"):
                href = f"{BASE_URL}/order/{href}" if not href.startswith("/") else f"{BASE_URL}{href}"
            if not href or href in seen_hrefs:
                continue
            seen_hrefs.add(href)
            link_text = clean_text(target_link.get_text())
            parent_td = target_link.find_parent("td")
            cell_text_raw = clean_text(parent_td.get_text()) if parent_td else link_text
            list_name, list_extra = split_list_prescription_display(cell_text_raw, link_text)

            # 목록 행에서 날짜 추출 (시간 포함) — 상세 방문 전 사전 필터용
            row_date = ""
            for td in row.find_all("td"):
                td_text = clean_text(td.get_text())
                dm = re.search(
                    r"\d{4}[-/]\d{2}[-/]\d{2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?",
                    td_text
                )
                if dm:
                    row_date = dm.group(0).replace("/", "-")
                    break

            detail_rows.append({
                "list_name": list_name,
                "list_extra": list_extra,
                "href": href,
                "row_date": row_date,
                "row_text": clean_text(row.get_text()),  # 한의원명 등 전체 행 텍스트
            })
        except Exception:
            continue
    return detail_rows


# ---------- 드라이버 생성 ----------
def create_pdf_driver():
    pdf_options = Options()
    pdf_options.add_argument("--headless=new")
    pdf_options.add_argument("--disable-gpu")
    pdf_options.add_argument("--window-size=1400,2000")
    pdf_options.add_argument("--no-sandbox")
    pdf_options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=pdf_options)


def get_printer_list() -> list:
    """현재 시스템에 설치된 프린터 목록을 반환한다."""
    printers = []
    try:
        if sys.platform == "win32":
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command",
                 "Get-Printer | Select-Object -ExpandProperty Name"],
                capture_output=True, text=True, timeout=10,
                creationflags=0x08000000  # CREATE_NO_WINDOW
            )
            if result.returncode == 0 and result.stdout.strip():
                for line in result.stdout.splitlines():
                    name = line.strip()
                    if name:
                        printers.append(name)
            else:
                # PowerShell 실패 시 wmic fallback
                result2 = subprocess.run(
                    ["wmic", "printer", "get", "name"],
                    capture_output=True, text=True, timeout=5
                )
                for line in result2.stdout.splitlines():
                    name = line.strip()
                    if name and name.lower() != "name":
                        printers.append(name)
        else:  # macOS / Linux
            result = subprocess.run(
                ["lpstat", "-p"],
                capture_output=True, text=True, timeout=5
            )
            for line in result.stdout.splitlines():
                # "printer HP_LaserJet is idle."  →  HP_LaserJet
                m = re.match(r"^printer\s+(\S+)", line)
                if m:
                    printers.append(m.group(1))
    except Exception as e:
        print(f"프린터 목록 조회 실패: {e}")
    return printers


def create_print_driver(printer_name: str = ""):
    import json as _json
    print_options = Options()
    print_options.add_argument("--kiosk-printing")
    if printer_name:
        prefs = {
            "printing.default_destination_selection_rules": _json.dumps({
                "kind": "local",
                "namePattern": printer_name,
            }),
            "printing.print_preview_sticky_settings.appState": _json.dumps({
                "recentDestinations": [
                    {"id": printer_name, "origin": "local", "account": ""}
                ],
                "selectedDestinationId": printer_name,
                "version": 2,
            }),
        }
        print_options.add_experimental_option("prefs", prefs)
    print_options.add_argument("--start-maximized")
    print_options.add_argument("--disable-gpu")
    print_options.add_argument("--no-sandbox")
    print_options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=print_options)


def create_download_driver(download_dir: str):
    download_options = Options()
    download_options.add_argument("--headless=new")
    download_options.add_argument("--disable-gpu")
    download_options.add_argument("--window-size=1400,2000")
    download_options.add_argument("--no-sandbox")
    download_options.add_argument("--disable-dev-shm-usage")
    prefs = {
        "download.default_directory": str(Path(download_dir).resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    download_options.add_experimental_option("prefs", prefs)
    drv = webdriver.Chrome(options=download_options)
    drv.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": str(Path(download_dir).resolve())}
    )
    return drv


def save_page_as_pdf(driver, url: str, save_path: str, hide_bottom_controls: bool = False):
    driver.get(url)
    time.sleep(1.5)
    if hide_bottom_controls:
        driver.execute_script("""
            const tables = Array.from(document.querySelectorAll("table"));
            tables.forEach(tbl => {
                const html = tbl.innerHTML || "";
                if (
                    html.includes("order_detail_view_excel.asp") &&
                    html.includes("printArea()") &&
                    html.includes("order_detail_view_03.asp")
                ) {
                    tbl.style.display = "none";
                }
            });
        """)
        time.sleep(0.2)
    driver.execute_script("""
        document.body.style.margin = "0";
        document.body.style.paddingTop = "20mm";
        document.body.style.paddingBottom = "8mm";
        const tables = Array.from(document.querySelectorAll("table"));
        let target = null;
        let maxArea = 0;
        tables.forEach(tbl => {
            const rect = tbl.getBoundingClientRect();
            const area = rect.width * rect.height;
            if (area > maxArea) {
                maxArea = area;
                target = tbl;
            }
        });
        if (target) {
            target.style.transform = "";
            target.style.width = "100%";
            target.style.marginLeft = "auto";
            target.style.marginRight = "auto";
        }
    """)
    time.sleep(0.5)
    pdf_data = driver.execute_cdp_cmd(
        "Page.printToPDF",
        {
            "printBackground": True,
            "paperWidth": 8.27,
            "paperHeight": 11.69,
            "marginTop": 0.4,
            "marginBottom": 0.4,
            "marginLeft": 0.4,
            "marginRight": 0.4,
            "scale": 1.0,
            "preferCSSPageSize": False,
        },
    )
    ensure_parent_dir(save_path)
    with open(save_path, "wb") as f:
        f.write(base64.b64decode(pdf_data["data"]))


def download_attachment_by_click(download_driver, page_url: str, link_text: str, final_path: str, base_download_dir: str):
    os.makedirs(os.path.dirname(final_path), exist_ok=True)
    before_files = set(os.listdir(base_download_dir))
    download_driver.get(page_url)
    time.sleep(1.2)
    anchors = download_driver.find_elements(By.TAG_NAME, "a")
    clicked = False
    for a in anchors:
        try:
            text = clean_text(a.text)
            href = clean_text(a.get_attribute("href") or "")
            if text == clean_text(link_text) and "download.asp?file=" in href:
                download_driver.execute_script("arguments[0].click();", a)
                clicked = True
                break
        except Exception:
            continue
    if not clicked:
        raise Exception(f"첨부파일 링크 클릭 실패: {link_text}")
    start = time.time()
    downloaded_name = ""
    while time.time() - start < 60:
        current_files = set(os.listdir(base_download_dir))
        new_files = list(current_files - before_files)
        completed = [f for f in new_files if not f.endswith(".crdownload")]
        downloading = any(f.endswith(".crdownload") for f in current_files)
        if completed and not downloading:
            completed.sort(key=lambda x: os.path.getmtime(os.path.join(base_download_dir, x)), reverse=True)
            downloaded_name = completed[0]
            break
        time.sleep(0.5)
    if not downloaded_name:
        raise Exception("다운로드된 파일을 찾지 못함")
    downloaded_path = os.path.join(base_download_dir, downloaded_name)
    if os.path.exists(final_path):
        os.remove(final_path)
    os.replace(downloaded_path, final_path)


def print_page(print_driver, url: str, expected_ordercode: str = "") -> dict:
    print_driver.get(url)
    time.sleep(1.5)
    if is_login_page(print_driver):
        raise Exception("로그인 세션이 만료되어 로그인 페이지로 이동했어요.")

    # 주문코드가 페이지에 나타날 때까지 최대 5초 추가 대기 (페이지 로딩 느릴 때 대비)
    page_text = clean_text(print_driver.find_element(By.TAG_NAME, "body").text)
    contains_ordercode = False
    if expected_ordercode:
        if expected_ordercode in page_text:
            contains_ordercode = True
        else:
            for _ in range(10):
                time.sleep(0.5)
                page_text = clean_text(print_driver.find_element(By.TAG_NAME, "body").text)
                if expected_ordercode in page_text:
                    contains_ordercode = True
                    break

    page_title = clean_text(print_driver.title)
    current_url = print_driver.current_url
    print_driver.execute_script("""
        const style = document.createElement('style');
        style.innerHTML = `
            @page {
                size: A4;
                margin-top: 30mm;
                margin-bottom: 10mm;
                margin-left: 15mm;
                margin-right: 15mm;
            }
            @media print {
                html, body {
                    overflow: hidden;
                    height: 227mm;
                }
            }
        `;
        document.head.appendChild(style);
    """)
    time.sleep(0.3)
    print_driver.execute_script("window.print();")
    time.sleep(PRINT_WAIT_SEC)
    return {
        "title": page_title,
        "url": current_url,
        "contains_ordercode": contains_ordercode,
    }


# ---------- requests 재시도 래퍼 ----------
def _post_with_retry(url: str, data: dict, cookies: dict,
                     timeout: int = 15, retries: int = 3, retry_delay: float = 2.0):
    """timeout + 재시도(최대 retries회) POST 요청. 연결 타임아웃 시 재시도."""
    last_exc = None
    for attempt in range(retries):
        try:
            response = requests.post(url, data=data, cookies=cookies, timeout=timeout)
            response.raise_for_status()
            body_lower = response.text.lower()
            if 'name="id"' in body_lower and 'name="pwd"' in body_lower:
                raise Exception("로그인 세션이 만료되어 로그인 페이지로 이동했어요.")
            return response
        except Exception as e:
            last_exc = e
            if attempt < retries - 1:
                time.sleep(retry_delay * (attempt + 1))  # 2초, 4초 간격으로 재시도
    raise last_exc


def fetch_order_snapshot(driver, href: str, wait_sec: float = 1.2) -> dict:
    driver.get(href)
    time.sleep(wait_sec)
    html_text = driver.page_source
    soup = BeautifulSoup(html_text, "html.parser")
    return {
        "html_text": html_text,
        "status_text": get_selected_option_text(soup, "order_ing"),
        "status_value": get_selected_option_value(soup, "order_ing"),
        "delivery_no": normalize_tracking_no(get_input_value(soup, "delivery_no")),
    }


def verify_order_status(driver, href: str, expected_value: str) -> dict:
    snapshot = fetch_order_snapshot(driver, href)
    actual_value = snapshot["status_value"]
    if actual_value != expected_value:
        expected_text = STATUS_TEXT_BY_VALUE.get(expected_value, expected_value)
        actual_text = snapshot["status_text"] or actual_value or "-"
        raise Exception(f"상태 변경 검증 실패: 기대={expected_text}, 실제={actual_text}")
    return snapshot


def verify_delivery_no(driver, href: str, expected_delivery_no: str) -> dict:
    snapshot = fetch_order_snapshot(driver, href)
    actual_delivery_no = snapshot["delivery_no"]
    expected_delivery_no = normalize_tracking_no(expected_delivery_no)
    if actual_delivery_no != expected_delivery_no:
        raise Exception(
            f"송장번호 검증 실패: 기대={expected_delivery_no or '-'}, 실제={actual_delivery_no or '-'}"
        )
    return snapshot


# ---------- 상태 전환 공통 ----------
def change_order_status(driver, html_text: str, href: str, order_ing_value: str) -> None:
    seqno = extract_seqno_from_detail_url(href)
    change_url = (
        f"{ORDER_CHANGE_BASE_URL}"
        f"?seqno={seqno}&page=&s_date=&e_date=&search=&s_string=&order_ings="
    )
    soup = BeautifulSoup(html_text, "html.parser")
    cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
    payload = {
        "ordercode": get_input_value(soup, "ordercode"),
        "handphone": get_input_value(soup, "handphone"),
        "name": "",
        "han_name": "",
        "payment": get_selected_option_value(soup, "payment"),
        "cash_bill": get_selected_option_value(soup, "cash_bill"),
        "order_ing": order_ing_value,
        "tak_sel": get_selected_option_value(soup, "tak_sel"),
        "delivery_no": get_input_value(soup, "delivery_no"),
        "realprice_memo": get_textarea_value(soup, "realprice_memo"),
    }
    _post_with_retry(change_url, data=payload, cookies=cookies)


# ---------- 송장번호 입력 ----------
def insert_delivery_no(driver, html_text: str, href: str, delivery_no: str) -> None:
    soup = BeautifulSoup(html_text, "html.parser")
    seqno = extract_seqno_from_detail_url(href)
    insert_url = (
        f"{DELIVERY_INSERT_BASE_URL}"
        f"?seqno={seqno}&page=&s_date=&e_date=&search=&s_string=&order_ings="
    )
    cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
    payload = {
        "ordercode": get_input_value(soup, "ordercode"),
        "handphone": get_input_value(soup, "handphone"),
        "name": "",
        "han_name": "",
        "payment": get_selected_option_value(soup, "payment"),
        "cash_bill": get_selected_option_value(soup, "cash_bill"),
        "order_ing": get_selected_option_value(soup, "order_ing"),
        "tak_sel": get_selected_option_value(soup, "tak_sel"),
        "delivery_no": delivery_no,
        "realprice_memo": get_textarea_value(soup, "realprice_memo"),
    }
    _post_with_retry(insert_url, data=payload, cookies=cookies)


# ---------- 라벨 인쇄용 엑셀 생성 ----------
def get_gorae_branch(text: str) -> str:
    text = str(text)
    if '관저' in text or '정재희' in text:
        return '고래한방_관저'
    elif '판암' in text or '최혁준' in text:
        return '고래한방_판암'
    elif '세종' in text or '박중현' in text:
        return '고래한방_세종'
    elif '오창' in text or '이형배' in text:
        return '고래한방_오창'
    return '고래한방_관저'  # fallback


GORAE_PANAK_CODE_MAP = {
    "위당귀수산": "WDGSS",
    "위당귀수산(평위산 합방)": "WDGSS",
    "위당귀수산(평위산합방)": "WDGSS",
    "위당귀수산 제1가감": "WDGSS1",
    "위당귀수산 제2가감": "WDGSS2",
    "당귀수산": "DGSS",
    "당귀수산 제1가감": "DGSS1",
    "당귀수산 제2가감": "DGSS2",
    "작약감초탕": "JYGC",
    "작약감초탕 제1가감": "JYGC1",
    "작약감초탕 제2가감": "JYGC2",
    "귀비온담탕": "GBOD",
    "귀비온담탕 제1가감": "GBOD1",
    "귀비온담탕 제2가감": "GBOD2",
    "감맥대조탕": "GMDJ",
    "감맥대조탕 제1가감": "GMDJ1",
    "감맥대조탕 제2가감": "GMDJ2",
    "오적산": "OJS",
    "오적산 제1가감": "OJS1",
    "오적산 제2가감": "OJS2",
}


def export_label_excel(xlsx_path: str):
    """통합 주문마스터 시트에서 라벨 인쇄용 엑셀 생성 (입원 제외)"""
    try:
        # 통합 주문마스터 시트 읽기 (탕전 컬럼 포함된 단일 시트)
        df = pd.read_excel(xlsx_path, sheet_name='주문마스터', dtype=str)
    except Exception as e:
        print(f"라벨 엑셀 생성 실패 (파일 읽기): {e}")
        return

    # 주소 매핑 (고래한방 지점 구분용 — 같은 df에서 추출)
    addr_map = {}
    if '주문코드' in df.columns:
        for _, r in df.iterrows():
            code = clean_text(str(r.get('주문코드', '') or ''))
            if code:
                parts = [
                    clean_text(str(r.get(c, '') or ''))
                    for c in ['보내는분', '보내는분_주소', '받는분', '받는분_주소']
                ]
                addr_map[code] = ' '.join(p for p in parts if p)

    # 취소 건 제외 (환불취소 상태)
    cancel_count = 0
    if '진행상태' in df.columns:
        cancel_excl = df['진행상태'].apply(lambda x: "취소" in clean_text(str(x or "")))
        cancel_count = int(cancel_excl.sum())
        df = df[~cancel_excl].copy()

    # 한의원 구분 (입원 제외 필터보다 먼저 계산 — 오창 감지에 사용)
    def classify_clinic(row):
        clinic = clean_text(str(row.get('한의원명', '') or ''))
        if '고래' in clinic:
            code = clean_text(str(row.get('주문코드', '') or ''))
            search_text = ' '.join(filter(None, [
                clinic,
                clean_text(str(row.get('회원명', '') or '')),
                clean_text(str(row.get('발송정보', '') or '')),
                clean_text(str(row.get('배송정보', '') or '')),
                addr_map.get(code, ''),
            ]))
            return get_gorae_branch(search_text)
        return clinic

    df['한의원_구분'] = df.apply(classify_clinic, axis=1)

    # 입원 제외 (조제지시사항 또는 복용첨부파일에 '입원' 포함)
    # 단, 오창점은 입원이어도 벌크 발송 대상 → 제외하지 않음
    ochang_mask = df['한의원_구분'].astype(str).str.contains('오창', na=False)

    excl = pd.Series([False] * len(df), index=df.index)
    for col in ['조제지시사항', '복용첨부파일']:
        if col in df.columns:
            excl |= df[col].str.contains('입원', na=False) & ~ochang_mask
    inpatient_count = int(excl.sum())
    df = df[~excl].copy()

    # 탕전일자: 주문날짜 컬럼을 parse_order_datetime_obj 로 파싱 후 포맷
    def to_tangjeon_date(val):
        dt = parse_order_datetime_obj(clean_text(str(val or '')))
        if dt:
            return dt.strftime('%Y.%m.%d')
        return ''
    df['탕전일자'] = df['주문날짜'].apply(to_tangjeon_date)

    df['용량'] = df['팩수'].astype(str) + '팩 / ' + df['파우치용량'].astype(str)

    # 벌크여부: 고래한방 + 한의원으로 택배 + 팩수<30 → "벌크", 나머지 → ""
    if '배송구분' in df.columns:
        df['벌크여부'] = df.apply(
            lambda row: "벌크"
            if is_bulk_delivery(
                str(row.get('배송구분', '') or ''),
                row.get('팩수', 0),
                str(row.get('한의원_구분', '') or ''),
            )
            else "",
            axis=1,
        )
    else:
        df['벌크여부'] = ''

    cols = ['한의원_구분', '환자명', '처방명', '팩수', '파우치용량', '용량', '탕전일자', '복용법', '복용첨부파일', '벌크여부', '박스포장', '파우치포장']
    label_df = df[[c for c in cols if c in df.columns]].reset_index(drop=True)

    if '팩수' in label_df.columns:
        label_df['팩수'] = pd.to_numeric(label_df['팩수'], errors='coerce')

    # ── 박스 분할: 팩수가 박스 용량 초과 시 라벨 복수 생성 ──
    import math as _math
    _split_rows = []
    for _, _row in label_df.iterrows():
        _clinic   = str(_row.get('한의원_구분', '') or '')
        _box_type = str(_row.get('박스포장', '') or '')
        _packs    = _row.get('팩수', None)
        _cap      = get_box_capacity(_clinic, _box_type)
        try:
            _packs_int = int(_packs)
        except (ValueError, TypeError):
            _packs_int = 0
        if _cap and _packs_int > _cap:
            _n = _math.ceil(_packs_int / _cap)
            for _i in range(1, _n + 1):
                _d = _row.to_dict()
                _d['박스번호'] = f'{_i}/{_n}'
                _split_rows.append(_d)
        else:
            _d = _row.to_dict()
            _d['박스번호'] = ''
            _split_rows.append(_d)
    label_df = pd.DataFrame(_split_rows).reset_index(drop=True)

    # 처방명에서 괄호 내용 분리: "오적산(소아)" → 처방명="오적산", 처방비고="소아"
    def _split_pres(v):
        s = clean_text(str(v or ''))
        # 1. 괄호 내용 분리: "오적산(소아)" → ("오적산", "소아")
        m = re.search(r'\(([^)]+)\)\s*$', s)
        if m:
            return s[:m.start()].strip(), m.group(1).strip()
        # 2. 한자 분리: 끝에 붙은 한자(漢字)를 비고로 "작약감초탕 芍藥甘草湯" → ("작약감초탕", "芍藥甘草湯")
        m2 = re.search(r'[一-鿿㐀-䶿豈-﫿]+\s*$', s)
        if m2:
            hanja = m2.group(0).strip()
            before = s[:m2.start()].strip()
            if before:  # 한자만 있는 처방명이 아닌 경우에만 분리
                return before, hanja
        return s, ''

    if '처방명' in label_df.columns:
        split_result = label_df['처방명'].apply(_split_pres)
        label_df['처방명'] = split_result.apply(lambda x: x[0])
        label_df['처방비고'] = split_result.apply(lambda x: x[1])
    else:
        label_df['처방비고'] = ''

    # 고래한방_판암 한정 영문코드 컬럼 추가 (맨 오른쪽)
    def get_panak_code(row):
        if row.get('한의원_구분') != '고래한방_판암':
            return ''
        pres = clean_text(str(row.get('처방명', '') or ''))
        code = GORAE_PANAK_CODE_MAP.get(pres, '')
        if not code:
            # _split_pres 가 처방명의 일부인 괄호까지 잘랐을 경우 재조합해서 재시도
            # 예: "위당귀수산(평위산합방)" → 처방명="위당귀수산", 처방비고="평위산합방" → "위당귀수산(평위산합방)"
            note = clean_text(str(row.get('처방비고', '') or ''))
            if note:
                code = GORAE_PANAK_CODE_MAP.get(f'{pres}({note})', '')
        return code
    label_df['라벨코드'] = label_df.apply(get_panak_code, axis=1)

    # 처방명에 대괄호 부착 ("위당귀수산" → "[위당귀수산]")
    if '처방명' in label_df.columns:
        def _wrap_pres(v):
            s = clean_text(str(v or ''))
            if not s:
                return ''
            if s.startswith('[') and s.endswith(']'):
                return s
            return f'[{s}]'
        label_df['처방명'] = label_df['처방명'].apply(_wrap_pres)

    # 열 순서 조정: 라벨코드 → 처방비고 → 벌크여부 → 박스번호 → 박스포장 → 파우치포장
    cols_order = list(label_df.columns)
    for col in ['처방비고', '벌크여부', '박스번호', '박스포장', '파우치포장']:
        if col in cols_order:
            cols_order.remove(col)
    if '라벨코드' in cols_order:
        idx = cols_order.index('라벨코드')
        for _tail in ['처방비고', '벌크여부', '박스번호', '박스포장', '파우치포장']:
            if _tail in label_df.columns:
                idx += 1
                cols_order.insert(idx, _tail)
    else:
        for _tail in ['처방비고', '벌크여부', '박스번호', '박스포장', '파우치포장']:
            if _tail in label_df.columns:
                cols_order.append(_tail)
    label_df = label_df[cols_order]

    # 타임스탬프(YYYYMMDD_HHMMSS) 추출 → "YYYYMMDD_HHMMSS_탕전 라벨 인쇄용.xlsx"
    stem = Path(xlsx_path).stem
    ts_m = re.match(r'^(\d{8}_\d{6})', stem)
    ts_prefix = ts_m.group(1) if ts_m else stem
    out_path = Path(xlsx_path).parent / f"{ts_prefix}_탕전 라벨 인쇄용.xlsx"
    label_df.to_excel(str(out_path), index=False)

    print(f"라벨 엑셀 저장: {out_path}")
    print(f"  총 {len(label_df)}건 (취소 {cancel_count}건, 입원 {inpatient_count}건 제외)")
    print("  [한의원별 건수]")
    for clinic, cnt in label_df['한의원_구분'].value_counts().items():
        print(f"    {clinic}: {cnt}건")


# ---------- 병원 대표 연락처 (CJ 업로드 양식 보내는분 자동 입력 / 폴백용) ----------
# 고래한방 지점 → branch 키 사용 (build_bulk_cj_upload_df)
# 기타 병원  → 한의원명 키워드 순서대로 매핑 (build_cj_upload_df 폴백)
GORAE_BRANCH_SENDER = {
    "관저": {"name": "고래한방병원 관저점", "phone": "042-542-1075", "address": "대전 서구 계백로 993"},
    "판암": {"name": "고래한방병원 판암점", "phone": "042-331-1005", "address": "대전 동구 옥전로 153"},
    "세종": {"name": "고래한방병원 세종점", "phone": "044-417-6637", "address": "세종특별자치시 보듬3로 158"},
    "오창": {"name": "고래한방병원 오창점", "phone": "043-714-3075", "address": "충북 청주시 청원구 오창읍 과학산업3로 216 508호"},
}

# 순서 중요: 긴/구체적 키워드를 앞에 (예: "청주필"이 "필한방"보다 먼저)
HOSPITAL_REP_PHONE: list = [
    ("청주필",  "043-715-2200"),
    ("필한방",  "042-336-1000"),
    ("약손",    "042-257-1546"),
    ("굿니스",  "042-719-7575"),
    ("개금365", "051-711-1575"),
    ("케이진",  "041-752-2224"),
    ("고래",    ""),  # 고래는 지점별로 다름 — 폴백 없음
]

def _lookup_rep_phone(hospital_name: str, sender_name: str = "") -> str:
    """한의원명/보내는분명에서 키워드 매핑으로 대표번호 반환. 없으면 ''."""
    haystack = clean_text(f"{hospital_name} {sender_name}")
    for kw, phone in HOSPITAL_REP_PHONE:
        if kw in haystack:
            return phone
    return ""

def _lookup_cj_sender_phone(row: dict) -> str:
    """CJ 보내는분전화번호에 넣을 한의원 대표번호를 찾는다."""
    member_name = clean_text(str(row.get("회원명", "") or ""))
    if member_name in GORAE_MEMBER_BRANCH_MAP:
        branch_info = GORAE_BRANCH_SENDER.get(GORAE_MEMBER_BRANCH_MAP[member_name], {})
        if branch_info.get("phone"):
            return branch_info["phone"]
    haystack = clean_text(
        " ".join(str(row.get(field, "") or "") for field in [
            "한의원명", "보내는분", "보내는분_주소", "받는분_주소", "회원명", "_hospital_folder"
        ])
    )
    for branch, info in GORAE_BRANCH_SENDER.items():
        if branch in haystack:
            return info.get("phone", "")
    return _lookup_rep_phone(
        f"{row.get('한의원명', '')} {row.get('_hospital_folder', '')}",
        row.get("보내는분", "")
    )

def _is_valid_phone(val: str) -> bool:
    """전화번호로 쓸 수 있는 값인지 (숫자 7자리 이상)"""
    digits = re.sub(r"\D", "", val)
    return len(digits) >= 7

def format_order_date_only(value: str) -> str:
    """주문날짜에서 시간 없이 YYYY-MM-DD만 반환한다."""
    dt = parse_order_datetime_obj(value)
    if dt:
        return dt.strftime("%Y-%m-%d")
    text = clean_text(str(value or ""))
    m = re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text)
    return m.group(0).replace("/", "-") if m else ""


# ---------- CJ 파일접수 양식 생성 ----------
# 실제 업로드 파일 기준 컬럼 순서
_CJ_COLUMNS = [
    "주문날짜",                  # A
    "고객주문번호",               # B
    "상호",                      # C
    "받는분성명",                 # D
    "받는분전화번호",             # E
    "받는분주소(전체, 분할)",      # F
    "품목명",                    # G
    "운임구분",                  # H
    "기본운임",                  # I
    "박스수량",                  # J
    "보내는분성명",              # K
    "보내는분전화번호",           # L
    "보내는분주소(전체, 분할)",   # M
    "배송메세지1",               # N
    "묶음여부",                  # O
]


def build_cj_upload_df(master_results: list, pdf_jobs: list) -> pd.DataFrame:
    """대한통운 파일 업로드 양식 생성 (실제 업로드 양식 기준, A-Q 17컬럼)"""
    # 주문코드 기준 중복 제거 (같은 주문이 두 번 수집된 경우 방지)
    _seen_codes: set = set()
    _deduped: list = []
    for _r in master_results:
        _code = clean_text(_r.get("주문코드", ""))
        if _code and _code in _seen_codes:
            print(f"[CJ중복제거] 주문코드 중복 제거: {_code}")
            continue
        if _code:
            _seen_codes.add(_code)
        _deduped.append(_r)
    master_results = _deduped

    dosage_file_map = {job["ordercode"]: job.get("dosage_file_name", "") for job in pdf_jobs}
    dosage_text_map = {job["ordercode"]: job.get("dosage_text_only", "") for job in pdf_jobs}

    def norm_phone(val):
        if not val:
            return ""
        return re.sub(r"[^0-9]", "", str(val)).strip()

    _FORCE_DOSAGE_KW = {"약손", "본가"}          # 배송방식 무관, 항상 복용법 첨부
    _FORCE_DOSAGE_KW_DIRECT = {"판암", "관저"}   # 직배송(환자택배)일 때만 복용법 첨부

    def is_yakson(row) -> bool:
        """복용법 강제 포함 병원 여부 판정.
        - 약손·본가: 배송 방식 무관 항상 첨부
        - 판암점: 한의원으로 택배(벌크)가 아닌 직배송일 때만 첨부"""
        delivery = clean_text(row.get("배송구분", "") or "")
        for field in ["한의원명", "보내는분", "보내는분_주소", "회원명", "_hospital_folder"]:
            val = clean_text(str(row.get(field, "") or ""))
            if any(kw in val for kw in _FORCE_DOSAGE_KW):
                return True
            if any(kw in val for kw in _FORCE_DOSAGE_KW_DIRECT):
                if delivery != "한의원으로 택배":
                    return True
        return False

    def 품명_part(ordercode, patient_name, force_dosage: bool = False) -> str:
        """환자 1명의 품명 파트. 이름 뒤에 '님' 자동 부착.
        force_dosage=True(약손한의원)이거나,
        복용첨부파일이 있거나,
        복용법 본문이 [프린트] 외 실제 내용이 있으면 _복용법 붙임."""
        has_dosage_file = bool(clean_text(dosage_file_map.get(ordercode, "")))
        dosage_text_raw = clean_text(dosage_text_map.get(ordercode, ""))
        has_dosage_text = bool(dosage_text_raw and dosage_text_raw != "[프린트]")
        name_with_honorific = f"{patient_name}님" if clean_text(patient_name) else ""
        if force_dosage or has_dosage_file or has_dosage_text:
            return f"{name_with_honorific}_복용법" if name_with_honorific else "복용법"
        return name_with_honorific

    def _dedup_receiver_name(group) -> str:
        """묶음 그룹의 받는분 성명: 모두 같은 이름이면 하나만, 다르면 ' / ' 연결"""
        names = [clean_text(r.get("받는분", "")) for r in group]
        unique = list(dict.fromkeys(names))   # 순서 보존 중복제거
        return unique[0] if len(unique) == 1 else " / ".join(unique)

    def make_row(base_row, ordercode_str: str, 품명_str: str,
                 받는분_override: str = "", is_bundle: bool = False) -> list:
        """실제 양식 기준 행 (list 반환)"""
        receiver_phone = clean_text(base_row.get("받는분_휴대폰", "") or base_row.get("받는분_전화", ""))
        original_sender_phone = clean_text(base_row.get("보내는분_전화", "") or base_row.get("보내는분_휴대폰", ""))
        sender_phone = _lookup_cj_sender_phone(base_row) or original_sender_phone
        if not _is_valid_phone(sender_phone):
            sender_phone = original_sender_phone if _is_valid_phone(original_sender_phone) else ""
        receiver_name = 받는분_override if 받는분_override else clean_text(base_row.get("받는분", ""))
        return [
            format_order_date_only(base_row.get("주문날짜", "")), # A 주문날짜
            ordercode_str,                                       # B 고객주문번호
            clean_text(base_row.get("한의원명", "")),             # C 상호
            receiver_name,                                       # D 받는분성명
            receiver_phone,                                      # E 받는분전화번호
            clean_text(base_row.get("받는분_주소", "")),          # F 받는분주소
            f"한약({품명_str})",                                   # G 품목명
            "신용",                                              # H 운임구분
            2800,                                               # I 기본운임
            1,                                                  # J 박스수량
            clean_text(base_row.get("보내는분", "")),             # K 보내는분성명
            sender_phone,                                        # L 보내는분전화번호
            clean_text(base_row.get("보내는분_주소", "")),         # M 보내는분주소
            clean_text(base_row.get("배송시메모", "")),           # N 배송메세지1
            "묶음" if is_bundle else "",                         # O 묶음여부
        ]

    # CJ 양식 제외 판정
    def should_skip_for_cj(row) -> bool:
        # 취소 건 제외 (환불취소 상태 또는 취소요청건 태그)
        if "취소" in clean_text(row.get("진행상태", "")):
            return True
        if "취소요청건" in clean_text(row.get("처방명_목록추가표시", "")):
            return True
        hospital = clean_text(row.get("한의원명", "") or row.get("보내는분", "") or "")
        delivery_type = clean_text(row.get("배송구분", ""))
        # 고래한방 + 한의원으로 택배 → 벌크 후보. 다만 예외:
        #   (1) 세종점은 자주 없어서 CJ 파일에 리스트업
        #   (2) 팩수 30 이상은 별도 박스 포장 → CJ 파일에 리스트업
        if "고래" in hospital and delivery_type == "한의원으로 택배":
            sender_addr = clean_text(row.get("보내는분_주소", "") or "")
            # 세종점은 CJ 포함
            if "세종" in sender_addr or "세종" in hospital:
                return False
            # 팩수 30 이상은 CJ 포함 (별도 박스 포장)
            pack_count_raw = row.get("팩수", "")
            try:
                if int(str(pack_count_raw).strip()) >= BULK_MAX_PACKS:
                    return False
            except (ValueError, TypeError):
                pass
            return True
        # 입원 건 제외 (오창점은 입원이어도 벌크 발송 → CJ 포함)
        if "입원" in clean_text(row.get("조제지시사항", "") or ""):
            for _f in ["한의원명", "보내는분", "보내는분_주소", "회원명", "_hospital_folder"]:
                if "오창" in clean_text(str(row.get(_f, "") or "")):
                    return False  # 오창점 입원 → 제외 안 함
            return True
        return False

    def _has_bundle_tag(row) -> bool:
        """처방명_목록추가표시에 [묶음 로 시작하는 태그가 있는지 확인 (숫자 붙은 형식도 허용: [묶음1], [묶음2] 등)"""
        return "[묶음" in clean_text(row.get("처방명_목록추가표시", ""))

    def norm_addr(val) -> str:
        """주소 정규화: 공백 모두 제거 (비교용)"""
        return re.sub(r'\s+', '', clean_text(str(val) if val else ""))

    # -----------------------------------------------------------
    # 1단계: bundle_key_set 구성 — 주소 키만 사용 (전화키 제거)
    # 같은 주소로 가는 [묶음] 태그 주문들을 하나의 그룹으로
    # -----------------------------------------------------------
    bundle_key_set: set = set()
    print("\n[CJ묶음] 1단계: 묶음 키 스캔")
    for row in master_results:
        code = clean_text(row.get("주문코드", ""))
        extra = clean_text(row.get("처방명_목록추가표시", ""))
        배송구분 = clean_text(row.get("배송구분", ""))
        if 배송구분 == "방문수령":
            if _has_bundle_tag(row):
                print(f"  {code}: [묶음] 태그 있으나 방문수령 → CJ 제외")
            continue
        if should_skip_for_cj(row):
            if _has_bundle_tag(row):
                print(f"  {code}: [묶음] 태그 있으나 CJ 제외 대상 (취소/고래벌크/입원) → 제외")
            continue
        if not _has_bundle_tag(row):
            continue
        addr = norm_addr(row.get("받는분_주소", ""))
        if addr:
            key = f"addr:{addr}"
            bundle_key_set.add(key)
            print(f"  {code}: 태그='{extra}' → 주소키 (주소 앞20: {addr[:20]})")
        else:
            print(f"  {code}: 태그='{extra}', 주소 없음 → 묶음 미적용")

    # -----------------------------------------------------------
    # 2단계: 묶음 그룹 구성 (bundle_key → 주문 목록)
    # 주소키: [묶음] 태그 있는 주문만 포함 (오주소 오결합 방지)
    # -----------------------------------------------------------
    bundle_groups: dict = {}
    print("\n[CJ묶음] 2단계: 묶음 그룹 구성")
    for row in master_results:
        if clean_text(row.get("배송구분", "")) == "방문수령":
            continue
        if should_skip_for_cj(row):
            continue
        if not _has_bundle_tag(row):
            continue
        # 주소 키 매칭: [묶음] 태그 있는 주문만
        addr = norm_addr(row.get("받는분_주소", ""))
        addr_key = f"addr:{addr}" if addr else ""
        if addr_key and addr_key in bundle_key_set:
            bundle_groups.setdefault(addr_key, []).append(row)

    for bkey, grp in bundle_groups.items():
        codes = [clean_text(r.get("주문코드", "")) for r in grp]
        names = [clean_text(r.get("환자명", "") or "") for r in grp]
        print(f"  키={bkey[:30]}: 주문={codes}, 환자={names}")

    # -----------------------------------------------------------
    # 주소 자동묶음 사전 처리 (태그 없어도 같은 주소면 2명까지 묶음)
    # 30포 이상 주문은 제외, 이미 bundle_groups에 있는 주문도 제외
    # -----------------------------------------------------------
    # id(row) → addr_bundle_key 매핑
    addr_bundle_map: dict = {}        # id(row) → str key
    addr_bundle_groups: dict = {}     # key → list[row]

    def is_addr_bundle_eligible(row) -> bool:
        """주소 자동묶음 대상 여부"""
        if should_skip_for_cj(row):
            return False
        if clean_text(row.get("배송구분", "")) == "방문수령":
            return False
        # 이미 [묶음] 태그 기반 bundle_groups에 포함된 주문은 제외
        if _has_bundle_tag(row):
            addr = norm_addr(row.get("받는분_주소", ""))
            addr_key = f"addr:{addr}" if addr else ""
            if addr_key and addr_key in bundle_key_set:
                return False
        # 팩수 30 이상이면 단독 처리
        packs = 0
        for field in ["처방명_목록추가표시", "처방명"]:
            val = clean_text(row.get(field, "") or "")
            m = re.search(r"(\d+)\s*포", val)
            if m:
                packs = int(m.group(1))
                break
        if packs >= 30:
            return False
        addr = norm_addr(row.get("받는분_주소", ""))
        return bool(addr)

    print("\n[CJ묶음] 주소자동묶음 사전 처리")
    eligible_rows = [r for r in master_results if is_addr_bundle_eligible(r)]
    # 주소 → list[row] 수집
    _addr_collect: dict = {}
    for r in eligible_rows:
        a = norm_addr(r.get("받는분_주소", ""))
        _addr_collect.setdefault(a, []).append(r)

    # 최대 2명씩 묶기 (N=2→[2], N=3→[2,1], N=4→[2,2], N=5→[2,2,1])
    _ab_counter = 0
    for addr, grp in _addr_collect.items():
        if len(grp) < 2:
            continue
        idx = 0
        while idx < len(grp):
            chunk = grp[idx: idx + 2]
            idx += 2
            if len(chunk) < 2:
                continue  # 홀로 남은 건 자동묶음 안 함
            _ab_counter += 1
            akey = f"autoaddr:{_ab_counter}"
            for r in chunk:
                addr_bundle_map[id(r)] = akey
            addr_bundle_groups[akey] = chunk
            codes = [clean_text(r.get("주문코드", "")) for r in chunk]
            names = [clean_text(r.get("환자명", "") or "") for r in chunk]
            print(f"  [주소자동묶음] 키={akey}, 주소={addr[:20]}, 주문={codes}, 환자={names}")

    # -----------------------------------------------------------
    # 헬퍼: 주문에 해당하는 bundle key 반환
    # -----------------------------------------------------------
    def get_bundle_key(row) -> str:
        if _has_bundle_tag(row):
            addr = norm_addr(row.get("받는분_주소", ""))
            addr_key = f"addr:{addr}" if addr else ""
            if addr_key and addr_key in bundle_key_set:
                return addr_key
        return ""

    # -----------------------------------------------------------
    # 3단계: 원래 순서 유지하며 행 생성
    # -----------------------------------------------------------
    rows = []         # list of (list_data, highlight_bool)
    bundle_keys_done: set = set()
    print("\n[CJ묶음] 3단계: 행 생성")

    for row in master_results:
        if clean_text(row.get("배송구분", "")) == "방문수령":
            continue
        if should_skip_for_cj(row):
            continue

        # 주소 자동묶음 우선 확인
        akey = addr_bundle_map.get(id(row), "")
        if akey:
            if akey in bundle_keys_done:
                continue
            bundle_keys_done.add(akey)
            group = addr_bundle_groups[akey]
            ordercode_str = "/".join(clean_text(r.get("주문코드", "")) for r in group)
            # 받는분 성명: 동일 이름이면 하나만, 다르면 " / " 연결
            names_override = _dedup_receiver_name(group)
            # 품목명: 각 환자명_복용법 조합을 ", " 로 연결
            명_str = ", ".join(
                품명_part(
                    clean_text(r.get("주문코드", "")),
                    clean_text(r.get("환자명", "") or ""),
                    force_dosage=is_yakson(r),
                )
                for r in group
            )
            print(f"  [주소자동묶음] 주문코드={ordercode_str}, 환자={명_str}")
            rows.append((make_row(group[0], ordercode_str, 명_str, 받는분_override=names_override, is_bundle=True), True))
            continue

        bkey = get_bundle_key(row)
        yakson = is_yakson(row)

        if bkey:
            if bkey in bundle_keys_done:
                continue  # 이미 묶음 행 생성됨
            bundle_keys_done.add(bkey)
            group = bundle_groups.get(bkey, [row])
            # 최대 2명씩 분할 (N=3→2+1, N=4→2+2)
            for ci in range(0, len(group), 2):
                chunk = group[ci: ci + 2]
                ordercode_str = "/".join(clean_text(r.get("주문코드", "")) for r in chunk)
                # 받는분 성명: 동일 이름이면 하나만, 다르면 " / " 연결
                names_override = _dedup_receiver_name(chunk)
                명_str = ", ".join(
                    품명_part(
                        clean_text(r.get("주문코드", "")),
                        clean_text(r.get("환자명", "") or ""),
                        force_dosage=is_yakson(r),
                    )
                    for r in chunk
                )
                print(f"  [주소묶음] 주문코드={ordercode_str}, 환자={명_str}")
                rows.append((make_row(chunk[0], ordercode_str, 명_str, 받는분_override=names_override, is_bundle=True), True))
        else:
            ordercode = clean_text(row.get("주문코드", ""))
            patient_name = clean_text(row.get("환자명", "") or "")
            print(f"  [단독] {ordercode} / {patient_name}")
            rows.append((make_row(row, ordercode, 품명_part(ordercode, patient_name, force_dosage=yakson)), False))

    data = [r for r, _ in rows]
    highlights = [h for _, h in rows]
    df = pd.DataFrame(data, columns=_CJ_COLUMNS)
    df["_highlight"] = highlights
    return df


# ---------- 로그 tee 헬퍼 ----------
class _TeeWriter:
    """sys.stdout을 콘솔 + StringIO 버퍼 양쪽에 동시 기록.
    PyInstaller windowed exe에서는 sys.stdout이 None이므로 None 가드 필수."""
    def __init__(self, original, buffer):
        self._orig = original  # windowed exe에서 None일 수 있음
        self._buf  = buffer
    def write(self, data):
        if self._orig is not None:
            try:
                self._orig.write(data)
            except Exception:
                pass
        self._buf.write(data)
    def flush(self):
        if self._orig is not None:
            try:
                self._orig.flush()
            except Exception:
                pass
    def isatty(self):
        return False


# ---------- 탕전주문 조회 실행 ----------
def run_job(settings: dict, progress_callback=None):
    import io as _io
    _log_buf    = _io.StringIO()
    _orig_stdout = sys.stdout
    sys.stdout   = _TeeWriter(_orig_stdout, _log_buf)
    output_root  = ""          # finally 에서 참조하기 위해 사전 초기화

    driver = None
    pdf_driver = None
    download_driver = None
    print_driver = None

    try:
        def update_progress(percent: int, message: str):
            if progress_callback:
                progress_callback(percent, message)

        target_statuses = settings["target_statuses"]
        max_page_limit = settings["max_page_limit"]
        hospital_filter = clean_text(settings.get("hospital_filter", "") or "")
        payment_filter = clean_text(settings.get("payment_filter", "") or "")
        search_target = settings.get("search_target", "주문자명")
        search_filter = clean_text(settings.get("search_filter", "") or "")
        # 프리셋 표시명이면 키워드 리스트로 변환, 직접입력이면 단일 키워드
        hospital_keywords = HOSPITAL_SEARCH_MAP.get(hospital_filter, [hospital_filter]) if hospital_filter else []
        search_keywords = [clean_text(p) for p in re.split(r"[,，]", search_filter) if clean_text(p)]
        search_member = search_target in ("주문자명", "주문자명+환자명")
        search_patient = search_target in ("환자명(복용자)", "주문자명+환자명")

        custom_output_dir = settings.get("output_dir", "").strip()
        if custom_output_dir and os.path.isdir(custom_output_dir):
            base_dir = custom_output_dir
        elif getattr(sys, "frozen", False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))

        run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_root = os.path.join(base_dir, f"결과_{run_timestamp}")
        decoction_pdf_dir = os.path.join(output_root, "주문내역서_pdf")
        dispense_pdf_dir = os.path.join(output_root, "조제지시서_pdf")
        dosage_pdf_dir = os.path.join(output_root, "복용법_pdf")
        dosage_attach_dir = os.path.join(output_root, "복용첨부파일")

        error_log_path = os.path.join(output_root, "오류_로그.txt")
        error_logs = []
        missed_print_logs = []   # 인쇄 누락 목록 (주문코드 공란 / URL 없음 / 출력 실패)
        auto_change_count = 0
        cancel_change_count = 0
        dispensing_change_count = 0
        pending_status_changes = []

        options_main = Options()
        options_main.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=options_main)
        wait = WebDriverWait(driver, 10)

        update_progress(5, "로그인 중...")
        login_driver(driver, ADMIN_ID, ADMIN_PW)
        update_progress(10, "주문 목록 조회 준비 중...")

        filter_start_dt = parse_filter_datetime(settings["start_date"])
        filter_end_dt = parse_filter_datetime(settings["end_date"])
        if filter_start_dt:
            print(f"⏰ 시작 시각 필터 적용: {settings['start_date']} 이전 주문은 추출에서 제외됩니다.")
        if filter_end_dt:
            print(f"⏰ 종료 시각 필터 적용: {settings['end_date']} 이후 주문은 추출에서 제외됩니다.")

        master_results = []
        decoction_results = []
        pdf_jobs = []
        global_collect_index = 0
        processed_count = 0
        effective_page_limit = min(max_page_limit, MAX_PAGE_SAFETY_LIMIT)
        seen_hrefs = set()

        # 전체 상태 선택이면 서버 필터 없이 한 번만 스캔, 아니면 상태별로 나눠서 스캔
        use_all = set(target_statuses) == set(ALL_STATUSES)
        scan_targets = [("", target_statuses)] if use_all else [
            (STATUS_VALUE_MAP[s], [s]) for s in target_statuses if s in STATUS_VALUE_MAP
        ]

        for order_ings_value, statuses_in_scan in scan_targets:
            if settings.get("cancel_event") and settings["cancel_event"].is_set():
                break

            if not use_all:
                print(f"\n===== 상태 필터: {statuses_in_scan[0]} (order_ings={order_ings_value}) =====")

            for page_no in range(1, effective_page_limit + 1):
                if settings.get("cancel_event") and settings["cancel_event"].is_set():
                    update_progress(0, "취소됨")
                    break

                if settings["test_limit"] is not None and processed_count >= settings["test_limit"]:
                    break

                list_url = build_order_list_url(page_no, settings["start_date"], settings["end_date"], order_ings_value)
                update_progress(15, f"{'['+statuses_in_scan[0]+'] ' if not use_all else ''}{page_no}p 목록 조회 중... ({processed_count}건)")
                print(f"\n===== {page_no}페이지 이동: {list_url} =====")

                driver.get(list_url)
                time.sleep(2)

                detail_rows = collect_detail_links_on_current_page(driver, wait)
                print(f"{page_no}페이지 상세링크 수: {len(detail_rows)}")

                if not detail_rows:
                    break

                for idx, item in enumerate(detail_rows, start=1):
                    if settings.get("cancel_event") and settings["cancel_event"].is_set():
                        update_progress(0, "취소됨")
                        break

                    if settings["test_limit"] is not None and processed_count >= settings["test_limit"]:
                        break

                    href = item["href"]
                    if href in seen_hrefs:
                        continue
                    seen_hrefs.add(href)

                    list_name = item["list_name"]
                    list_extra = item["list_extra"]

                    # 목록 날짜로 사전 필터 — 날짜(연월일)만 비교해 명확히 범위 밖인 경우만 제외
                    # ※ 시간까지 비교하면 목록의 row_date가 주문날짜가 아닌 다른 날짜 컬럼을
                    #   잘못 잡았을 때 정상 주문이 오제외되는 버그 있음 → 날짜만 비교
                    if (filter_start_dt or filter_end_dt) and item.get("row_date"):
                        row_dt = parse_order_datetime_obj(item["row_date"])
                        if row_dt:
                            row_date_only = row_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if filter_start_dt:
                                start_date_only = filter_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                if row_date_only < start_date_only:
                                    print(f"  -> 목록 날짜 범위 전(날짜 기준), 건너뜀: {item['row_date']}")
                                    continue
                            if filter_end_dt:
                                end_date_only = filter_end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                if row_date_only > end_date_only:
                                    print(f"  -> 목록 날짜 범위 후(날짜 기준), 건너뜀: {item['row_date']}")
                                    continue

                    print(f"  [{page_no}페이지 {idx}/{len(detail_rows)}] 확인 중: {list_name} {list_extra}")
                    update_progress(40, f"데이터 수집 중... ({processed_count}건 완료 | {page_no}p {idx}/{len(detail_rows)})")

                    driver.get(href)
                    time.sleep(1.5)

                    global_collect_index += 1
                    html_text = driver.page_source
                    master_data = parse_detail_html(html_text, href, global_collect_index, page_no)
                    master_data["처방명_목록"] = list_name
                    master_data["처방명_목록추가표시"] = list_extra

                    status = master_data.get("진행상태", "")
                    prescription_name = clean_text(master_data.get("처방명", ""))

                    if status not in target_statuses:
                        print(f"    -> 제외: 진행상태 = {status}")
                        continue

                    # 취소요청건 → 환불취소 자동 전환 예약
                    # 목록 스캔 중 상태를 바꾸면 사이트 페이지네이션이 밀려 주문이 누락될 수 있어
                    # 모든 주문을 먼저 수집한 뒤 일괄 전환한다.
                    if settings.get("auto_cancel_status") and "취소요청건" in clean_text(list_extra):
                        _cancel_code = clean_text(master_data.get("주문코드", ""))
                        pending_status_changes.append({
                            "kind": "cancel",
                            "href": href,
                            "ordercode": _cancel_code,
                            "target_status": "7",
                            "label": "환불취소",
                        })
                        print(f"    -> 환불취소 전환 예약: {_cancel_code}")

                    # 취소요청건 제외 (엑셀·PDF 모두 제외)
                    if "취소요청건" in clean_text(list_extra):
                        print(f"    -> 제외: 취소요청건")
                        continue

                    if filter_start_dt or filter_end_dt:
                        order_dt = parse_order_datetime_obj(master_data.get("주문날짜", ""))
                        if order_dt:
                            if filter_start_dt and order_dt < filter_start_dt:
                                print(f"    -> 제외: 주문날짜 범위 밖 ({master_data.get('주문날짜', '')})")
                                continue
                            if filter_end_dt and order_dt > filter_end_dt:
                                print(f"    -> 제외: 주문날짜 범위 밖 ({master_data.get('주문날짜', '')})")
                                continue

                    # 한의원 필터 (부분 일치, 지점 주소까지 포함해 검사)
                    if hospital_keywords:
                        _hn = clean_text(master_data.get("한의원명", "") or "")
                        _sender = clean_text(master_data.get("보내는분", "") or "")
                        _sender_addr = clean_text(master_data.get("보내는분_주소", "") or "")
                        _receiver_addr = clean_text(master_data.get("받는분_주소", "") or "")
                        # 회원명은 master_data에서 직접 읽어야 함 (이전 루프 값 사용 방지)
                        _cur_member = clean_text(master_data.get("회원명", "") or "")
                        _member_branch = GORAE_MEMBER_BRANCH_MAP.get(_cur_member, "")
                        # 목록 행 전체 텍스트도 포함 (한의원명이 상세페이지에 없을 때 보완)
                        _row_text = clean_text(item.get("row_text", "") or "")
                        _haystack = f"{_hn} {_sender} {_sender_addr} {_receiver_addr} {_member_branch} {_row_text}"
                        if not all(kw in _haystack for kw in hospital_keywords):
                            print(f"    -> 제외: 한의원 필터 '{hospital_filter}' 미일치 ({_hn or _sender})")
                            continue

                    # 결제방법 필터
                    if payment_filter:
                        _pay = clean_text(master_data.get("결제방법", "") or "")
                        if payment_filter not in _pay:
                            print(f"    -> 제외: 결제방법 필터 '{payment_filter}' 미일치 ({_pay})")
                            continue

                    ordercode = clean_text(master_data.get("주문코드", ""))

                    decoction_url = build_decoction_url_from_detail_url(href)
                    dispense_url = build_dispense_url_from_detail_url(href)
                    latest_decoction = {}

                    if decoction_url and (not search_keywords or search_patient):
                        try:
                            driver.get(decoction_url)
                            time.sleep(1.5)
                            decoction_html = driver.page_source
                            latest_decoction = parse_decoction_html(decoction_html, master_data, decoction_url)
                        except Exception as e:
                            msg = f"탕전주문내역서 저장 실패 ({ordercode}): {e}"
                            print(f"    -> {msg}")
                            error_logs.append(msg)

                    if search_keywords:
                        _member = clean_text(master_data.get("회원명", "") or "")
                        patient_haystack = clean_text(" ".join([
                            str(master_data.get("환자명", "") or ""),
                            str(latest_decoction.get("복용자", "") or "") if latest_decoction else "",
                            str(item.get("row_text", "") or ""),
                        ]))
                        member_match = search_member and any(kw in _member for kw in search_keywords)
                        patient_match = search_patient and any(kw in patient_haystack for kw in search_keywords)
                        if not (member_match or patient_match):
                            print(
                                f"    -> 제외: {search_target} 필터 '{search_filter}' 미일치 "
                                f"(주문자={_member}, 환자={patient_haystack})"
                            )
                            continue

                    master_results.append(master_data)
                    processed_count += 1
                    print(f"    -> 주문마스터 저장: {status} / 주문코드 {ordercode}")
                    if latest_decoction:
                        decoction_results.append(latest_decoction)
                        print(f"    -> 탕전주문내역서 저장: 주문코드 {ordercode}")

                    patient_name = sanitize_filename(master_data.get("환자명", "") or "환자명없음")
                    hospital_name = clean_text(master_data.get("한의원명", "") or "한의원명없음")
                    member_name = clean_text(master_data.get("회원명", "") or "")

                    sender_name = clean_text(master_data.get("보내는분", "") or "")
                    sender_address = clean_text(master_data.get("보내는분_주소", "") or "")
                    receiver_name = clean_text(master_data.get("받는분", "") or "")
                    receiver_address = clean_text(master_data.get("받는분_주소", "") or "")

                    if hospital_name in sender_name:
                        branch_address = sender_address
                    elif hospital_name in receiver_name:
                        branch_address = receiver_address
                    else:
                        branch_address = receiver_address

                    hospital_folder_name = build_hospital_folder_name(hospital_name, branch_address, member_name)
                    master_data["_hospital_folder"] = hospital_folder_name  # 팝업 집계용

                    dispensing_note = clean_text(latest_decoction.get("조제지시사항", "")) if latest_decoction else ""
                    dosage_text_only = clean_text(latest_decoction.get("복용법_본문", "")) if latest_decoction else ""
                    dosage_file_name = clean_text(latest_decoction.get("복용첨부파일명", "")) if latest_decoction else ""
                    order_datetime_file = format_order_datetime_for_filename(master_data.get("주문날짜", ""))
                    inpatient = is_inpatient_dispense(hospital_folder_name, dispensing_note)

                    # 고래한방병원인데 dispensing_note 공란이면 입원 판정 불가 경고
                    if "고래" in hospital_name and not dispensing_note:
                        msg = f"⚠ 조제지시사항 공란 — 입원 판정 불가 (탕전페이지 취득 실패 가능성): {ordercode} {patient_name}"
                        print(f"    -> {msg}")

                    # 접수대기 → 입금대기 자동 전환 예약 (입원 건)
                    if settings.get("auto_change_status") and status == "접수대기" and "입원" in dispensing_note:
                        pending_status_changes.append({
                            "kind": "inpatient",
                            "href": href,
                            "ordercode": ordercode,
                            "target_status": "2",
                            "label": "입금대기",
                        })
                        print(f"    -> 입금대기 전환 예약: {ordercode}")

                    # 정상 주문 → 조제중 자동 전환 예약 (입원·취소 제외)
                    elif settings.get("auto_dispensing_status") and not inpatient:
                        _skip_for_dispensing = {"발송", "예약발송", "완료", "환불취소"}
                        _existing_trk = normalize_tracking_no(master_data.get("송장번호", ""))
                        if status in _skip_for_dispensing or _existing_trk:
                            print(f"    -> 조제중 전환 스킵 (상태={status}, 송장={_existing_trk or '-'}): {ordercode}")
                        else:
                            pending_status_changes.append({
                                "kind": "dispensing",
                                "href": href,
                                "ordercode": ordercode,
                                "target_status": "3",
                                "label": "조제중",
                            })
                            print(f"    -> 조제중 전환 예약: {ordercode}")

                    # ★ 팩수는 order_view.asp에 없고 탕전주문내역서에만 있음
                    # → latest_decoction에서 보완 (is_bulk_delivery·should_skip_for_cj 모두 master_data 참조)
                    if latest_decoction and not master_data.get("팩수"):
                        master_data["팩수"] = clean_text(str(latest_decoction.get("팩수", "") or ""))

                    delivery_type = clean_text(master_data.get("배송구분", ""))
                    pack_count = master_data.get("팩수", "")
                    bulk = is_bulk_delivery(delivery_type, pack_count, hospital_folder_name)
                    master_data["_bulk"] = bulk  # 벌크 내역 추출 저장용

                    pdf_jobs.append({
                        "ordercode": ordercode,
                        "patient_name": patient_name,
                        "hospital_folder_name": hospital_folder_name,
                        "decoction_url": decoction_url,
                        "dispense_url": dispense_url,
                        "dispensing_note": dispensing_note,
                        "dosage_text_only": dosage_text_only,
                        "dosage_file_name": dosage_file_name,
                        "order_datetime_file": order_datetime_file,
                        "order_datetime_raw": master_data.get("주문날짜", ""),
                        "collect_index": global_collect_index,
                        "is_inpatient": inpatient,
                        "is_bulk": bulk,
                        "delivery_type": delivery_type,
                        "status": status,
                    })

        if pending_status_changes:
            update_progress(52, f"상태 자동 전환 중... ({len(pending_status_changes)}건)")
            print(f"\n[상태 자동 전환] 목록 수집 완료 후 {len(pending_status_changes)}건 처리")
            changed_keys = set()
            for ch in pending_status_changes:
                if settings.get("cancel_event") and settings["cancel_event"].is_set():
                    break
                key = (ch["kind"], ch["ordercode"], ch["target_status"])
                if key in changed_keys:
                    continue
                changed_keys.add(key)
                try:
                    driver.get(ch["href"])
                    time.sleep(1.5)
                    change_order_status(driver, driver.page_source, ch["href"], ch["target_status"])
                    if ch["kind"] == "cancel":
                        cancel_change_count += 1
                    elif ch["kind"] == "dispensing":
                        dispensing_change_count += 1
                    else:
                        auto_change_count += 1
                    print(f"  -> {ch['label']} 전환 완료: {ch['ordercode']}")
                except Exception as e:
                    msg = f"{ch['label']} 전환 실패 ({ch['ordercode']}): {e}"
                    print(f"  -> {msg}")
                    error_logs.append(msg)

        master_df = pd.DataFrame(master_results)
        if master_df.empty and not pending_status_changes:
            update_progress(100, "조회 완료 - 조건에 맞는 주문 없음")
            print("조회 조건에 맞는 주문이 없습니다.")
            return {
                "master_results": [],
                "decoction_results": [],
                "output_root": "",
                "auto_change_count": auto_change_count,
                "cancel_change_count": cancel_change_count,
                "dispensing_change_count": dispensing_change_count,
                "no_data": True,
            }

        if not master_df.empty:
            # 취소요청건 제외 (처방명_목록추가표시 태그 기준)
            cancel_req_mask = master_df["처방명_목록추가표시"].apply(
                lambda v: "취소요청건" in clean_text(str(v) if v else "")
            )
            cancel_req_count = int(cancel_req_mask.sum())
            if cancel_req_count:
                master_df = master_df[~cancel_req_mask].reset_index(drop=True)
                print(f"  취소요청건 제외: {cancel_req_count}건")

            oldest_first = settings.get("sort_oldest_first", False)
            # datetime64[ns]로 변환 → object dtype 비교 불안정 문제 해소
            master_df["_sort_dt"] = pd.to_datetime(
                master_df["주문날짜"].apply(parse_order_datetime_obj),
                errors="coerce"
            )
            # 주문코드는 순번 배정이 아님 → tiebreaker로 사용 불가
            # 같은 분(minute) 내 동률은 수집순번으로 결정
            # 서버는 최신순 반환 → 수집순번 낮을수록 최신
            # 최신순(oldest_first=False): 수집순번 오름차순 / 오래된순: 내림차순
            sn_asc = not oldest_first
            master_df.sort_values(
                by=["_sort_dt", "수집순번"],
                ascending=[oldest_first, sn_asc],
                inplace=True, na_position="last"
            )
            master_df.drop(columns=["_sort_dt"], inplace=True)
            master_df.drop_duplicates(subset=["주문코드"], keep="first", inplace=True)
            master_df.reset_index(drop=True, inplace=True)
            master_df.insert(0, "no.", range(1, len(master_df) + 1))
            master_column_order = [
                "no.", "수집순번", "페이지", "주문코드",
                "회원명", "한의원명", "환자명", "성별", "나이",
                "대표연락처_전화", "대표연락처_휴대폰", "이메일",
                "처방명", "처방명_목록", "처방명_목록추가표시", "주문날짜", "처방번호",
                "결제방법", "입금상태", "영수증신청", "계산서/영수증상태",
                "결제금액", "약재비", "탕전비", "배송비", "총비용", "할인금액", "총결제비용",
                "진행상태", "배송구분", "택배사", "송장번호",
                "카드거래번호", "카드주문번호", "카드승인번호",
                "보내는분", "보내는분_전화", "보내는분_휴대폰", "보내는분_우편번호", "보내는분_주소",
                "받는분", "받는분_전화", "받는분_휴대폰", "받는분_우편번호", "받는분_주소",
                "배송시메모", "주문시요청사항", "관리자메모",
                "상세URL"
            ]
            master_df = master_df[[c for c in master_column_order if c in master_df.columns]]

        decoction_df = pd.DataFrame(decoction_results)
        if not decoction_df.empty:
            if "주문날짜" in decoction_df.columns:
                decoction_df["_sort_dt"] = pd.to_datetime(
                    decoction_df["주문날짜"].apply(parse_order_datetime_obj),
                    errors="coerce"
                )
                decoction_df.sort_values(
                    by=["_sort_dt", "수집순번"],
                    ascending=[oldest_first, sn_asc],
                    inplace=True, na_position="last"
                )
                decoction_df.drop(columns=["_sort_dt"], inplace=True)
            elif "수집순번" in decoction_df.columns:
                decoction_df.sort_values(by="수집순번", ascending=oldest_first, inplace=True)
            decoction_df.drop_duplicates(subset=["주문코드"], keep="first", inplace=True)
            decoction_df.reset_index(drop=True, inplace=True)
            decoction_df.insert(0, "no.", range(1, len(decoction_df) + 1))
            decoction_column_order = [
                "no.",
                "주문코드", "회원명", "환자명", "처방명", "주문날짜",
                "회원명(처방자)", "한의원명", "복용자", "성별", "나이",
                "주문번호", "주문날짜_내역서", "주문상품",
                "첩수", "팩수", "파우치용량", "탕전방식", "묶음배송",
                "배송비구분", "주수상반", "스티로폼포장", "박스포장", "파우치포장",
                "약재비", "탕전비", "주수상반비용", "증류탕전비용", "배송비", "할인금액", "총결제",
                "조제지시사항", "조제첨부파일",
                "복용법", "복용첨부파일",
                "발송정보", "배송정보", "배송시메모", "주문시요청사항", "관리자메모",
                "탕전주문내역서_원문텍스트",
                "탕전주문내역서URL"
            ]
            decoction_df = decoction_df[[c for c in decoction_column_order if c in decoction_df.columns]]

        sort_suffix = "_오래된순서" if oldest_first else ""
        excel_filename = f"{run_timestamp}_케이진 탕전주문 마스터{sort_suffix}.xlsx"
        excel_path = os.path.join(output_root, excel_filename)

        # 탕전 추가 컬럼 (주문마스터 오른쪽에 붙일 순서)
        _DECOCTION_MERGE_COLS = [
            "첩수", "팩수", "파우치용량", "탕전방식", "묶음배송",
            "박스포장", "파우치포장", "조제지시사항", "조제첨부파일", "복용법", "복용첨부파일",
            "배송비구분", "주수상반", "스티로폼포장",
        ]

        excel_master_df = master_df.copy()
        excel_decoction_df = decoction_df.copy()

        if settings.get("save_excel_exclude_inpatient"):
            inpatient_ordercodes = {
                clean_text(job.get("ordercode", "") or "")
                for job in pdf_jobs
                if job.get("is_inpatient")
            }
            inpatient_ordercodes.discard("")
            if inpatient_ordercodes:
                removed_master = 0
                removed_decoction = 0
                if not excel_master_df.empty and "주문코드" in excel_master_df.columns:
                    before = len(excel_master_df)
                    excel_master_df = excel_master_df[
                        ~excel_master_df["주문코드"].astype(str).map(clean_text).isin(inpatient_ordercodes)
                    ].reset_index(drop=True)
                    if "no." in excel_master_df.columns:
                        excel_master_df["no."] = range(1, len(excel_master_df) + 1)
                    removed_master = before - len(excel_master_df)
                if not excel_decoction_df.empty and "주문코드" in excel_decoction_df.columns:
                    before = len(excel_decoction_df)
                    excel_decoction_df = excel_decoction_df[
                        ~excel_decoction_df["주문코드"].astype(str).map(clean_text).isin(inpatient_ordercodes)
                    ].reset_index(drop=True)
                    if "no." in excel_decoction_df.columns:
                        excel_decoction_df["no."] = range(1, len(excel_decoction_df) + 1)
                    removed_decoction = before - len(excel_decoction_df)
                print(f"  입원 케이스 제외 후 엑셀 저장: 주문마스터 {removed_master}건, 탕전시트 {removed_decoction}건")

        update_progress(55, "엑셀 저장 중...")
        if settings["save_excel"]:
            ensure_dir(output_root)
            # 주문마스터 + 탕전 핵심 컬럼 통합 (주문코드 기준 left join)
            if not excel_master_df.empty:
                if not excel_decoction_df.empty:
                    join_cols = ["주문코드"] + [c for c in _DECOCTION_MERGE_COLS if c in excel_decoction_df.columns]
                    merged_df = excel_master_df.merge(
                        excel_decoction_df[join_cols], on="주문코드", how="left", suffixes=("", "_dec")
                    )
                else:
                    merged_df = excel_master_df.copy()
                # 탕전 파싱 실패 / 비어있는 컬럼은 빈 컬럼으로 보장
                for col in _DECOCTION_MERGE_COLS:
                    if col not in merged_df.columns:
                        merged_df[col] = ""
                # 컬럼 순서: 마스터 컬럼 → 탕전 추가 컬럼
                final_cols = [c for c in master_column_order if c in merged_df.columns] + \
                             [c for c in _DECOCTION_MERGE_COLS if c in merged_df.columns]
                merged_df = merged_df[final_cols]
            else:
                merged_df = pd.DataFrame(columns=["no."])

            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                merged_df.to_excel(writer, sheet_name="주문마스터", index=False)
                # 탕전주문내역서 시트는 옵션 체크 시에만 저장
                if settings.get("save_decoction_sheet") and not excel_decoction_df.empty:
                    excel_decoction_df.to_excel(writer, sheet_name="탕전주문내역서", index=False)
            print(f"\n엑셀 저장 완료: {excel_path}")
            if settings.get("save_label_excel"):
                export_label_excel(excel_path)

        if settings.get("save_cj_excel") and master_results:
            ensure_dir(output_root)
            cj_df = build_cj_upload_df(master_results, pdf_jobs)
            cj_filename = f"{run_timestamp}_대한통운_파일업로드_양식.xlsx"
            cj_path = os.path.join(output_root, cj_filename)
            highlight_flags = cj_df["_highlight"].tolist() if "_highlight" in cj_df.columns else []
            cj_df_save = cj_df.drop(columns=["_highlight"], errors="ignore")
            cj_df_save.to_excel(cj_path, index=False)
            # 노란색 배경 적용 (묶음 행)
            if any(highlight_flags):
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill
                    _yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    _wb = load_workbook(cj_path)
                    _ws = _wb.active
                    for _i, _flag in enumerate(highlight_flags):
                        if _flag:
                            _excel_row = _i + 2  # header=row1, data starts row2
                            for _col in range(1, len(cj_df_save.columns) + 1):
                                _ws.cell(row=_excel_row, column=_col).fill = _yellow
                    _wb.save(cj_path)
                except Exception as _e:
                    print(f"[경고] 노란색 하이라이트 적용 실패: {_e}")
            print(f"대한통운 파일 업로드 양식 저장: {cj_path}")

        if settings.get("save_bulk_excel") and master_results:
            ensure_dir(output_root)
            _BULK_COLS = [
                "주문코드", "회원명", "한의원명", "환자명", "처방명", "주문날짜",
                "진행상태", "배송구분", "송장번호",
                "보내는분", "보내는분_주소", "보내는분_전화", "보내는분_휴대폰",
                "받는분", "받는분_주소", "받는분_전화", "받는분_휴대폰", "배송시메모",
            ]
            bulk_판암 = [r for r in master_results
                         if r.get("_bulk") and "판암" in (r.get("_hospital_folder") or "")]
            bulk_관저 = [r for r in master_results
                         if r.get("_bulk") and "관저" in (r.get("_hospital_folder") or "")]
            bulk_세종 = [r for r in master_results
                         if r.get("_bulk") and "세종" in (r.get("_hospital_folder") or "")]
            bulk_오창 = [r for r in master_results
                         if r.get("_bulk") and "오창" in (r.get("_hospital_folder") or "")]
            if bulk_판암 or bulk_관저 or bulk_세종 or bulk_오창:
                def _bulk_to_df(rows):
                    df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")} for r in rows])
                    return df[[c for c in _BULK_COLS if c in df.columns]]
                bulk_filename = f"{run_timestamp}_고래한방_벌크내역.xlsx"
                bulk_path = os.path.join(output_root, bulk_filename)
                with pd.ExcelWriter(bulk_path, engine="openpyxl") as writer:
                    if bulk_판암:
                        _bulk_to_df(bulk_판암).to_excel(writer, sheet_name="판암", index=False)
                    if bulk_관저:
                        _bulk_to_df(bulk_관저).to_excel(writer, sheet_name="관저", index=False)
                    if bulk_세종:
                        _bulk_to_df(bulk_세종).to_excel(writer, sheet_name="세종", index=False)
                    if bulk_오창:
                        _bulk_to_df(bulk_오창).to_excel(writer, sheet_name="오창", index=False)
                print(f"벌크 내역 저장: {bulk_path} (판암 {len(bulk_판암)}건, 관저 {len(bulk_관저)}건, 세종 {len(bulk_세종)}건, 오창 {len(bulk_오창)}건)")
            else:
                print("벌크 내역 없음 (판암·관저·세종·오창)")

        if settings["save_decoction_pdf"] or settings["save_dispense_pdf"] or settings["save_dosage_text_pdf"]:
            ensure_dir(output_root)
            update_progress(65, "PDF 저장 준비 중...")
            pdf_driver = create_pdf_driver()
            login_driver(pdf_driver, ADMIN_ID, ADMIN_PW)

        if settings["save_dosage_attachment"]:
            ensure_dir(dosage_attach_dir)
            download_driver = create_download_driver(dosage_attach_dir)
            login_driver(download_driver, ADMIN_ID, ADMIN_PW)

        if settings["print_dispense"] or settings.get("print_decoction"):
            print_driver = create_print_driver(settings.get("print_printer_name", ""))
            login_driver(print_driver, ADMIN_ID, ADMIN_PW)

        seen_pdf_ordercodes = set()
        print_queue = []  # 인쇄 대기 목록 — 수집 후 정렬해서 일괄 출력
        total_pdf_jobs = len(pdf_jobs)

        for _pdf_idx, job in enumerate(pdf_jobs, 1):
            ordercode = job["ordercode"]
            patient_name = job["patient_name"]
            _pdf_pct = 65 + int((_pdf_idx / total_pdf_jobs) * 20) if total_pdf_jobs else 75
            update_progress(_pdf_pct, f"PDF 저장 중... ({_pdf_idx}/{total_pdf_jobs}) {patient_name}")
            hospital_folder_name = job["hospital_folder_name"]
            dosage_text_only = clean_text(job.get("dosage_text_only", ""))
            dosage_file_name = clean_text(job.get("dosage_file_name", ""))
            order_datetime_file = job.get("order_datetime_file", "날짜없음")
            inpatient = bool(job.get("is_inpatient", False))

            if inpatient:
                inpatient_sub = "[입원]"
                bulk_sub = ""
            else:
                inpatient_sub = ""
                bulk_sub = "[벌크]" if job.get("is_bulk") else ""

            if not ordercode:
                msg = f"⚠ 주문코드 공란 — PDF/인쇄 건너뜀: {patient_name} ({hospital_folder_name})"
                print(f"    -> {msg}")
                missed_print_logs.append(f"[주문코드 공란] {patient_name} / {hospital_folder_name}")
                continue
            if ordercode in seen_pdf_ordercodes:
                print(f"    -> 중복 ordercode 스킵: {ordercode} ({patient_name})")
                continue
            seen_pdf_ordercodes.add(ordercode)

            if settings["save_decoction_pdf"] and job["decoction_url"]:
                try:
                    decoction_pdf_name = f"{order_datetime_file}_{ordercode}_{patient_name}_주문내역서.pdf"
                    decoction_pdf_path = os.path.join(decoction_pdf_dir, hospital_folder_name, inpatient_sub, bulk_sub, decoction_pdf_name)
                    save_page_as_pdf(pdf_driver, job["decoction_url"], decoction_pdf_path, hide_bottom_controls=False)
                    print(f"주문내역서 PDF 저장: {hospital_folder_name}\\{inpatient_sub}\\{bulk_sub}\\{decoction_pdf_name}")
                except Exception as e:
                    print(f"주문내역서 PDF 저장 실패 ({ordercode}): {e}")
                    error_logs.append(f"주문내역서 PDF 저장 실패 ({ordercode}): {e}")

            if settings["save_dispense_pdf"] and job["dispense_url"]:
                try:
                    if inpatient:
                        dispense_pdf_name = f"{order_datetime_file}_{ordercode}_{patient_name}_(입원)_조제지시서.pdf"
                    else:
                        dispense_pdf_name = f"{order_datetime_file}_{ordercode}_{patient_name}_조제지시서.pdf"
                    dispense_pdf_path = os.path.join(dispense_pdf_dir, hospital_folder_name, inpatient_sub, bulk_sub, dispense_pdf_name)
                    save_page_as_pdf(pdf_driver, job["dispense_url"], dispense_pdf_path, hide_bottom_controls=True)
                    print(f"조제지시서 PDF 저장: {hospital_folder_name}\\{inpatient_sub}\\{bulk_sub}\\{dispense_pdf_name}")
                except Exception as e:
                    print(f"조제지시서 PDF 저장 실패 ({ordercode}): {e}")
                    error_logs.append(f"조제지시서 PDF 저장 실패 ({ordercode}): {e}")

            if settings["save_dosage_text_pdf"] and dosage_text_only:
                try:
                    dosage_pdf_name = f"{ordercode}_{patient_name}_복용법.pdf"
                    dosage_pdf_path = os.path.join(dosage_pdf_dir, hospital_folder_name, dosage_pdf_name)
                    save_text_as_pdf(pdf_driver, f"{patient_name} 복용법", dosage_text_only, dosage_pdf_path)
                    print(f"복용법 PDF 저장: {hospital_folder_name}\\{dosage_pdf_name}")
                except Exception as e:
                    print(f"복용법 PDF 저장 실패 ({ordercode}): {e}")
                    error_logs.append(f"복용법 PDF 저장 실패 ({ordercode}): {e}")

            if settings["save_dosage_attachment"] and dosage_file_name and job["decoction_url"]:
                update_progress(90, "첨부파일 다운로드 중...")
                try:
                    attach_ext = os.path.splitext(dosage_file_name)[1] or ".bin"
                    dosage_attach_name = f"{ordercode}_{patient_name}_복용첨부파일{attach_ext}"
                    dosage_attach_path = os.path.join(dosage_attach_dir, hospital_folder_name, dosage_attach_name)
                    download_attachment_by_click(
                        download_driver, job["decoction_url"],
                        dosage_file_name, dosage_attach_path, dosage_attach_dir
                    )
                    print(f"복용첨부파일 저장: {hospital_folder_name}\\{dosage_attach_name}")
                except Exception as e:
                    print(f"복용첨부파일 저장 실패 ({ordercode}): {e} / 파일명={dosage_file_name}")
                    error_logs.append(f"복용첨부파일 저장 실패 ({ordercode}): {e} / 파일명={dosage_file_name}")

            # 인쇄 대기열에 추가 (실제 출력은 루프 후 정렬해서 일괄 처리)
            job_status = job.get("status", "")
            if settings.get("print_decoction"):
                if job["decoction_url"]:
                    print_queue.append({
                        "kind": "decoction",
                        "url": job["decoction_url"],
                        "label": f"{order_datetime_file}_{ordercode}_{patient_name}_탕전주문내역서",
                        "hospital": hospital_folder_name,
                        "inpatient": inpatient,
                        "bulk": bool(job.get("is_bulk")),
                        "delivery_type": job.get("delivery_type", ""),
                        "ordercode": ordercode,
                        "status": job_status,
                        "order_datetime": order_datetime_file,
                        "order_datetime_raw": job.get("order_datetime_raw", ""),
                        "collect_index": job.get("collect_index", 0),
                    })
                else:
                    print(f"  -> 탕전주문내역서 URL 없음, 출력 건너뜀: {ordercode}")

            if settings["print_dispense"]:
                if job["dispense_url"] and should_print_dispense(inpatient):
                    suffix = "(입원)_조제지시서" if inpatient else "조제지시서"
                    print_queue.append({
                        "kind": "dispense",
                        "url": job["dispense_url"],
                        "label": f"{order_datetime_file}_{ordercode}_{patient_name}_{suffix}",
                        "hospital": hospital_folder_name,
                        "inpatient": inpatient,
                        "bulk": bool(job.get("is_bulk")),
                        "delivery_type": job.get("delivery_type", ""),
                        "ordercode": ordercode,
                        "status": job_status,
                        "order_datetime": order_datetime_file,
                        "order_datetime_raw": job.get("order_datetime_raw", ""),
                        "collect_index": job.get("collect_index", 0),
                    })
                elif not job["dispense_url"]:
                    print(f"  -> 조제지시서 URL 없음, 출력 건너뜀: {ordercode}")

        # ── URL 없음으로 print_queue에서 빠진 건 누락 로그 ──
        if settings.get("print_dispense"):
            for job in pdf_jobs:
                oc = job.get("ordercode", "")
                pn = job.get("patient_name", "")
                if oc and oc in seen_pdf_ordercodes and not job.get("dispense_url"):
                    missed_print_logs.append(f"[조제지시서 URL 없음] {oc} / {pn}")

        # ── 인쇄 대기열 정렬 → 일괄 출력 ──
        if print_queue and print_driver:
            ensure_dir(output_root)
            # ── 1단계: datetime 객체로 안전하게 날짜순 정렬 (한국 로케일에서 strptime %p 오작동 방지)
            # 기본: 최신순 / 체크박스(sort_oldest_first=True): 오래된순  ← 엑셀 정렬과 동일
            # 서버는 최신순 반환 → collect_index 작을수록 최신, 클수록 오래됨
            # 동일 dt 동률: 최신순이면 collect_index 오름차순(작은=최신 우선) / 오래된순이면 내림차순(큰=오래됨 우선)
            _epoch = datetime(1970, 1, 1)
            oldest_first = settings.get("sort_oldest_first", False)
            def _print_sort_key(p):
                dt = parse_order_datetime_obj(p.get("order_datetime_raw", ""))
                ts = (dt - _epoch).total_seconds() if dt else 0.0
                idx = p.get("collect_index", 0)
                if oldest_first:
                    return (ts, -idx)   # dt 오름차순, 동률은 idx 내림차순(큰=오래됨 앞)
                else:
                    return (-ts, idx)   # dt 내림차순, 동률은 idx 오름차순(작은=최신 앞)
            print_queue.sort(key=_print_sort_key)
            # ── 2단계: 인쇄 묶음 정렬
            # 고래한방은 옵션과 무관하게 직배송(환자) 묶음 → 벌크 묶음 순서로 모은다.
            # 그 외 한의원은 "한의원별로 묶어서 출력" 옵션이 켜진 경우에만 묶는다.
            def _print_group_sort_key(p):
                hospital = clean_text(p.get("hospital", ""))
                is_gorae = is_gorae_hospital_folder(hospital)
                if is_gorae:
                    delivery_type = clean_text(p.get("delivery_type", ""))
                    if p.get("bulk"):
                        delivery_group = 1       # 벌크
                    elif "한의원으로 택배" in delivery_type:
                        delivery_group = 2       # 30포 이상 등 별도 박스 한의원택배
                    else:
                        delivery_group = 0       # 환자 직배송
                    branch = extract_gorae_branch(hospital) or "기타"
                    return (0, delivery_group, branch)
                if settings.get("print_by_hospital"):
                    return (1, 0, hospital)
                return (1, 0, "")

            print_queue.sort(key=_print_group_sort_key)
            # ── 3단계: 입원(입금대기) 건 맨 마지막 (옵션) — 입원이거나 입금대기인 건 후순위
            if settings.get("print_inpatient_last"):
                normal_q = [p for p in print_queue
                            if not p["inpatient"] and p.get("status") != "입금대기"]
                last_q   = [p for p in print_queue
                            if p["inpatient"] or p.get("status") == "입금대기"]
                print_queue = normal_q + last_q

            # ── 인쇄 시작 전 요약 ──
            _inpatient_q = [p for p in print_queue if p["inpatient"]]
            _normal_q    = [p for p in print_queue if not p["inpatient"]]
            print(f"\n[인쇄 대기열] 총 {len(print_queue)}건 — 일반 {len(_normal_q)}건 / 입원 {len(_inpatient_q)}건")
            for _pj in print_queue:
                _flag = "(입원)" if _pj["inpatient"] else ""
                _gorae_branch = extract_gorae_branch(_pj.get("hospital", ""))
                _delivery_type = clean_text(_pj.get("delivery_type", ""))
                if _pj.get("bulk"):
                    _delivery_group = "벌크"
                elif "한의원으로 택배" in _delivery_type:
                    _delivery_group = "한의원택배"
                else:
                    _delivery_group = "직배송"
                _gorae_info = f"[고래/{_gorae_branch}/{_delivery_group}] " if _gorae_branch else ""
                print(f"  - {_gorae_info}{_pj['ordercode']} {_flag} {_pj.get('label','')}")

            print_request_log_path = os.path.join(output_root, f"{run_timestamp}_인쇄요청목록.txt")
            with open(print_request_log_path, "w", encoding="utf-8") as f:
                f.write(f"인쇄 요청 목록 ({len(print_queue)}건)\n")
                f.write("=" * 40 + "\n")
                for i, _pj in enumerate(print_queue, 1):
                    _flag = "입원" if _pj["inpatient"] else "일반"
                    _bulk = "벌크" if _pj.get("bulk") else "-"
                    _gorae = "고래" if "고래" in clean_text(_pj.get("hospital", "")) else "-"
                    _branch = extract_gorae_branch(_pj.get("hospital", "")) or "-"
                    _delivery_type = clean_text(_pj.get("delivery_type", ""))
                    if _pj.get("bulk"):
                        _delivery_group = "벌크"
                    elif "한의원으로 택배" in _delivery_type:
                        _delivery_group = "한의원택배"
                    else:
                        _delivery_group = "직배송"
                    _kind = "탕전주문내역서" if _pj["kind"] == "decoction" else "조제지시서"
                    f.write(
                        f"{i}. [{_kind}] [{_flag}] "
                        f"[{_bulk}] [{_gorae}] [지점={_branch}] [구분={_delivery_group}] "
                        f"{_pj['ordercode']} / 상태={_pj.get('status', '')} / "
                        f"수집순번={_pj.get('collect_index', '')} / 주문일={_pj.get('order_datetime_raw', '')}\n"
                        f"   병원={_pj.get('hospital', '')}\n"
                        f"   라벨={_pj.get('label', '')}\n"
                        f"   URL={_pj.get('url', '')}\n"
                    )
            print(f"인쇄 요청 목록 저장: {print_request_log_path}")

            kind_label = {"decoction": "탕전주문내역서", "dispense": "조제지시서"}
            _total_print = len(print_queue)
            print_result_logs = []
            for _pr_idx, pj in enumerate(print_queue, 1):
                if settings.get("cancel_event") and settings["cancel_event"].is_set():
                    break
                kl = kind_label.get(pj["kind"], pj["kind"])
                # label에서 환자명 추출 시도 (label 예: "날짜_주문코드_환자명_탕전주문내역서")
                _label_parts = str(pj.get("label", "")).split("_")
                _pname = _label_parts[2] if len(_label_parts) >= 3 else ""
                _pr_pct = 90 + int((_pr_idx / _total_print) * 8) if _total_print else 95
                update_progress(_pr_pct, f"{kl} 출력 중... ({_pr_idx}/{_total_print}) {_pname}")
                try:
                    print_info = print_page(print_driver, pj["url"], pj.get("ordercode", ""))
                    _verify = "주문코드확인" if print_info.get("contains_ordercode") else "주문코드미확인"
                    print(f"{kl} 출력 요청: {pj['hospital']}\\{pj['label']} ({_verify})")
                    if not print_info.get("contains_ordercode") and pj.get("ordercode"):
                        _flag = "(입원) " if pj["inpatient"] else ""
                        missed_print_logs.append(
                            f"[페이지 로딩 미확인] {pj['ordercode']} {_flag}/ {pj.get('label','')}"
                        )
                        print(f"    -> ⚠ 주문코드가 페이지에서 확인되지 않음 — 인쇄 내용이 다를 수 있음")
                    print_result_logs.append(
                        f"{_pr_idx}. [요청완료] [{kl}] [{_verify}] "
                        f"{pj['ordercode']} / {pj.get('hospital', '')} / {pj.get('label', '')}\n"
                        f"   page_title={print_info.get('title', '')}\n"
                        f"   current_url={print_info.get('url', '')}"
                    )
                except Exception as e:
                    _flag = "(입원) " if pj["inpatient"] else ""
                    msg = f"{kl} 출력 실패 {_flag}({pj['ordercode']}): {e}"
                    print(f"    -> {msg}")
                    error_logs.append(msg)
                    missed_print_logs.append(f"[출력 실패] {pj['ordercode']} {_flag}/ {pj.get('label','')}")
                    print_result_logs.append(
                        f"{_pr_idx}. [실패] [{kl}] "
                        f"{pj['ordercode']} / {pj.get('hospital', '')} / {pj.get('label', '')}\n"
                        f"   error={e}"
                    )

            print_result_log_path = os.path.join(output_root, f"{run_timestamp}_인쇄결과목록.txt")
            with open(print_result_log_path, "w", encoding="utf-8") as f:
                f.write(f"인쇄 결과 목록 ({len(print_result_logs)}건)\n")
                f.write("=" * 40 + "\n")
                f.write("\n".join(print_result_logs))
            print(f"인쇄 결과 목록 저장: {print_result_log_path}")

        is_cancelled = settings.get("cancel_event") and settings["cancel_event"].is_set()

        if error_logs:
            ensure_dir(output_root)
            with open(error_log_path, "w", encoding="utf-8") as f:
                f.write("\n".join(error_logs))
            print(f"오류 로그 저장: {error_log_path}")

        if missed_print_logs:
            ensure_dir(output_root)
            missed_log_path = os.path.join(output_root, f"{run_timestamp}_인쇄누락목록.txt")
            with open(missed_log_path, "w", encoding="utf-8") as f:
                f.write(f"인쇄 누락 목록 ({len(missed_print_logs)}건)\n")
                f.write("=" * 40 + "\n")
                f.write("\n".join(missed_print_logs))
            print(f"인쇄 누락 목록 저장: {missed_log_path} ({len(missed_print_logs)}건)")

        if is_cancelled:
            remove_empty_dir_tree(output_root)
            update_progress(0, "취소됨")
        else:
            if has_any_file(output_root):
                if sys.platform == "win32":
                    os.startfile(output_root)
                elif sys.platform == "darwin":
                    os.system(f"open '{output_root}'")
            else:
                remove_empty_dir_tree(output_root)
                output_root = ""
            update_progress(100, "작업 완료")

        print("\n완료")
        print(f"출력 폴더: {output_root}")
        print(f"주문마스터 건수: {len(master_df)}")
        print(f"탕전주문내역서 건수: {len(decoction_df)}")

        return {
            "master_results": master_results,
            "auto_change_count": auto_change_count,
            "cancel_change_count": cancel_change_count,
            "dispensing_change_count": dispensing_change_count,
            "output_root": output_root,
        }

    finally:
        sys.stdout = _orig_stdout
        # 실행 로그 파일 저장
        try:
            _log_text = _log_buf.getvalue()
            if _log_text.strip() and output_root:
                ensure_dir(output_root)
                _log_path = os.path.join(output_root, "실행로그.txt")
                with open(_log_path, "w", encoding="utf-8") as _lf:
                    _lf.write(_log_text)
                print(f"실행 로그 저장: {_log_path}")
        except Exception:
            pass

        for drv in [driver, pdf_driver, download_driver, print_driver]:
            try:
                if drv:
                    drv.quit()
            except Exception:
                pass


# ---------- 송장번호 입력 실행 ----------
DELIVERY_TARGET_STATUSES = ["접수대기", "입금대기", "조제중", "탕전중"]


def run_delivery_job(detail_excel_path: str, start_date: str = "", end_date: str = "", log_callback=None, progress_callback=None, cancel_check=None):
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    def update_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    driver = None
    try:
        # 1. 상세내역 엑셀 읽기
        update_progress(5, "파일 읽는 중...")
        detail_df = pd.read_excel(detail_excel_path, header=0, dtype=str)

        tracking_col = None
        ordercode_col = None
        for col in detail_df.columns:
            col_clean = clean_text(str(col))
            if "운송장번호" in col_clean:
                tracking_col = col
            if "고객주문번호" in col_clean:
                ordercode_col = col

        if not tracking_col:
            raise Exception(f"운송장번호 컬럼을 찾을 수 없어요.\n찾은 컬럼: {list(detail_df.columns)}")

        use_ordercode_match = ordercode_col is not None
        if use_ordercode_match:
            log("매칭 방식: 고객주문번호(주문코드)")
        else:
            log("⚠ 고객주문번호 컬럼 없음 → 받는분전화번호로 매칭")

        from collections import defaultdict

        def normalize_phone(val):
            if not val or str(val).strip() in ("nan", ""):
                return ""
            return re.sub(r"[^0-9]", "", str(val)).strip()

        # 상세내역 파싱 (주문코드 or 전화번호)
        if use_ordercode_match:
            detail_by_ordercode = {}
            for _, row in detail_df.iterrows():
                tracking_raw = str(row.get(tracking_col, "")).strip()
                ordercode_raw = clean_text(str(row.get(ordercode_col, "")))
                if tracking_raw in ("nan", "") or not ordercode_raw or ordercode_raw == "nan":
                    continue
                tracking_clean = re.sub(r"[-\s]", "", tracking_raw).strip()
                if tracking_clean:
                    # 묶음 주문코드 "CODE1/CODE2" → 각각 동일 송장번호로 매핑
                    for code in ordercode_raw.split("/"):
                        code = code.strip()
                        if code:
                            detail_by_ordercode[code] = tracking_clean
            log(f"상세내역 파일: {len(detail_by_ordercode)}건 읽음 (주문코드 기준)")
        else:
            phone_col = None
            for col in detail_df.columns:
                col_clean = clean_text(str(col))
                if "받는분전화번호" in col_clean or "받는분 전화번호" in col_clean:
                    phone_col = col
            if not phone_col:
                raise Exception(f"운송장번호 또는 받는분전화번호 컬럼을 찾을 수 없어요.\n찾은 컬럼: {list(detail_df.columns)}")
            detail_by_phone = defaultdict(list)
            for _, row in detail_df.iterrows():
                tracking_raw = str(row.get(tracking_col, "")).strip()
                phone_raw = str(row.get(phone_col, "")).strip()
                if tracking_raw in ("nan", "") or phone_raw in ("nan", ""):
                    continue
                tracking_clean = re.sub(r"[-\s]", "", tracking_raw).strip()
                phone_clean = normalize_phone(phone_raw)
                if tracking_clean and phone_clean:
                    detail_by_phone[phone_clean].append(tracking_clean)
            log(f"상세내역 파일: {sum(len(v) for v in detail_by_phone.values())}건 읽음 (전화번호 기준)")

        filter_start_dt = parse_filter_datetime(start_date)
        filter_end_dt = parse_filter_datetime(end_date)
        if filter_start_dt:
            log(f"⏰ 시작 시각 필터 적용: {start_date} 이전 주문은 추출에서 제외됩니다.")
        if filter_end_dt:
            log(f"⏰ 종료 시각 필터 적용: {end_date} 이후 주문은 추출에서 제외됩니다.")

        # 2. 사이트 로그인 및 주문 목록 수집
        update_progress(10, "로그인 중...")
        options_main = Options()
        options_main.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=options_main)
        wait = WebDriverWait(driver, 10)
        login_driver(driver, ADMIN_ID, ADMIN_PW)

        update_progress(15, "사이트 주문 목록 수집 중...")
        site_by_ordercode = {}  # ordercode → {"url", "list_extra", "ordercode", "patient_name", "html_text"}
        site_by_phone = {}      # phone → [같은 구조] (전화번호 매칭 폴백용)

        lookup_keys = set(detail_by_ordercode.keys()) if use_ordercode_match else set(detail_by_phone.keys())

        for page_no in range(1, MAX_PAGE_SAFETY_LIMIT + 1):
            list_url = build_order_list_url(page_no, start_date, end_date)
            driver.get(list_url)
            time.sleep(2)

            detail_rows = collect_detail_links_on_current_page(driver, wait)
            if not detail_rows:
                break

            for item in detail_rows:
                href = item["href"]
                list_extra = item["list_extra"]

                # 목록 날짜로 사전 필터 — 날짜(연월일)만 비교
                if (filter_start_dt or filter_end_dt) and item.get("row_date"):
                    row_dt = parse_order_datetime_obj(item["row_date"])
                    if row_dt:
                        row_date_only = row_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                        if filter_start_dt:
                            start_date_only = filter_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if row_date_only < start_date_only:
                                continue
                        if filter_end_dt:
                            end_date_only = filter_end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if row_date_only > end_date_only:
                                continue

                driver.get(href)
                time.sleep(1.5)
                html_text = driver.page_source
                master_data = parse_detail_html(html_text, href)

                status = master_data.get("진행상태", "")
                existing_tracking = normalize_tracking_no(master_data.get("송장번호", ""))
                already_done = (status == "발송" and bool(existing_tracking))
                tracking_only_pending = (status == "발송" and not existing_tracking)
                # 발송+송장번호 없음, 송장번호만 있고 아직 발송 전인 건도 재실행 대상으로 포함
                if status not in DELIVERY_TARGET_STATUSES and not already_done and not tracking_only_pending:
                    continue

                if filter_start_dt or filter_end_dt:
                    order_dt = parse_order_datetime_obj(master_data.get("주문날짜", ""))
                    if order_dt:
                        if filter_start_dt and order_dt < filter_start_dt:
                            continue
                        if filter_end_dt and order_dt > filter_end_dt:
                            continue

                ordercode = clean_text(master_data.get("주문코드", ""))
                order_info = {
                    "url": href,
                    "list_extra": list_extra,
                    "ordercode": ordercode,
                    "patient_name": clean_text(master_data.get("환자명", "")),
                    "status": status,
                    "html_text": html_text,
                    "already_done": already_done,
                    "existing_tracking": existing_tracking,
                }
                site_by_ordercode[ordercode] = order_info

                phone = normalize_phone(master_data.get("받는분_휴대폰", ""))
                if phone:
                    if phone not in site_by_phone:
                        site_by_phone[phone] = []
                    site_by_phone[phone].append(order_info)

            collected = len(site_by_ordercode) if use_ordercode_match else len(site_by_phone)
            log(f"{page_no}페이지 수집 완료 (누적 {collected}건)")

            if use_ordercode_match:
                if all(k in site_by_ordercode for k in lookup_keys):
                    break
            else:
                if all(p in site_by_phone for p in lookup_keys):
                    break

        log(f"사이트 주문 수집 완료: {len(site_by_ordercode)}건")

        # 3. 매칭 및 처리 계획
        to_process = []
        skipped = []

        if use_ordercode_match:
            for ordercode, tracking in detail_by_ordercode.items():
                order = site_by_ordercode.get(ordercode)
                if not order:
                    log(f"⚠ 매칭 실패 (사이트에 없음): {ordercode}")
                    skipped.append(ordercode)
                    continue
                if order.get("already_done"):
                    log(f"↩ 스킵 (이미 발송처리됨 / 송장: {order.get('existing_tracking', '-')}): {ordercode} {order.get('patient_name', '')}")
                    continue
                to_process.append({
                    "url": order["url"],
                    "tracking": tracking,
                    "ordercode": order["ordercode"],
                    "patient_name": order["patient_name"],
                    "status": order["status"],
                    "html_text": order["html_text"],
                    "existing_tracking": order.get("existing_tracking", ""),
                })
        else:
            for phone, trackings in detail_by_phone.items():
                orders = site_by_phone.get(phone, [])
                if not orders:
                    log(f"⚠ 매칭 실패 (사이트에 없음): {phone}")
                    skipped.append(phone)
                    continue

                is_bundle = any("[묶음" in o["list_extra"] for o in orders)

                # 이미 발송처리된 건 스킵
                done_orders = [o for o in orders if o.get("already_done")]
                pending_orders = [o for o in orders if not o.get("already_done")]
                for o in done_orders:
                    log(f"↩ 스킵 (이미 발송처리됨 / 송장: {o.get('existing_tracking', '-')}): {o['ordercode']} {o.get('patient_name', '')}")
                orders = pending_orders
                if not orders:
                    continue

                if is_bundle:
                    for order in orders:
                        to_process.append({
                            "url": order["url"],
                            "tracking": trackings[0],
                            "ordercode": order["ordercode"],
                            "patient_name": order["patient_name"],
                            "status": order["status"],
                            "html_text": order["html_text"],
                            "existing_tracking": order.get("existing_tracking", ""),
                        })
                    log(f"묶음 처리: {phone} / 송장 {trackings[0]} → {len(orders)}건")
                elif len(orders) > 1 and len(trackings) != len(orders):
                    log(f"⚠ 스킵 (주문 {len(orders)}건 / 송장 {len(trackings)}개 불일치): {phone} / {[o['patient_name'] for o in orders]}")
                    skipped.append(phone)
                    continue
                else:
                    for i, order in enumerate(orders):
                        tracking = trackings[i] if i < len(trackings) else trackings[0]
                        to_process.append({
                            "url": order["url"],
                            "tracking": tracking,
                            "ordercode": order["ordercode"],
                            "patient_name": order["patient_name"],
                            "status": order["status"],
                            "html_text": order["html_text"],
                            "existing_tracking": order.get("existing_tracking", ""),
                        })

        log(f"\n처리 대상: {len(to_process)}건 / 스킵: {len(skipped)}건")

        if not to_process:
            log("처리할 항목이 없어요.")
            update_progress(100, "완료 - 처리할 항목 없음")
            raise NoWorkFound("조회 조건에 맞는 송장 입력 대상이 없어요.")

        # 4. 송장번호 입력
        success_count = 0
        fail_count = 0

        for i, item in enumerate(to_process):
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                update_progress(0, "취소됨")
                break
            pct = 50 + int(45 * (i + 1) / len(to_process))
            update_progress(pct, f"송장번호 입력 중... ({i+1}/{len(to_process)})")

            try:
                html_text = item["html_text"]
                current_status = clean_text(item.get("status", ""))
                existing_tracking = normalize_tracking_no(item.get("existing_tracking", ""))

                if not existing_tracking:
                    insert_delivery_no(driver, html_text, item["url"], item["tracking"])
                    snapshot = verify_delivery_no(driver, item["url"], item["tracking"])
                    html_text = snapshot["html_text"]
                    current_status = snapshot["status_text"] or current_status
                else:
                    target_tracking = normalize_tracking_no(item["tracking"])
                    if existing_tracking != target_tracking:
                        raise Exception(
                            f"기존 송장번호({existing_tracking})와 업로드 송장번호({target_tracking})가 달라요."
                        )
                    snapshot = fetch_order_snapshot(driver, item["url"])
                    html_text = snapshot["html_text"]
                    current_status = snapshot["status_text"] or current_status
                    log(f"↪ 기존 송장번호 유지, 상태만 확인/전환: {item['ordercode']} {item['patient_name']} / {existing_tracking}")

                if current_status != "발송":
                    change_order_status(driver, html_text, item["url"], "5")
                    verify_order_status(driver, item["url"], "5")

                success_count += 1
                log(f"✓ 완료: {item['ordercode']} {item['patient_name']} → {item['tracking']}")
            except Exception as e:
                fail_count += 1
                log(f"✗ 실패: {item['ordercode']} {item['patient_name']} → {e}")

        update_progress(100, "완료")
        log(f"\n완료: 성공 {success_count}건 / 실패 {fail_count}건 / 스킵 {len(skipped)}건")

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


# ---------- 벌크 발송처리 ----------
BULK_BOX_SIZE = 11      # 한 박스당 주문 수
BULK_SCAN_STATUSES = ["접수대기", "조제중", "탕전중"]
GORAE_BRANCHES = ["관저", "판암", "세종"]


def extract_gorae_branch(hospital_folder_name: str) -> str:
    """고래한방병원_관저 → '관저', 매칭 없으면 '기타'"""
    name = clean_text(hospital_folder_name)
    if name.startswith("고래한방병원_"):
        branch = name[len("고래한방병원_"):]
        return branch if branch in GORAE_BRANCHES else "기타"
    return ""


def is_gorae_hospital_folder(hospital_folder_name: str) -> bool:
    return clean_text(hospital_folder_name).startswith("고래한방병원_")


def run_bulk_scan(start_date: str = "", end_date: str = "",
                  log_callback=None, progress_callback=None) -> dict:
    """
    고래한방 벌크 미발송 주문을 스캔해 지점별 11개 그룹으로 반환.
    Returns: {"관저": [[order,...], ...], "판암": [...], "세종": [...]}
    """
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    def prog(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    driver = None
    try:
        options = Options()
        options.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 10)

        driver.set_page_load_timeout(30)
        prog(5, "로그인 중...")
        login_driver(driver, ADMIN_ID, ADMIN_PW)

        filter_start_dt = parse_filter_datetime(start_date)
        filter_end_dt = parse_filter_datetime(end_date)
        if filter_start_dt:
            prog(5, f"⏰ 시작 시각 필터: {start_date} 이전 제외")
            print(f"⏰ 시작 시각 필터 적용: {start_date} 이전 주문은 추출에서 제외됩니다.")
        if filter_end_dt:
            print(f"⏰ 종료 시각 필터 적용: {end_date} 이후 주문은 추출에서 제외됩니다.")

        branch_orders = {b: [] for b in GORAE_BRANCHES}
        seen_ordercodes = set()
        session_dead = False  # 크롬 세션 끊김 플래그

        total = len(BULK_SCAN_STATUSES)
        for si, status_name in enumerate(BULK_SCAN_STATUSES):
            if session_dead:
                break
            status_value = STATUS_VALUE_MAP.get(status_name, "")
            if not status_value:
                continue
            log(f"\n[{status_name}] 스캔 중...")
            prog(int(5 + si / total * 75), f"{status_name} 스캔 중...")

            for page_no in range(1, MAX_PAGE_SAFETY_LIMIT + 1):
                if session_dead:
                    break
                try:
                    list_url = build_order_list_url(page_no, start_date, end_date, status_value)
                    driver.get(list_url)
                    time.sleep(1.5)
                    detail_rows = collect_detail_links_on_current_page(driver, wait)
                except Exception as e:
                    err = str(e)
                    if any(k in err for k in ("invalid session", "session deleted", "disconnected", "no such window")):
                        log(f"⚠ 브라우저 세션 끊김 — 지금까지 수집된 결과로 진행합니다. ({err[:80]})")
                        session_dead = True
                    else:
                        log(f"⚠ 목록 페이지 로드 실패 (스킵): {e}")
                    break

                if not detail_rows:
                    break

                page_exhausted = False
                for item in detail_rows:
                    if session_dead:
                        break

                    # 날짜 사전 필터 — 날짜(연월일)만 비교
                    if (filter_start_dt or filter_end_dt) and item.get("row_date"):
                        row_dt = parse_order_datetime_obj(item["row_date"])
                        if row_dt:
                            row_date_only = row_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if filter_start_dt:
                                start_date_only = filter_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                if row_date_only < start_date_only:
                                    continue
                            if filter_end_dt:
                                end_date_only = filter_end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                if row_date_only > end_date_only:
                                    page_exhausted = True
                                    break

                    # 고래한방 사전 필터 — 전체 행 텍스트(한의원명 열 포함)로 검사
                    row_text = item.get("row_text") or (item.get("list_name", "") + " " + item.get("list_extra", ""))
                    if "고래" not in row_text:
                        continue

                    href = item["href"]
                    try:
                        driver.get(href)
                        time.sleep(1.0)
                        master_data = parse_detail_html(driver.page_source, href)
                    except Exception as e:
                        err = str(e)
                        if any(k in err for k in ("invalid session", "session deleted", "disconnected", "no such window")):
                            log(f"⚠ 브라우저 세션 끊김 — 지금까지 수집된 결과로 진행합니다.")
                            session_dead = True
                            break
                        log(f"⚠ 상세 페이지 오류 (스킵): {href[:60]} → {err[:60]}")
                        continue

                    delivery_type = clean_text(master_data.get("배송구분", ""))
                    hospital_name = clean_text(master_data.get("한의원명", "") or "")
                    if "고래" not in hospital_name:
                        continue
                    # 배송구분이 "한의원으로 택배"가 아니면 벌크 대상 아님 → 탕전 페이지 조회 생략
                    if "한의원으로 택배" not in delivery_type:
                        continue

                    receiver_name  = clean_text(master_data.get("받는분", "") or "")
                    receiver_addr  = clean_text(master_data.get("받는분_주소", "") or "")
                    sender_addr    = clean_text(master_data.get("보내는분_주소", "") or "")
                    member_name    = clean_text(master_data.get("회원명", "") or "")
                    # 지점 감지: 받는분 이름 + 받는분 주소 + 보내는분 주소 + 한의원명 통합 검색
                    combined = f"{hospital_name} {receiver_name} {receiver_addr} {sender_addr}"

                    # ★ 핵심: 팩수는 주문 상세 페이지(order_view.asp)에 없고
                    #         탕전 페이지(order_detail_view.asp)에만 있음 → 반드시 별도 조회
                    pack_count = ""
                    dispensing_note = ""
                    decoction_url = build_decoction_url_from_detail_url(href)
                    if decoction_url:
                        try:
                            driver.get(decoction_url)
                            time.sleep(0.8)
                            decoction_data = parse_decoction_html(driver.page_source, master_data, decoction_url)
                            pack_count = clean_text(str(decoction_data.get("팩수", "")))
                            dispensing_note = clean_text(decoction_data.get("조제지시사항", "") or "")
                        except Exception as e:
                            err = str(e)
                            if any(k in err for k in ("invalid session", "session deleted", "disconnected", "no such window")):
                                log(f"⚠ 브라우저 세션 끊김 — 지금까지 수집된 결과로 진행합니다.")
                                session_dead = True
                                break
                            log(f"  ⚠ 탕전 페이지 오류 (팩수 불명, 스킵): {err[:60]}")
                            continue
                    if session_dead:
                        break

                    ordercode = clean_text(master_data.get("주문코드", "") or "")
                    hospital_folder_name = build_hospital_folder_name(hospital_name, sender_addr or receiver_addr, member_name)
                    if is_inpatient_dispense(hospital_folder_name, dispensing_note):
                        log(f"  ✗ 벌크 제외: 입원 건 ({ordercode or '-'})")
                        continue

                    if not is_bulk_delivery(delivery_type, pack_count, hospital_name, combined):
                        log(f"  ✗ 벌크 제외: 배송={delivery_type}, 팩수={pack_count}")
                        continue

                    if not ordercode or ordercode in seen_ordercodes:
                        continue
                    seen_ordercodes.add(ordercode)

                    # 지점 결정 — combined 텍스트 검색 → 없으면 회원명 fallback
                    branch = "기타"
                    for b in GORAE_BRANCHES:  # ["관저", "판암", "세종"]
                        if b in combined:
                            branch = b
                            break
                    if branch == "기타" and member_name in GORAE_MEMBER_BRANCH_MAP:
                        branch = GORAE_MEMBER_BRANCH_MAP[member_name]

                    patient_name = clean_text(master_data.get("환자명", "") or master_data.get("복용자", "") or "")
                    log(f"  ✓ 벌크: {ordercode} {patient_name} ({branch}, {pack_count}팩, {status_name})")
                    branch_orders.setdefault(branch, []).append({
                        "ordercode": ordercode,
                        "href": href,
                        "branch": branch,
                        "patient_name": patient_name,
                        "order_datetime": format_order_datetime_for_filename(master_data.get("주문날짜", "")),
                        "order_datetime_raw": master_data.get("주문날짜", ""),
                        "pack_count": pack_count,
                        "receiver_name": receiver_name,
                        "receiver_address": receiver_addr,
                        "receiver_phone": clean_text(master_data.get("받는분_휴대폰", "") or master_data.get("받는분_전화", "") or ""),
                        "status": status_name,
                    })

                if page_exhausted:
                    break

        prog(90, "결과 정리 중...")

        # 그룹핑은 발송처리/CJ양식 생성 시점에 '박스당 주문 수' 설정에 따라 수행
        # → 여기서는 지점별 flat 리스트만 반환
        branch_orders_flat = {}
        for branch in GORAE_BRANCHES + ["기타"]:
            orders = branch_orders.get(branch, [])
            if not orders:
                continue
            orders.sort(key=lambda x: x["order_datetime"])
            branch_orders_flat[branch] = orders
            log(f"[{branch}] {len(orders)}개 수집 완료")

        if not branch_orders_flat:
            prog(100, "완료 - 벌크 주문 없음")
            raise NoWorkFound("조회 조건에 맞는 벌크 주문이 없어요.")

        prog(100, "스캔 완료")
        return branch_orders_flat

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


def build_bulk_cj_upload_df(branch_groups: dict) -> pd.DataFrame:
    """벌크 전용 CJ 업로드 양식 — 그룹 1행, 고객주문번호=고래{지점}_{번호}"""
    rows = []
    for branch in GORAE_BRANCHES + ["기타"]:
        groups = branch_groups.get(branch, [])
        sender_info = GORAE_BRANCH_SENDER.get(branch, {})
        for gi, group in enumerate(groups, start=1):
            first = group[0]
            group_id = f"고래{branch}_{gi}"
            receiver_phone = first["receiver_phone"]
            if not _is_valid_phone(receiver_phone):
                receiver_phone = ""
            rows.append([
                format_order_date_only(first.get("order_datetime_raw", "") or first.get("order_datetime", "")), # A 주문날짜
                group_id,                                             # B 고객주문번호
                f"고래한방병원_{branch}",                              # C 상호
                first["receiver_name"],                               # D 받는분성명
                receiver_phone,                                       # E 받는분전화번호
                first["receiver_address"],                            # F 받는분주소
                "한약",                                               # G 품목명
                "",                                                   # H 운임구분
                "",                                                   # I 기본운임
                1,                                                    # J 박스수량
                sender_info.get("name", f"고래한방병원_{branch}"),    # K 보내는분성명
                sender_info.get("phone", ""),                         # L 보내는분전화번호
                sender_info.get("address", ""),                       # M 보내는분주소
                "",                                                   # N 배송메시지
            ])
    return pd.DataFrame(rows, columns=_CJ_COLUMNS)


def run_bulk_ship(branch_groups: dict, tracking_by_group: dict,
                  log_callback=None, progress_callback=None, cancel_check=None):
    """
    벌크 발송처리: 그룹별 운송장 번호 입력 + 발송 상태로 전환.
    tracking_by_group: {"고래관저_1": "123456789012", "고래판암_1": "...", ...}
    """
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    def prog(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    driver = None
    try:
        options = Options()
        options.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 10)

        prog(5, "로그인 중...")
        login_driver(driver, ADMIN_ID, ADMIN_PW)

        all_jobs = []
        for branch in GORAE_BRANCHES + ["기타"]:
            groups = branch_groups.get(branch, [])
            for gi, group in enumerate(groups, start=1):
                group_id = f"고래{branch}_{gi}"
                tracking = normalize_tracking_no(tracking_by_group.get(group_id, ""))
                if not tracking:
                    log(f"⚠ {group_id}: 운송장 번호 없음, 스킵")
                    continue
                for order in group:
                    all_jobs.append({**order, "tracking": tracking, "group_id": group_id})

        total = len(all_jobs)
        if total == 0:
            log("처리할 주문이 없습니다.")
            prog(100, "완료")
            raise NoWorkFound("발송 처리할 벌크 주문이 없어요.")

        success, fail = 0, 0
        for i, job in enumerate(all_jobs):
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                prog(0, "취소됨")
                break
            prog(int(10 + i / total * 85), f"처리 중 ({i+1}/{total})...")
            try:
                driver.get(job["href"])
                time.sleep(1.5)
                html_text = driver.page_source
                master_data = parse_detail_html(html_text, job["href"])
                current_status = clean_text(master_data.get("진행상태", ""))
                current_tracking = normalize_tracking_no(master_data.get("송장번호", ""))
                target_tracking = normalize_tracking_no(job["tracking"])

                if current_tracking:
                    if current_tracking != target_tracking:
                        raise Exception(
                            f"기존 송장번호({current_tracking})와 입력 송장번호({target_tracking})가 달라요."
                        )
                    snapshot = fetch_order_snapshot(driver, job["href"])
                    html_text = snapshot["html_text"]
                    current_status = snapshot["status_text"] or current_status
                    log(f"↪ 기존 송장번호 유지: {job['ordercode']} / {current_tracking}")
                else:
                    insert_delivery_no(driver, html_text, job["href"], target_tracking)
                    snapshot = verify_delivery_no(driver, job["href"], target_tracking)
                    html_text = snapshot["html_text"]
                    current_status = snapshot["status_text"] or current_status

                if current_status != "발송":
                    change_order_status(driver, html_text, job["href"], "5")
                    verify_order_status(driver, job["href"], "5")

                success += 1
                log(f"✓ {job['group_id']} | {job['ordercode']} {job['patient_name']} → {job['tracking']}")
            except Exception as e:
                fail += 1
                log(f"✗ 실패: {job['ordercode']} {job['patient_name']} → {e}")

        prog(100, "완료")
        log(f"\n완료: 성공 {success}건 / 실패 {fail}건")

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


# ---------- 배송완료 전환 실행 ----------
def run_complete_job(start_date: str = "", end_date: str = "", ignore_status: bool = False,
                     log_callback=None, progress_callback=None, cancel_check=None):
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    def update_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    filter_start_dt = parse_filter_datetime(start_date)
    filter_end_dt   = parse_filter_datetime(end_date)
    if filter_start_dt:
        log(f"⏰ 시작 시각 필터 적용: {start_date} 이전 주문은 추출에서 제외됩니다.")
    if filter_end_dt:
        log(f"⏰ 종료 시각 필터 적용: {end_date} 이후 주문은 추출에서 제외됩니다.")

    driver = None
    try:
        update_progress(5, "로그인 중...")
        options_main = Options()
        options_main.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=options_main)
        wait = WebDriverWait(driver, 10)
        login_driver(driver, ADMIN_ID, ADMIN_PW)

        # 1. 발송(또는 전체) 상태 주문 목록 수집
        scan_ings = "" if ignore_status else STATUS_VALUE_MAP["발송"]
        scan_label = "전체(송장번호 있는 주문)" if ignore_status else "발송"
        update_progress(10, f"{scan_label} 주문 목록 수집 중...")
        if ignore_status:
            log("ℹ 상태 무관 조회: 송장번호가 있는 모든 주문을 확인합니다.")
        shipped_orders = []
        seen_seqnos = set()   # 중첩 행 중복 방지

        for page_no in range(1, MAX_PAGE_SAFETY_LIMIT + 1):
            list_url = build_order_list_url(page_no, start_date, end_date, scan_ings)
            driver.get(list_url)
            time.sleep(2)

            soup = BeautifulSoup(driver.page_source, "html.parser")
            rows = soup.find_all("tr")
            found_on_page = 0

            for row in rows:
                # 처방명 링크에서 seqno 추출
                order_link = None
                for a in row.find_all("a", href=True):
                    if "order_view.asp" in a["href"] and "seqno=" in a["href"]:
                        order_link = a
                        break

                if not order_link:
                    continue

                seqno = extract_seqno_from_detail_url(urljoin(f"{BASE_URL}/order/", order_link["href"]))
                if not seqno:
                    continue

                # 중첩 <tr> 중복 제거 — 같은 seqno는 한 번만 수집
                if seqno in seen_seqnos:
                    continue
                seen_seqnos.add(seqno)

                # 송장번호 추출
                tracking = ""
                for a in row.find_all("a", href=True):
                    href_val = a.get("href", "")
                    if "delivery_view" in href_val:
                        m = re.search(r"delivery_view\('([^']+)'\)", href_val)
                        if m:
                            tracking = m.group(1).strip()
                        break

                if not tracking:
                    continue

                # 주문코드, 환자명 (cells[0]=주문코드, cells[4]=환자명)
                cells = row.find_all("td")
                ordercode_raw = clean_text(cells[0].get_text()) if cells else ""
                # 주문코드는 10~20자리 숫자 형식 — 중첩 행의 garbage 텍스트 제거
                ordercode = ordercode_raw if re.match(r"^\d{10,20}$", ordercode_raw) else seqno
                patient = clean_text(cells[4].get_text()) if len(cells) > 4 else ""
                # 환자명이 너무 길면 (중첩 행) 비움
                if len(patient) > 30:
                    patient = ""

                # 날짜 필터 — 목록 행에서 날짜 패턴 탐색 (날짜 기준으로만 비교)
                if filter_start_dt or filter_end_dt:
                    row_date_str = ""
                    for cell in cells:
                        row_date_str = extract_datetime_text(cell.get_text())
                        if row_date_str:
                            break
                    if row_date_str:
                        order_dt = parse_order_datetime_obj(row_date_str)
                        if order_dt:
                            order_date_only = order_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if filter_start_dt:
                                start_date_only = filter_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                if order_date_only < start_date_only:
                                    continue
                            if filter_end_dt:
                                end_date_only = filter_end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                                if order_date_only > end_date_only:
                                    continue

                shipped_orders.append({
                    "seqno": seqno,
                    "tracking": tracking,
                    "patient": patient,
                    "ordercode": ordercode,
                })
                found_on_page += 1

            log(f"{page_no}페이지: {scan_label} {found_on_page}건 수집")

            order_links = soup.find_all("a", href=re.compile(r"order_view\.asp"))
            if not order_links:
                break

        log(f"\n{scan_label} 주문 총 {len(shipped_orders)}건 수집 완료")

        if not shipped_orders:
            log(f"처리할 {scan_label} 주문이 없어요.")
            update_progress(100, "완료 - 처리할 발송 주문 없음")
            raise NoWorkFound("배송완료 전환 대상 주문이 없어요.")

        # 2. 배송완료 확인 및 전환
        success_count = 0
        fail_count = 0
        skip_count = 0

        for i, order in enumerate(shipped_orders):
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                update_progress(0, "취소됨")
                break
            pct = 20 + int(75 * (i + 1) / len(shipped_orders))
            update_progress(pct, f"배송현황 확인 중... ({i+1}/{len(shipped_orders)})")

            try:
                tracking_url = f"https://trace.cjlogistics.com/next/tracking.html?wblNo={order['tracking']}"
                driver.get(tracking_url)

                # statusDetail 안에 실제 내용이 로드될 때까지 대기
                try:
                    WebDriverWait(driver, 15).until(
                        lambda d: len(d.find_element(By.ID, "statusDetail").text.strip()) > 0
                    )
                except Exception:
                    time.sleep(3)

                page_text = driver.page_source
                if "배송완료" in page_text:
                    # ── 상태 변경: order_change.asp 폼 → "완료" 선택 → 제출 ──
                    change_page_url = (
                        f"{ORDER_CHANGE_BASE_URL}"
                        f"?seqno={order['seqno']}"
                        f"&page=&s_date=&e_date=&search=&s_string=&order_ings="
                    )
                    try:
                        driver.get(change_page_url)
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.NAME, "order_ing"))
                        )
                        sel_el = driver.find_element(By.NAME, "order_ing")
                        sel_obj = Select(sel_el)

                        # 현재 선택값 확인 후 "완료" 선택
                        before_val = sel_obj.first_selected_option.text.strip()
                        if before_val == "완료":
                            skip_count += 1
                            log(f"- 이미 완료: {order['ordercode']} {order['patient']} (송장: {order['tracking']})")
                            continue
                        sel_obj.select_by_visible_text("완료")
                        time.sleep(0.3)

                        # 폼 제출 — JS order_change() 함수 사용
                        driver.execute_script(
                            "if(typeof order_change==='function'){ order_change(); }"
                            "else{ document.forms[0].submit(); }"
                        )
                        time.sleep(2)

                        # 제출 후 상태 검증 — order_change.asp 재방문해서 실제 저장값 확인
                        try:
                            driver.get(change_page_url)
                            WebDriverWait(driver, 8).until(
                                EC.presence_of_element_located((By.NAME, "order_ing"))
                            )
                            after_val = Select(driver.find_element(By.NAME, "order_ing")).first_selected_option.text.strip()
                            if after_val != "완료":
                                log(f"⚠ 상태 검증 실패 ({order['ordercode']}): 제출 후 상태={after_val} (이전={before_val}) → 재시도")
                                # 재시도 1회
                                sel_obj2 = Select(driver.find_element(By.NAME, "order_ing"))
                                sel_obj2.select_by_visible_text("완료")
                                time.sleep(0.3)
                                driver.execute_script(
                                    "if(typeof order_change==='function'){ order_change(); }"
                                    "else{ document.forms[0].submit(); }"
                                )
                                time.sleep(2)
                        except Exception:
                            pass

                    except Exception as se:
                        log(f"✗ 상태 전환 실패 ({order['ordercode']}): {se}")
                        fail_count += 1
                        continue

                    success_count += 1
                    log(f"✓ 완료 전환: {order['ordercode']} {order['patient']} (송장: {order['tracking']})")
                else:
                    skip_count += 1
                    log(f"- 배송중: {order['ordercode']} {order['patient']} (송장: {order['tracking']})")

            except Exception as e:
                fail_count += 1
                log(f"✗ 실패: {order['ordercode']} {order['patient']} → {e}")

        update_progress(100, "완료")
        log(f"\n완료: 전환 {success_count}건 / 배송중 {skip_count}건 / 실패 {fail_count}건")

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


# ---------- 약속처방주문관리 ----------

def _parse_shop_order_list_page(html: str) -> list:
    """목록 페이지 HTML에서 주문 행 목록 반환"""
    from bs4 import BeautifulSoup
    import re as _re
    soup = BeautifulSoup(html, "html.parser")
    results = []
    # onmouseover 속성 있는 tr만 데이터 행
    for tr in soup.find_all("tr", onmouseover=True):
        cells = tr.find_all("td")
        if len(cells) < 14:
            continue
        def _ct(cell):
            return clean_text(cell.get_text(" ", strip=True))
        # 경로: H or C
        route_text = _ct(cells[0])
        route = "C" if "C" in route_text else "H"
        # 주문번호
        order_no = _ct(cells[1])
        # 상품명 + 상세 URL
        link = cells[2].find("a")
        product_name = clean_text(link.get_text(" ", strip=True)) if link else _ct(cells[2])
        detail_href = link["href"] if link and link.get("href") else ""
        # 주문자
        orderer_link = cells[3].find("a")
        orderer = clean_text(orderer_link.get_text(strip=True)) if orderer_link else _ct(cells[3])
        # 등급
        grade = _ct(cells[4])
        # 수취인
        receiver = _ct(cells[5])
        # 한의원명
        clinic = _ct(cells[6])
        # 금액 (숫자만)
        amount_text = _ct(cells[7]).replace(",", "").replace("원", "").strip()
        # 입금 (select selected)
        pay_sel = cells[8].find("select")
        pay_status = ""
        if pay_sel:
            sel_opt = pay_sel.find("option", selected=True)
            pay_status = clean_text(sel_opt.get_text(strip=True)) if sel_opt else ""
        # 결제방법
        pay_method = _ct(cells[9])
        # 계산서
        receipt = _ct(cells[10])
        # 주문일
        order_date = _ct(cells[11])
        # 진행상태 (select selected)
        status_sel = cells[12].find("select")
        status = ""
        if status_sel:
            sel_opt = status_sel.find("option", selected=True)
            status = clean_text(sel_opt.get_text(strip=True)) if sel_opt else ""
        # 송장번호 (delivery_view('xxx'))
        delivery_text = _ct(cells[13])
        delivery_no = ""
        dlink = cells[13].find("a")
        if dlink:
            m = _re.search(r"delivery_view\('([^']+)'\)", dlink.get("href", ""))
            if m:
                delivery_no = m.group(1)
            else:
                delivery_no = clean_text(dlink.get_text(strip=True))
        # 메세지
        message = _ct(cells[14]) if len(cells) > 14 else ""

        results.append({
            "주문번호": order_no,
            "경로": route,
            "상품명_목록": product_name,
            "detail_href": detail_href,
            "주문자": orderer,
            "등급": grade,
            "수취인": receiver,
            "한의원명": clinic,
            "금액": amount_text,
            "입금": pay_status,
            "결제방법": pay_method,
            "계산서": receipt,
            "주문일": order_date,
            "진행상태": status,
            "송장번호": delivery_no,
            "메세지": message,
        })
    return results


def _parse_payment_str(raw: str) -> dict:
    """결제금액 문자열 → 총액/처방비용/배송료/포인트 분리.

    예시 H형: "561,000 원 (처방비용 : 561,000 | 배송료: 0 | 포인트: 0)"
    예시 C형: "390000 (처방비용 : 390,000 | 배송료: 0)"
    """
    import re
    result = {"결제금액_총액": "", "처방비용": "", "배송료": "", "포인트": ""}
    if not raw:
        return result
    # 총액: 맨 앞 숫자(쉼표 포함) 추출
    m_total = re.match(r"[\s]*([\d,]+)", raw)
    if m_total:
        result["결제금액_총액"] = m_total.group(1).replace(",", "")
    # 처방비용
    m_rx = re.search(r"처방비용\s*:\s*([\d,]+)", raw)
    if m_rx:
        result["처방비용"] = m_rx.group(1).replace(",", "")
    # 배송료
    m_ship = re.search(r"배송료\s*:\s*([\d,]+)", raw)
    if m_ship:
        result["배송료"] = m_ship.group(1).replace(",", "")
    # 포인트
    m_pt = re.search(r"포인트\s*:\s*([\d,]+)", raw)
    if m_pt:
        result["포인트"] = m_pt.group(1).replace(",", "")
    return result


def _split_contact(s: str):
    """'전화 / 휴대전화' 형식 → (전화, 휴대전화) 튜플. / 없으면 휴대전화로 분류."""
    s = s.strip()
    if " / " in s:
        parts = s.split(" / ", 1)
        return parts[0].strip(), parts[1].strip()
    return "", s  # 구분자 없으면 전부 휴대전화로


def _parse_shop_order_detail_h(html: str) -> dict:
    """order_view.asp (H형/사이트주문) 상세 파싱"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    data = {}

    def _label_value_map(soup_obj):
        """bgcolor=D8E5F0 셀을 레이블, 그 다음 셀을 값으로 매핑"""
        mapping = {}
        for tr in soup_obj.find_all("tr"):
            cells = tr.find_all("td")
            i = 0
            while i < len(cells):
                cell = cells[i]
                bg = (cell.get("bgcolor") or cell.get("style") or "").upper()
                if "D8E5F0" in bg or "d8e5f0" in (cell.get("bgcolor") or "").lower():
                    label = clean_text(cell.get_text(strip=True))
                    if label and i + 1 < len(cells):
                        mapping[label] = cells[i + 1]
                    i += 2
                else:
                    i += 1
        return mapping

    lv = _label_value_map(soup)

    def _text(key):
        cell = lv.get(key)
        return clean_text(cell.get_text(" ", strip=True)) if cell else ""

    # 주문자 (회원명 행에서 [변경] 링크 제외)
    name_cell = lv.get("회원명")
    if name_cell:
        raw = clean_text(name_cell.get_text(" ", strip=True))
        # "[변경]" 이후 실제 이름
        if "[변경]" in raw:
            raw = raw.split("[변경]")[-1].strip()
        data["주문자"] = raw
    else:
        data["주문자"] = ""

    data["한의원명"]   = _text("한의원명")
    _ct_tel, _ct_mob  = _split_contact(_text("연락처"))
    data["연락처_전화"] = _ct_tel
    data["연락처_휴대"] = _ct_mob
    data["이메일"]     = _text("이메일")
    data["사업자번호"] = ""

    # 주문상품 파싱: "제품명 : 주문내용" 형식, 값 있는 항목만 추출
    prod_cell = lv.get("주문상품")
    if prod_cell:
        items = []
        for tr in prod_cell.find_all("tr"):
            cells = tr.find_all("td")
            if len(cells) >= 2:
                name = clean_text(cells[0].get_text(" ", strip=True)).rstrip(":")
                val  = clean_text(cells[1].get_text(" ", strip=True)).lstrip(":").strip()
                if val:
                    items.append(f"{name}: {val}" if name else val)
            else:
                txt = clean_text(tr.get_text(" ", strip=True))
                if ":" in txt:
                    k, _, v = txt.partition(":")
                    if v.strip():
                        items.append(f"{k.strip()}: {v.strip()}")
                elif txt:
                    items.append(txt)
        data["주문상품"] = " / ".join(items) if items else clean_text(prod_cell.get_text(" ", strip=True))
    else:
        data["주문상품"] = ""

    # 결제금액 → 총액 / 처방비용 / 배송료 / 포인트 분리
    data.update(_parse_payment_str(_text("결제금액")))

    # 결제방법
    pay_method_cell = lv.get("결제방법")
    data["결제방법"] = clean_text(pay_method_cell.get_text(strip=True)) if pay_method_cell else ""

    # 입금
    pay_ing_cell = lv.get("입금")
    if pay_ing_cell:
        sel = pay_ing_cell.find("select")
        if sel:
            opt = sel.find("option", selected=True)
            data["입금"] = clean_text(opt.get_text(strip=True)) if opt else ""
        else:
            data["입금"] = clean_text(pay_ing_cell.get_text(strip=True))
    else:
        data["입금"] = ""
    data["입금자명"] = ""

    # 계산서
    receipt_cell = lv.get("계산서/영수증")
    if receipt_cell:
        sel = receipt_cell.find("select")
        if sel:
            opt = sel.find("option", selected=True)
            data["계산서"] = clean_text(opt.get_text(strip=True)) if opt else ""
        else:
            data["계산서"] = clean_text(receipt_cell.get_text(strip=True))
    else:
        data["계산서"] = ""

    # 진행상태
    status_cell = lv.get("진행상태")
    if status_cell:
        sel = status_cell.find("select")
        if sel:
            opt = sel.find("option", selected=True)
            data["진행상태_상세"] = clean_text(opt.get_text(strip=True)) if opt else ""
        else:
            data["진행상태_상세"] = clean_text(status_cell.get_text(strip=True))
    else:
        data["진행상태_상세"] = ""

    # 송장번호
    delivery_cell = lv.get("송장번호")
    if delivery_cell:
        inp = delivery_cell.find("input", {"name": "deliveryno"})
        data["송장번호_상세"] = clean_text(inp["value"]) if inp and inp.get("value") else ""
    else:
        data["송장번호_상세"] = ""

    # 관리자메모
    memo_cell = lv.get("관리자메모")
    data["관리자메모"] = clean_text(memo_cell.get_text(" ", strip=True)) if memo_cell else ""

    # 보내는분
    data["보내는분"]      = _text("이름")   # 첫 번째 이름 셀
    data["보내는분_전화"] = ""
    data["보내는분_휴대"] = ""
    data["보내는분_주소"]  = ""
    data["받는분"]        = ""
    data["받는분_전화"]   = ""
    data["받는분_휴대"]   = ""
    data["받는분_주소"]   = ""
    data["배송시메모"]    = ""
    data["주문시요청사항"] = ""

    # 보내는분/받는분 블록 순서대로 파싱
    _in_sender = False
    _in_receiver = False
    for tr in soup.find_all("tr"):
        txt = clean_text(tr.get_text(" ", strip=True))
        if "보내시는분" in txt and tr.find("b"):
            _in_sender = True
            _in_receiver = False
            continue
        if "받으시는분" in txt and tr.find("b"):
            _in_sender = False
            _in_receiver = True
            continue
        cells = tr.find_all("td")
        if len(cells) < 2:
            continue
        for i, cell in enumerate(cells):
            bg = (cell.get("bgcolor") or "").lower()
            if "d8e5f0" in bg:
                label = clean_text(cell.get_text(strip=True))
                val_cell = cells[i + 1] if i + 1 < len(cells) else None
                val = clean_text(val_cell.get_text(" ", strip=True)) if val_cell else ""
                if _in_sender:
                    if label == "이름":       data["보내는분"] = val
                    elif label == "연락처":
                        t, m = _split_contact(val)
                        data["보내는분_전화"] = t; data["보내는분_휴대"] = m
                    elif label == "주소":     data["보내는분_주소"] = val
                elif _in_receiver:
                    if label == "이름":       data["받는분"] = val
                    elif label == "연락처":
                        t, m = _split_contact(val)
                        data["받는분_전화"] = t; data["받는분_휴대"] = m
                    elif label == "주소":     data["받는분_주소"] = val
                    elif label == "배송시메모":   data["배송시메모"] = val
                    elif label == "주문시 요청사항": data["주문시요청사항"] = val
    return data


def _parse_shop_order_detail_c(html: str) -> dict:
    """order_view22.asp (C형/카페주문) 상세 파싱"""
    from bs4 import BeautifulSoup
    import re as _re
    soup = BeautifulSoup(html, "html.parser")
    data = {}

    def _label_value_map(soup_obj):
        mapping = {}
        for tr in soup_obj.find_all("tr"):
            cells = tr.find_all("td")
            i = 0
            while i < len(cells):
                cell = cells[i]
                bg = (cell.get("bgcolor") or "").lower()
                if "d8e5f0" in bg:
                    label = clean_text(cell.get_text(strip=True))
                    if label and i + 1 < len(cells):
                        mapping[label] = cells[i + 1]
                    i += 2
                else:
                    i += 1
        return mapping

    lv = _label_value_map(soup)

    def _text(key):
        cell = lv.get(key)
        return clean_text(cell.get_text(" ", strip=True)) if cell else ""

    # 주문자 (name_change 링크 제외)
    name_cell = lv.get("주문자")
    if name_cell:
        raw = clean_text(name_cell.get_text(" ", strip=True))
        if "[변경]" in raw:
            raw = raw.split("[변경]")[-1].strip()
        data["주문자"] = raw
    else:
        data["주문자"] = ""

    data["한의원명"]   = _text("한의원명")
    _ct_tel, _ct_mob  = _split_contact(_text("주문자연락처"))
    data["연락처_전화"] = _ct_tel
    data["연락처_휴대"] = _ct_mob
    data["이메일"]     = _text("이메일")

    # 사업자번호
    biz_cell = lv.get("사업자번호")
    if biz_cell:
        raw = clean_text(biz_cell.get_text(" ", strip=True))
        # "*사업자 번호를 입력한 주문건..." 이전 부분만
        raw = raw.split("*")[0].strip()
        data["사업자번호"] = raw
    else:
        data["사업자번호"] = ""

    # 주문상품: "제품명 : 주문내용" 형식, 내용 있는 항목만 추출
    # (목록에 여러 제품이 나열되고 실제 주문한 제품에만 값이 기재됨)
    prod_cell = lv.get("주문상품")
    if prod_cell:
        items = []
        trs = prod_cell.find_all("tr")
        if trs:
            # 내부 테이블 구조: 각 TR에서 name셀 / value셀 파싱
            for tr in trs:
                cells = tr.find_all("td")
                if len(cells) >= 2:
                    # 셀이 2개 이상: 첫 번째=제품명, 두 번째=값 (":"로 시작할 수 있음)
                    name = clean_text(cells[0].get_text(" ", strip=True)).rstrip(":")
                    val  = clean_text(cells[1].get_text(" ", strip=True)).lstrip(":").strip()
                    if val:
                        items.append(f"{name}: {val}" if name else val)
                else:
                    # 셀이 1개: "제품명 : 값" 한 줄 형식
                    txt = clean_text(tr.get_text(" ", strip=True))
                    if ":" in txt:
                        k, _, v = txt.partition(":")
                        if v.strip():
                            items.append(f"{k.strip()}: {v.strip()}")
        else:
            # TR 없음 — 줄 단위 텍스트 파싱 (": " 기준으로 값 있는 것만)
            for ln in [l.strip() for l in prod_cell.get_text("\n", strip=True).split("\n") if l.strip()]:
                if ":" in ln:
                    k, _, v = ln.partition(":")
                    if v.strip():
                        items.append(f"{k.strip()}: {v.strip()}")
        data["주문상품"] = " / ".join(items) if items else ""
    else:
        data["주문상품"] = ""

    # 결제금액 → 총액 / 처방비용 / 배송료 / 포인트 분리
    price_cell = lv.get("금액확인")
    if price_cell:
        raw = clean_text(price_cell.get_text(" ", strip=True))
        raw = raw.replace("[입력]", "").strip()
    else:
        raw = ""
    data.update(_parse_payment_str(raw))

    data["결제방법"] = _text("결제방법")

    # 입금 / 입금자명
    pay_cell = lv.get("입금 / 입금자명")
    if pay_cell:
        sel = pay_cell.find("select")
        pay_status = ""
        if sel:
            opt = sel.find("option", selected=True)
            pay_status = clean_text(opt.get_text(strip=True)) if opt else ""
        raw = clean_text(pay_cell.get_text(" ", strip=True))
        # "입금 / 천심한의원" → 입금자명 추출
        depositor = ""
        if "/" in raw:
            depositor = raw.split("/", 1)[-1].strip()
        data["입금"] = pay_status
        data["입금자명"] = depositor
    else:
        data["입금"] = ""
        data["입금자명"] = ""

    # 세금계산서
    tax_cell = lv.get("세금계산서")
    if tax_cell:
        sel = tax_cell.find("select")
        if sel:
            opt = sel.find("option", selected=True)
            data["계산서"] = clean_text(opt.get_text(strip=True)) if opt else ""
        else:
            data["계산서"] = clean_text(tax_cell.get_text(strip=True))
    else:
        data["계산서"] = ""

    # 진행상태
    status_cell = lv.get("진행상태")
    if status_cell:
        sel = status_cell.find("select")
        if sel:
            opt = sel.find("option", selected=True)
            data["진행상태_상세"] = clean_text(opt.get_text(strip=True)) if opt else ""
        else:
            data["진행상태_상세"] = clean_text(status_cell.get_text(strip=True))
    else:
        data["진행상태_상세"] = ""

    # 송장번호
    delivery_cell = lv.get("송장번호")
    if delivery_cell:
        inp = delivery_cell.find("input", {"name": "deliveryno"})
        data["송장번호_상세"] = clean_text(inp["value"]) if inp and inp.get("value") else ""
    else:
        data["송장번호_상세"] = ""

    data["관리자메모"] = _text("관리자메모")

    # 보내는분/받는분
    data["보내는분"]       = ""
    data["보내는분_전화"]  = ""
    data["보내는분_휴대"]  = ""
    data["보내는분_주소"]   = ""
    data["받는분"]         = ""
    data["받는분_전화"]    = ""
    data["받는분_휴대"]    = ""
    data["받는분_주소"]    = ""
    data["배송시메모"]     = ""
    data["주문시요청사항"]  = ""

    _in_sender = False
    _in_receiver = False
    for tr in soup.find_all("tr"):
        txt = clean_text(tr.get_text(" ", strip=True))
        if "보내시는분" in txt and tr.find("b"):
            _in_sender = True
            _in_receiver = False
            continue
        if "받으시는분" in txt and tr.find("b"):
            _in_sender = False
            _in_receiver = True
            continue
        cells = tr.find_all("td")
        if len(cells) < 2:
            continue
        for i, cell in enumerate(cells):
            bg = (cell.get("bgcolor") or "").lower()
            if "d8e5f0" in bg:
                label = clean_text(cell.get_text(strip=True))
                val_cell = cells[i + 1] if i + 1 < len(cells) else None
                val = clean_text(val_cell.get_text(" ", strip=True)) if val_cell else ""
                if _in_sender:
                    if label == "이름":       data["보내는분"] = val
                    elif label == "연락처":
                        t, m = _split_contact(val)
                        data["보내는분_전화"] = t; data["보내는분_휴대"] = m
                    elif label == "주소":     data["보내는분_주소"] = val
                elif _in_receiver:
                    if label == "이름":       data["받는분"] = val
                    elif label == "연락처":
                        t, m = _split_contact(val)
                        data["받는분_전화"] = t; data["받는분_휴대"] = m
                    elif label == "주소":     data["받는분_주소"] = val
                    elif label == "배송시메모":     data["배송시메모"] = val
                    elif label == "주문시 요청사항": data["주문시요청사항"] = val
    return data


_SHOP_ORDER_EXCEL_COLS = [
    "주문번호", "경로", "주문일", "주문자", "한의원명", "연락처_전화", "연락처_휴대", "이메일",
    "사업자번호", "상품명_목록", "주문상품",
    "결제금액_총액", "처방비용", "배송료", "포인트", "결제방법",
    "입금", "입금자명", "계산서", "진행상태", "송장번호",
    "보내는분", "보내는분_전화", "보내는분_휴대", "보내는분_주소",
    "받는분", "받는분_전화", "받는분_휴대", "받는분_주소",
    "배송시메모", "주문시요청사항", "관리자메모", "주문링크",
    # 면세/과세 열은 품목 수에 따라 동적 생성: 상품1_공급가액, 상품2_공급가액, ...
]

_SHOP_CJ_COLUMNS = [
    "주문날짜", "고객주문번호", "상호", "받는분성명", "받는분전화번호",
    "받는분주소", "품목명", "운임구분", "운임", "수량",
    "보내는분", "보내는분전화", "보내는분주소", "배송시메모",
]

# 케이진바이오 본사 발송 정보 (C형/보내는분 없는 경우 폴백)
_KGINBIO_SENDER = {
    "name": "케이진바이오",
    "phone": "042-861-9759",
    "address": "충청남도 천안시 서북구 성환읍 봉강리 111-35",
}


# 약속처방 주문상품명 → CJ 파일 품목명 변환 맵 (필요 시 추가)
# 형식: "주문상품 원문": "표기할 품목명"
_SHOP_PRODUCT_NAME_MAP: dict = {
    # 예: "공진단(목향) : 1개": "공진단(목향)",
}


# 카페 상품 → 약속처방 상품 수동 매핑 (퍼지 매칭이 실패하거나 잘못 매칭될 때 사용)
# 형식: "카페 상품명에 포함된 키워드": "약속처방 상품명에 포함된 키워드"
# 카페 상품명에 키워드가 포함되면, 약속처방 캐시에서 대응 키워드를 포함한 항목을 우선 매칭
_CAFE_TO_SHOP_NAME_MAP: dict = {
    "낙산균": "프리바이오틱스 유산균",       # 카페 "낙산균 xxx" → 약속처방 "프리바이오틱스 유산균"
    "프리바이오틱스": "낙산균 프리미엄",     # 카페 "프리바이오틱스" → 약속처방 "낙산균 프리미엄-S"
    "경옥고 30환": "30환박스",               # 카페 "경옥고 30환" → 약속처방 "30환박스"
    "경옥고 벌크": "150환",                  # 카페 "경옥고 벌크 150" → 약속처방 "150환"
    "경옥고 스틱": "30포박스",               # 카페 "경옥고 스틱 30포" → 약속처방 "30포박스"
}

# 주문 상세의 상품명 → product_tax_map 검색 키워드 수동 매핑
# C형 주문에서 자동 퍼지 매칭이 실패하는 경우에 적용
_ORDER_PRODUCT_NAME_MAP: dict = {
    "유산균": "프리바이오틱스",    # 카페 주문 "프리바이오틱스: 유산균" 에서 추출된 "유산균" → 프리바이오틱스 제품
    "사향25방": "사향공진단",      # "사향25방" → 사향공진단 25방
}


def _goods_tax_cache_path() -> str:
    """goods_tax_cache.json 파일 경로 (exe 옆 또는 소스 파일 옆)"""
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "goods_tax_cache.json")


def _load_goods_tax_cache() -> tuple:
    """캐시 로드. Returns (product_dict, meta_dict). 없거나 실패 시 ({}, {})"""
    import json as _json
    path = _goods_tax_cache_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = _json.load(f)
            return data.get("products", {}), data.get("_meta", {})
        except Exception:
            pass
    return {}, {}


def _save_goods_tax_cache(product_dict: dict):
    """상품 면세/과세 맵을 JSON 캐시로 저장"""
    import json as _json
    path = _goods_tax_cache_path()
    data = {
        "_meta": {
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "count": len(product_dict),
        },
        "products": product_dict,
    }
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(data, f, ensure_ascii=False, indent=2)


def _sync_cache_from_github(log=None) -> str:
    """GitHub Releases 최신 릴리즈에서 goods_tax_cache.json을 내려받아
    로컬 캐시보다 새로우면 덮어씀. 결과 메시지 반환."""
    import json as _json
    if log is None:
        log = lambda x: None
    try:
        cache_url = (
            f"https://github.com/{GITHUB_REPO}/releases/latest/download/"
            f"{GITHUB_CACHE_FILE_NAME}"
        )
        log("GitHub 상품 캐시 확인 중...")
        resp = requests.get(cache_url, timeout=15)
        if resp.status_code == 404:
            return "GitHub에 캐시 파일 없음 (아직 배포 전)"
        resp.raise_for_status()
        remote_data = resp.json()
        remote_updated = remote_data.get("_meta", {}).get("updated_at", "")
        remote_count   = len(remote_data.get("products", {}))

        _local, _local_meta = _load_goods_tax_cache()
        local_updated = _local_meta.get("updated_at", "")

        if remote_updated > local_updated:
            with open(_goods_tax_cache_path(), "w", encoding="utf-8") as f:
                _json.dump(remote_data, f, ensure_ascii=False, indent=2)
            msg = f"GitHub 캐시 업데이트 완료: {remote_count}개 ({remote_updated})"
        else:
            msg = f"이미 최신 상태 ({local_updated or '로컬 캐시 없음'})"
        log(msg)
        return msg
    except Exception as _e:
        msg = f"GitHub 동기화 실패: {_e}"
        log(msg)
        return msg


def _upload_cache_to_github(token: str, log=None) -> str:
    """로컬 goods_tax_cache.json을 GitHub Releases 최신 릴리즈에 업로드."""
    if log is None:
        log = lambda x: None
    try:
        cache_path = _goods_tax_cache_path()
        if not os.path.exists(cache_path):
            return "로컬 캐시 파일이 없습니다. 먼저 다운로드/업데이트를 실행하세요."
        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
        }
        log("GitHub 최신 릴리즈 조회 중...")
        resp = requests.get(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            headers=headers, timeout=15,
        )
        resp.raise_for_status()
        release = resp.json()
        upload_url_base = release["upload_url"].split("{")[0]

        # 기존 동일 asset 삭제
        for asset in release.get("assets", []):
            if asset["name"] == GITHUB_CACHE_FILE_NAME:
                log(f"기존 캐시 asset 삭제 중...")
                requests.delete(
                    f"https://api.github.com/repos/{GITHUB_REPO}/releases/assets/{asset['id']}",
                    headers=headers, timeout=15,
                )

        with open(cache_path, "rb") as f:
            content = f.read()
        log("GitHub에 캐시 파일 업로드 중...")
        up_resp = requests.post(
            f"{upload_url_base}?name={GITHUB_CACHE_FILE_NAME}",
            headers={**headers, "Content-Type": "application/json"},
            data=content, timeout=30,
        )
        up_resp.raise_for_status()
        msg = f"GitHub 배포 완료 ({len(content):,} bytes)"
        log(msg)
        return msg
    except Exception as _e:
        msg = f"GitHub 배포 실패: {_e}"
        log(msg)
        return msg


def _parse_order_products(주문상품_str: str) -> list:
    """주문상품 문자열 → [(상품명, 수량), ...] 파싱.

    형식 예:
      "경옥고 30환: 1개 / 공진단- 2개"  →  [("경옥고 30환", "1개"), ("공진단", "2개")]
    구분자는 ': ' 또는 '- ' 모두 처리.
    """
    result = []
    if not 주문상품_str:
        return result
    for part in 주문상품_str.split(" / "):
        part = part.strip()
        if not part:
            continue
        # 끝부분 "- N개" 또는 ": N개" 패턴으로 이름/수량 분리
        m = re.match(r'^(.+?)\s*[-:]\s*(\d+)\s*(개)?\s*$', part)
        if m:
            name = m.group(1).strip()
            qty  = f"{m.group(2)}{m.group(3) or '개'}"
        else:
            name = part
            qty  = ""
        result.append((name, qty))
    return result


def _parse_qty_int(qty_str: str) -> int:
    """수량 문자열 ('10개', '3', '') → 정수. 못 읽으면 1."""
    if not qty_str:
        return 1
    m = re.match(r"\s*(\d+)", str(qty_str))
    return int(m.group(1)) if m else 1


def _extract_product_names_for_tax(merged: dict) -> list:
    """주문 행에서 상품명 리스트 추출 (면세/과세 캐시 매핑용).

    H형 (사이트 직접주문) 형식: "경옥고 30환: 1개 / 공진단- 2개"
      → ["경옥고 30환", "공진단"]

    C형 (카페 주문) 형식: "프리바이오틱스: 유산균 / 인삼: , 백절삼600g"
      여기서 앞 부분(콜론 앞)은 카테고리, 뒷 부분(콜론 뒤)이 실제 상품명.
      → ["유산균", "백절삼600g"]

    상품명_목록: 목록 페이지의 단일 상품명 (폴백용)
    """
    주문상품 = clean_text(merged.get("주문상품", "") or "")
    if 주문상품:
        names = []
        for raw_name, _ in _parse_order_products(주문상품):
            if not raw_name:
                continue
            # C형 카페 주문: "카테고리: 상품명텍스트" 형식 감지
            # — _parse_order_products 에서 숫자 패턴이 매칭 안 되면 콜론이 name에 그대로 남음
            # — 콜론 뒤가 비숫자 텍스트이면 뒷 부분이 실제 상품명
            if ":" in raw_name:
                _k, _, _v = raw_name.partition(":")
                _v = _v.strip().lstrip(",").strip()
                # 뒤가 "N개" 형태의 수량이 아닌 텍스트일 때만 C형으로 처리
                if _v and not re.match(r"^\d+개?$", _v):
                    names.append(_v)
                    continue
            names.append(raw_name)
        if names:
            return names
    상품명_목록 = clean_text(merged.get("상품명_목록", "") or "")
    return [상품명_목록] if 상품명_목록 else []


def _fetch_shop_product_tax_map(driver, log=None) -> dict:
    """약속처방 상품관리 목록에서 상품별 처방비용(공급가액/부가세/면세금액/과세구분) 수집.

    Returns:
        dict: {상품명(str): {"공급가액": int, "부가세": int, "면세금액": int, "과세구분": str}}
    """
    import re as _re
    import time as _time
    from bs4 import BeautifulSoup as _BS
    from urllib.parse import urljoin as _urljoin

    if log is None:
        log = print

    product_tax: dict = {}

    try:
        # --- 목록 페이지 순회: p_seq 수집 ---
        products: list = []   # [(p_seq, name), ...]
        seen_pseqs: set = set()
        page = 1
        while True:
            url = SHOP_GOODS_LIST_URL if page == 1 else f"{SHOP_GOODS_LIST_URL}?page={page}&s_string=&p_bigpart=&p_smallpart="
            driver.get(url)
            _time.sleep(0.8)
            html = driver.page_source
            soup = _BS(html, "html.parser")

            # p_seq 링크 수집
            found_on_page = 0
            for a in soup.find_all("a", href=True):
                href = a["href"]
                m = _re.search(r"goods_view\.asp\?(?:.*?&)?p_seq=(\d+)", href)
                if not m:
                    continue
                p_seq = m.group(1)
                if p_seq in seen_pseqs:
                    continue
                seen_pseqs.add(p_seq)
                found_on_page += 1
                # 상품명: 링크 텍스트 우선, 없으면 부모 td
                name = clean_text(a.get_text())
                if not name:
                    td = a.find_parent("td")
                    if td:
                        name = clean_text(td.get_text())
                if name:
                    products.append((p_seq, name))

            if found_on_page == 0:
                break  # 더 이상 데이터 없음

            # 다음 페이지 링크 확인
            next_link = soup.find("a", href=lambda h: h and f"page={page + 1}" in h)
            if not next_link:
                break
            page += 1

        log(f"상품관리: {len(products)}개 상품 발견")

        # --- 상세 페이지에서 처방비용 파싱 ---
        # 목록에서 클릭해서 들어가는 goods_view.asp 페이지에서 추출
        for p_seq, name in products:
            detail_url = f"{SHOP_GOODS_VIEW_URL}?p_seq={p_seq}"
            try:
                driver.get(detail_url)
                _time.sleep(0.5)
                raw_html = driver.page_source
                stripped = _re.sub(r"<[^>]+>", " ", raw_html)
                stripped = stripped.replace("&nbsp;", " ").replace(",", "")

                m = _re.search(
                    r"총액\s*([\d]+).*?공급가액\s*([\d]+).*?부가세\s*([\d]+).*?면세금액\s*([\d]+)",
                    stripped,
                )
                if m:
                    공급가액 = int(m.group(2))
                    부가세   = int(m.group(3))
                    면세금액 = int(m.group(4))

                    if 면세금액 > 0 and 공급가액 == 0:
                        과세구분 = "면세"
                    elif 공급가액 > 0 and 면세금액 == 0:
                        과세구분 = "과세"
                    elif 공급가액 > 0 and 면세금액 > 0:
                        과세구분 = "혼합"
                    else:
                        과세구분 = ""

                    product_tax[name] = {
                        "공급가액": 공급가액,
                        "부가세": 부가세,
                        "면세금액": 면세금액,
                        "과세구분": 과세구분,
                    }
                    log(f"  [{name}] 공급가액={공급가액}, 부가세={부가세}, 면세금액={면세금액} → {과세구분}")
                else:
                    log(f"  [{name}] 처방비용 파싱 실패 (p_seq={p_seq})")
            except Exception as _e:
                log(f"  [{name}] 상세 조회 실패 (p_seq={p_seq}): {_e}")

    except Exception as _e:
        log(f"상품관리 조회 오류: {_e}")
        import traceback as _tb
        log(_tb.format_exc())

    return product_tax


def _fetch_cafe_product_tax_map(driver, shop_map: dict = None, log=None) -> dict:
    """카페 상품관리 목록에서 상품별 처방비용 수집.

    카페 페이지는 단일 <input name="price" value="N"> 필드만 있고
    공급가액/부가세/면세금액 세목이 없음.
    → shop_map(약속처방 캐시)에 퍼지 매칭하여 과세구분을 상속받고,
      price로 공급가액/부가세/면세금액을 역산해 저장.
    """
    import re as _re
    import time as _time
    import difflib as _dl
    from bs4 import BeautifulSoup as _BS

    if log is None:
        log = print
    if shop_map is None:
        shop_map = {}

    # ── 퍼지 매칭: 카페 상품명 → 약속처방 캐시 키 ──
    def _fuzzy_match(cafe_name: str):
        """카페 상품명을 shop_map 키에 퍼지 매칭. (matched_key, ref_entry) 반환."""
        if not shop_map:
            return None, None
        cafe_norm = _re.sub(r"\s+", "", cafe_name).lower()

        # ── 0단계: 수동 매핑 사전 우선 적용 ──
        for _cafe_kw, _shop_kw in _CAFE_TO_SHOP_NAME_MAP.items():
            _cafe_kw_norm = _re.sub(r"\s+", "", _cafe_kw).lower()
            if _cafe_kw_norm and _cafe_kw_norm in cafe_norm:
                _shop_kw_norm = _re.sub(r"\s+", "", _shop_kw).lower()
                # 약속처방 캐시에서 대응 키워드 포함하는 첫 항목
                for shop_key in shop_map:
                    if _shop_kw_norm in _re.sub(r"\s+", "", shop_key).lower():
                        return shop_key, shop_map[shop_key]

        # ── 1단계: 퍼지 매칭 (부분 문자열 + SequenceMatcher) ──
        best_key = None
        best_score = 0.0
        for shop_key in shop_map:
            shop_norm = _re.sub(r"\s+", "", shop_key).lower()
            # 한쪽이 다른 쪽의 부분 문자열이면 강한 매칭
            if (len(cafe_norm) >= 3 and len(shop_norm) >= 3
                    and (cafe_norm in shop_norm or shop_norm in cafe_norm)):
                score = 0.9
            else:
                score = _dl.SequenceMatcher(None, cafe_norm, shop_norm).ratio()
            if score > best_score:
                best_score = score
                best_key = shop_key
        if best_key and best_score >= 0.55:
            return best_key, shop_map[best_key]
        return None, None

    # ── price + 과세구분 → 공급가액/부가세/면세금액 역산 ──
    def _compute_breakdown(price: int, 과세구분: str, ref: dict = None) -> dict:
        if 과세구분 == "면세":
            return {"공급가액": 0, "부가세": 0, "면세금액": price, "과세구분": "면세"}
        if 과세구분 == "과세":
            공급가액 = round(price / 1.1)
            부가세 = price - 공급가액
            return {"공급가액": 공급가액, "부가세": 부가세, "면세금액": 0, "과세구분": "과세"}
        if 과세구분 == "혼합" and ref:
            ref_taxable = ref.get("공급가액", 0)
            ref_exempt  = ref.get("면세금액", 0)
            ref_total   = ref_taxable + ref_exempt
            if ref_total > 0:
                면세금액 = round(price * ref_exempt / ref_total)
                과세부분 = price - 면세금액
                공급가액 = round(과세부분 / 1.1)
                부가세 = 과세부분 - 공급가액
                return {"공급가액": 공급가액, "부가세": 부가세, "면세금액": 면세금액, "과세구분": "혼합"}
        # 매칭 실패 or 알 수 없음
        return {"공급가액": 0, "부가세": 0, "면세금액": 0, "과세구분": "?"}

    product_tax: dict = {}

    try:
        # ── 목록 페이지 순회: seqno 수집 ──
        seqnos: list = []
        seen_seqs: set = set()
        page = 1
        while True:
            url = (CAFE_GOODS_LIST_URL if page == 1
                   else f"{CAFE_GOODS_LIST_URL}?page={page}&s_string=&p_bigpart=&p_smallpart=")
            driver.get(url)
            _time.sleep(0.8)
            soup = _BS(driver.page_source, "html.parser")

            found_on_page = 0
            for a in soup.find_all("a", href=True):
                href = a["href"]
                m = _re.search(r"goods_write\.asp\?(?:.*?&)?seqno=(\d+)", href)
                if not m:
                    continue
                seqno = m.group(1)
                if seqno in seen_seqs:
                    continue
                seen_seqs.add(seqno)
                found_on_page += 1
                seqnos.append(seqno)

            if found_on_page == 0:
                break
            if not soup.find("a", href=lambda h: h and f"page={page + 1}" in h):
                break
            page += 1

        log(f"카페 상품관리: {len(seqnos)}개 상품 발견")

        # ── 각 상품 수정 폼에서 goods_name + price input 추출 ──
        for seqno in seqnos:
            detail_url = f"{CAFE_GOODS_WRITE_URL}?seqno={seqno}"
            try:
                driver.get(detail_url)
                _time.sleep(0.5)
                soup2 = _BS(driver.page_source, "html.parser")

                # 상품명: <input name="goods_name" value="...">
                name_inp = soup2.find("input", {"name": "goods_name"})
                goods_name = (name_inp.get("value", "").strip() if name_inp else "")

                # 가격: <input name="price" value="...">
                price_inp = soup2.find("input", {"name": "price"})
                price_str = (price_inp.get("value", "0").replace(",", "") if price_inp else "0")
                try:
                    price = int(price_str)
                except ValueError:
                    price = 0

                if not goods_name:
                    log(f"  [카페] seqno={seqno}: 상품명 추출 실패, 건너뜀")
                    continue
                if price <= 0:
                    log(f"  [카페][{goods_name}] price=0, 건너뜀")
                    continue

                # 약속처방 캐시에 퍼지 매칭하여 과세구분 상속
                matched_key, ref_entry = _fuzzy_match(goods_name)
                if matched_key:
                    과세구분 = ref_entry.get("과세구분", "")
                    entry = _compute_breakdown(price, 과세구분, ref_entry)
                    log(f"  [카페][{goods_name}] price={price:,} → 매칭=[{matched_key}]({과세구분}) "
                        f"공급가액={entry['공급가액']:,} 부가세={entry['부가세']:,} 면세금액={entry['면세금액']:,}")
                else:
                    entry = _compute_breakdown(price, "")
                    log(f"  [카페][{goods_name}] price={price:,} → 약속처방 매칭 실패, 과세구분=?")

                product_tax[goods_name] = entry

            except Exception as _e:
                log(f"  [카페] seqno={seqno} 상세 조회 실패: {_e}")

    except Exception as _e:
        log(f"카페 상품관리 조회 오류: {_e}")
        import traceback as _tb
        log(_tb.format_exc())

    return product_tax


def _fetch_all_product_tax_map(driver, log=None) -> dict:
    """약속처방 + 카페 상품 모두 수집하여 통합 dict 반환."""
    if log is None:
        log = print
    log("=== 약속처방 상품 수집 시작 ===")
    shop_map = _fetch_shop_product_tax_map(driver, log=log)
    log("=== 카페 상품 수집 시작 ===")
    cafe_map = _fetch_cafe_product_tax_map(driver, shop_map=shop_map, log=log)
    merged = dict(shop_map)
    merged.update(cafe_map)
    log(f"=== 전체 수집 완료: 약속처방 {len(shop_map)}개 + 카페 {len(cafe_map)}개 → 통합 {len(merged)}개 ===")
    return merged


def _fetch_shop_product_all_details(driver, log=None) -> list:
    """약속처방 상품관리 목록 전체 상세 내역 수집 → 엑셀 저장용.

    Returns:
        list[dict]: 상품별 전체 필드 딕셔너리 목록.
                    각 dict 키는 페이지에서 동적으로 추출되므로 상품마다 다를 수 있음.
    """
    import re as _re
    import time as _time
    from bs4 import BeautifulSoup as _BS

    if log is None:
        log = print

    all_products: list = []

    try:
        # ── 1. 목록 페이지 순회: 테이블 전체 행 데이터 + p_seq 수집 ──
        products: list = []   # list[dict] — 목록 테이블 컬럼 값 포함
        seen_pseqs: set = set()
        page = 1
        while True:
            url = (SHOP_GOODS_LIST_URL if page == 1
                   else f"{SHOP_GOODS_LIST_URL}?page={page}&s_string=&p_bigpart=&p_smallpart=")
            driver.get(url)
            _time.sleep(0.8)
            soup = _BS(driver.page_source, "html.parser")

            found_on_page = 0
            for tr in soup.find_all("tr"):
                # goods_view.asp?p_seq=... 링크가 있는 행만 처리
                a_tags = [a for a in tr.find_all("a", href=True)
                          if _re.search(r"goods_view\.asp\?.*?p_seq=(\d+)", a["href"])]
                if not a_tags:
                    continue
                m_seq = _re.search(r"goods_view\.asp\?.*?p_seq=(\d+)", a_tags[0]["href"])
                if not m_seq:
                    continue
                p_seq = m_seq.group(1)
                if p_seq in seen_pseqs:
                    continue
                seen_pseqs.add(p_seq)
                found_on_page += 1

                row_data: dict = {"p_seq": p_seq}

                # 상품명: 링크 텍스트
                row_data["상품명"] = clean_text(a_tags[0].get_text()) or ""

                # 목록 구조: 번호 | 상품명(링크) | 판매가 | 처방상태 | 활성 | 등록일 | ...
                # 처방상태/활성/가격은 상세 페이지에서 정확히 수집하므로 목록에서는 등록일(4번째 형제)만 추출
                name_td = a_tags[0].find_parent("td")
                if name_td:
                    sibs = name_td.find_next_siblings("td")
                    if len(sibs) >= 4:
                        val = clean_text(sibs[3].get_text())
                        if val and len(val) <= 30:
                            row_data["등록일"] = val

                products.append(row_data)

            if found_on_page == 0:
                break
            if not soup.find("a", href=lambda h: h and f"page={page + 1}" in h):
                break
            page += 1

        log(f"상품관리: {len(products)}개 상품 발견 — 상세 수집 시작")

        # ── 2. 각 상품 상세 페이지에서 추가 필드 + 처방비용 파싱 ──
        for idx, row_data in enumerate(products, 1):
            p_seq = row_data["p_seq"]
            name = row_data.get("상품명", "") or p_seq
            log(f"  [{idx}/{len(products)}] {name} (p_seq={p_seq})")
            detail_url = f"{SHOP_GOODS_VIEW_URL}?p_seq={p_seq}"
            try:
                driver.get(detail_url)
                _time.sleep(0.5)
                soup2 = _BS(driver.page_source, "html.parser")

                # 상세 페이지 파싱: bgcolor="D8E5F0" 셀이 라벨
                # 제외 항목: 이미지(이진), 구성(재료 select), 내용(HTML img), 처방비용(regex로 별도 추출)
                _SKIP_KEYS = {"이미지", "구성", "내용", "처방비용"}
                for td_lbl in soup2.find_all("td", bgcolor="D8E5F0"):
                    key = clean_text(td_lbl.get_text())
                    if not key or key in _SKIP_KEYS:
                        continue
                    td_val = td_lbl.find_next_sibling("td")
                    if td_val is None or td_val.get("bgcolor", "").upper() == "D8E5F0":
                        continue
                    sel = td_val.find("select")
                    if sel:
                        opts = [clean_text(o.get_text()) for o in sel.find_all("option")
                                if clean_text(o.get_text()) and "없습니다" not in o.get_text()]
                        val = ", ".join(opts) if opts else ""
                    else:
                        val = clean_text(td_val.get_text())
                    if key and key not in row_data:
                        row_data[key] = val

                # 처방비용 숫자 보정 (쉼표 제거 후 regex)
                raw = _re.sub(r"<[^>]+>", " ", driver.page_source)
                raw = raw.replace("&nbsp;", " ").replace(",", "")
                mc = _re.search(
                    r"총액\s*([\d]+).*?공급가액\s*([\d]+).*?부가세\s*([\d]+).*?면세금액\s*([\d]+)",
                    raw,
                )
                if mc:
                    row_data["처방비용_총액"]    = int(mc.group(1))
                    row_data["처방비용_공급가액"] = int(mc.group(2))
                    row_data["처방비용_부가세"]   = int(mc.group(3))
                    row_data["처방비용_면세금액"] = int(mc.group(4))
                    공급 = int(mc.group(2))
                    면세 = int(mc.group(4))
                    if 면세 > 0 and 공급 == 0:
                        row_data["처방비용_과세구분"] = "면세"
                    elif 공급 > 0 and 면세 == 0:
                        row_data["처방비용_과세구분"] = "과세"
                    elif 공급 > 0 and 면세 > 0:
                        row_data["처방비용_과세구분"] = "혼합"
                    else:
                        row_data["처방비용_과세구분"] = ""

            except Exception as _e:
                log(f"    상세 조회 실패: {_e}")
                row_data["_오류"] = str(_e)

            row_data["구분"] = "H"
            all_products.append(row_data)

    except Exception as _e:
        log(f"상품 전체 내역 조회 오류: {_e}")
        import traceback as _tb
        log(_tb.format_exc())

    return all_products


def _fetch_cafe_product_all_details(driver, shop_map: dict = None, log=None) -> list:
    """카페 상품관리 목록 전체 상세 내역 수집 → 엑셀 저장용.

    - goods_write.asp 폼에서 bgcolor=D8E5F0 라벨 기준으로 필드 추출
    - 처방비용(price)을 shop_map 퍼지 매칭으로 과세구분 판단 후 약속처방과 동일한
      처방비용_총액 / 처방비용_공급가액 / 처방비용_부가세 / 처방비용_면세금액 / 처방비용_과세구분 산출
    구분 열: 'C'
    """
    import re as _re
    import time as _time
    import difflib as _dl
    from bs4 import BeautifulSoup as _BS

    if log is None:
        log = print
    if shop_map is None:
        shop_map = {}

    # ── 퍼지 매칭: 카페 상품명 → shop_map 키 ──
    def _fmatch(cafe_name: str):
        if not shop_map:
            return None, None
        cn = _re.sub(r"\s+", "", cafe_name).lower()
        for kw, skw in _CAFE_TO_SHOP_NAME_MAP.items():
            kn = _re.sub(r"\s+", "", kw).lower()
            if kn and kn in cn:
                sn = _re.sub(r"\s+", "", skw).lower()
                for k, v in shop_map.items():
                    if sn in _re.sub(r"\s+", "", k).lower():
                        return k, v
        best_k, best_s = None, 0.0
        for k, v in shop_map.items():
            kn = _re.sub(r"\s+", "", k).lower()
            if len(cn) >= 3 and len(kn) >= 3:
                s = 0.9 if (cn in kn or kn in cn) else _dl.SequenceMatcher(None, cn, kn).ratio()
            else:
                s = 0.0
            if s > best_s:
                best_s, best_k = s, k
        return (best_k, shop_map[best_k]) if best_k and best_s >= 0.55 else (None, None)

    # ── 과세구분 → 공급가액/부가세/면세금액 역산 ──
    def _breakdown(price: int, 과세구분: str, ref: dict = None) -> dict:
        base = {"처방비용_총액": price}
        if 과세구분 == "면세":
            return {**base, "처방비용_공급가액": 0, "처방비용_부가세": 0,
                    "처방비용_면세금액": price, "처방비용_과세구분": "면세"}
        if 과세구분 == "과세":
            공급 = round(price / 1.1)
            return {**base, "처방비용_공급가액": 공급, "처방비용_부가세": price - 공급,
                    "처방비용_면세금액": 0, "처방비용_과세구분": "과세"}
        if 과세구분 == "혼합" and ref:
            rt = ref.get("공급가액", 0); re_ = ref.get("면세금액", 0); rtot = rt + re_
            if rtot > 0:
                exempt = round(price * re_ / rtot)
                taxable = price - exempt
                공급 = round(taxable / 1.1)
                return {**base, "처방비용_공급가액": 공급, "처방비용_부가세": taxable - 공급,
                        "처방비용_면세금액": exempt, "처방비용_과세구분": "혼합"}
        return {**base, "처방비용_공급가액": 0, "처방비용_부가세": 0,
                "처방비용_면세금액": 0, "처방비용_과세구분": "?"}

    all_products: list = []

    try:
        # ── 1. 목록 페이지 순회: seqno 수집 ──
        seqnos: list = []
        seen_seqs: set = set()
        page = 1
        while True:
            url = (CAFE_GOODS_LIST_URL if page == 1
                   else f"{CAFE_GOODS_LIST_URL}?page={page}&s_string=&p_bigpart=&p_smallpart=")
            driver.get(url)
            _time.sleep(0.8)
            soup = _BS(driver.page_source, "html.parser")

            found_on_page = 0
            for a in soup.find_all("a", href=True):
                m = _re.search(r"goods_write\.asp\?(?:.*?&)?seqno=(\d+)", a["href"])
                if not m:
                    continue
                seqno = m.group(1)
                if seqno in seen_seqs:
                    continue
                seen_seqs.add(seqno)
                found_on_page += 1
                seqnos.append(seqno)

            if found_on_page == 0:
                break
            if not soup.find("a", href=lambda h: h and f"page={page + 1}" in h):
                break
            page += 1

        log(f"카페 상품관리: {len(seqnos)}개 상품 발견 — 상세 수집 시작")

        # ── 2. 각 상품 수정 폼에서 필드 추출 + 처방비용 계산 ──
        for idx, seqno in enumerate(seqnos, 1):
            row_data: dict = {"seqno": seqno, "구분": "C"}
            try:
                driver.get(f"{CAFE_GOODS_WRITE_URL}?seqno={seqno}")
                _time.sleep(0.5)
                soup2 = _BS(driver.page_source, "html.parser")

                # bgcolor="D8E5F0" 라벨 → 다음 형제 td 값 (input/select/text)
                for td_lbl in soup2.find_all("td", bgcolor="D8E5F0"):
                    key = clean_text(td_lbl.get_text())
                    if not key:
                        continue
                    td_val = td_lbl.find_next_sibling("td")
                    if td_val is None or td_val.get("bgcolor", "").upper() == "D8E5F0":
                        continue
                    sel = td_val.find("select")
                    inp = td_val.find("input")
                    if sel:
                        opt = sel.find("option", selected=True)
                        val = clean_text(opt.get_text()) if opt else ""
                    elif inp:
                        val = (inp.get("value") or "").strip()
                    else:
                        val = clean_text(td_val.get_text())
                    if key and key not in row_data:
                        row_data[key] = val

                # 할인적용(N개이상) / 할인_금액 별도 추출
                if "할인적용" in row_data:
                    row_data["할인_N개이상"] = row_data.pop("할인적용")
                # sale_price input: N개이상 주문시 할인 금액
                sp_inp = soup2.find("input", {"name": "sale_price"})
                if sp_inp:
                    sp_val = (sp_inp.get("value") or "").strip()
                    if sp_val:
                        row_data["할인_금액"] = sp_val

                # 처방명 → 상품명 표준 열
                if "처방명" in row_data:
                    row_data["상품명"] = row_data["처방명"]

                # 처방비용(price) → 과세구분 매칭 후 공급가액/부가세 역산
                price_raw = row_data.get("처방비용", "0") or "0"
                try:
                    price = int(_re.sub(r"[^\d]", "", str(price_raw)))
                except ValueError:
                    price = 0

                cafe_name = row_data.get("상품명", "")
                matched_key, ref_entry = _fmatch(cafe_name)
                과세구분 = ref_entry.get("과세구분", "") if ref_entry else ""
                if matched_key:
                    log(f"  [카페][{idx}/{len(seqnos)}] {cafe_name} → [{matched_key}]({과세구분})")
                else:
                    log(f"  [카페][{idx}/{len(seqnos)}] {cafe_name} → 매칭 실패(과세구분=?)")
                    과세구분 = "?"
                row_data.update(_breakdown(price, 과세구분, ref_entry))

            except Exception as _e:
                log(f"  [카페] seqno={seqno} 상세 조회 실패: {_e}")
                row_data["_오류"] = str(_e)

            all_products.append(row_data)

    except Exception as _e:
        log(f"카페 상품 전체 내역 조회 오류: {_e}")
        import traceback as _tb
        log(_tb.format_exc())

    return all_products


def build_shop_order_cj_df(results: list) -> pd.DataFrame:
    """약속처방 주문 목록 → 대한통운 파일 업로드 양식 DataFrame"""
    rows = []
    for r in results:
        # 취소/환불 제외
        status = clean_text(r.get("진행상태") or r.get("진행상태_상세") or "")
        if "취소" in status or "환불" in status:
            continue
        # 받는분 주소 없으면 제외
        recv_addr = clean_text(r.get("받는분_주소", "") or "")
        if not recv_addr:
            continue

        order_no = clean_text(r.get("주문번호", ""))
        order_date = clean_text(r.get("주문일", ""))
        # 날짜 형식 변환 "26/05/11 16:16" → "2026-05-11"
        try:
            from datetime import datetime as _dt
            dt = _dt.strptime(order_date[:8], "%y/%m/%d")
            order_date = dt.strftime("%Y-%m-%d")
        except Exception:
            pass

        clinic = clean_text(r.get("한의원명", "") or "")
        recv_name = clean_text(r.get("받는분", "") or "")
        # 받는분 전화: 휴대 우선, 없으면 전화
        recv_phone = clean_text(r.get("받는분_휴대", "") or r.get("받는분_전화", "") or "")

        # 실제 주문 내용(주문상품) 우선, 없으면 상품명_목록 폴백
        _raw_product = clean_text(r.get("주문상품", "") or r.get("상품명_목록", "") or "")
        product = _SHOP_PRODUCT_NAME_MAP.get(_raw_product, _raw_product)

        # 보내는분 정보
        sender_name = clean_text(r.get("보내는분", "") or "")
        sender_phone = clean_text(r.get("보내는분_휴대", "") or r.get("보내는분_전화", "") or "")
        sender_addr = clean_text(r.get("보내는분_주소", "") or "")
        if not sender_name:
            sender_name  = _KGINBIO_SENDER["name"]
            sender_phone = _KGINBIO_SENDER["phone"]
            sender_addr  = _KGINBIO_SENDER["address"]

        memo = clean_text(r.get("배송시메모", "") or "")

        rows.append([
            order_date, order_no, clinic,
            recv_name, recv_phone, recv_addr,
            product, "신용", 2800, 1,
            sender_name, sender_phone, sender_addr,
            memo,
        ])
    return pd.DataFrame(rows, columns=_SHOP_CJ_COLUMNS)


def run_sp_delivery_job(detail_excel_path: str, start_date: str = "", end_date: str = "",
                        log_callback=None, progress_callback=None, cancel_check=None):
    """약속처방 송장번호 입력: CJ 엑셀 → 배송준비 주문 매칭 → 송장 입력 + 배송중(3) 전환"""
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    def update_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    driver = None
    try:
        # 1. CJ 엑셀 읽기
        update_progress(5, "파일 읽는 중...")
        detail_df = pd.read_excel(detail_excel_path, header=0, dtype=str)

        tracking_col = None
        ordercode_col = None
        for col in detail_df.columns:
            col_c = clean_text(str(col))
            if "운송장번호" in col_c:
                tracking_col = col
            if "고객주문번호" in col_c:
                ordercode_col = col

        if not tracking_col:
            raise Exception(f"운송장번호 컬럼을 찾을 수 없어요.\n찾은 컬럼: {list(detail_df.columns)}")
        if not ordercode_col:
            raise Exception("고객주문번호 컬럼을 찾을 수 없어요. (약속처방은 주문번호 기준 매칭)")

        detail_by_ordercode = {}
        for _, row in detail_df.iterrows():
            tracking_raw = str(row.get(tracking_col, "")).strip()
            ordercode_raw = clean_text(str(row.get(ordercode_col, "")))
            if tracking_raw in ("nan", "") or not ordercode_raw or ordercode_raw == "nan":
                continue
            tracking_clean = re.sub(r"[-\s]", "", tracking_raw).strip()
            if tracking_clean:
                for code in ordercode_raw.split("/"):
                    code = code.strip()
                    if code:
                        detail_by_ordercode[code] = tracking_clean
        log(f"CJ 파일: {len(detail_by_ordercode)}건 읽음 (주문번호 기준)")

        if not detail_by_ordercode:
            raise NoWorkFound("CJ 파일에서 매칭 가능한 주문이 없어요.")

        filter_start_dt = parse_filter_datetime(start_date)
        filter_end_dt   = parse_filter_datetime(end_date)
        if filter_start_dt:
            log(f"⏰ 시작 필터: {start_date}")
        if filter_end_dt:
            log(f"⏰ 종료 필터: {end_date}")

        # 2. 로그인 + 약속처방 배송준비(2) 목록 수집
        update_progress(10, "로그인 중...")
        _opts = Options()
        _opts.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=_opts)
        login_driver(driver, ADMIN_ID, ADMIN_PW)

        update_progress(15, "약속처방 주문 목록 수집 중...")

        driver.get(SHOP_ORDER_LIST_URL)
        time.sleep(0.8)
        driver.execute_script(f"""
            var f = document.form2 || document.forms[0];
            if (!f) return;
            f.s_date.value     = '{start_date}';
            f.e_date.value     = '{end_date}';
            f.search.value     = 'name';
            f.s_string.value   = '';
            f.order_ings.value = '2';
        """)
        driver.execute_script("(document.form2 || document.forms[0]).submit();")
        time.sleep(1.2)

        all_sp_orders = []
        seen_order_nos = set()
        page = 1
        while page <= MAX_PAGE_SAFETY_LIMIT:
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                break
            html = driver.page_source
            rows = _parse_shop_order_list_page(html)
            if not rows:
                log(f"페이지 {page}: 데이터 없음 → 수집 완료")
                break
            new_rows = [r for r in rows if r["주문번호"] not in seen_order_nos]
            for r in new_rows:
                seen_order_nos.add(r["주문번호"])
            all_sp_orders.extend(new_rows)
            log(f"페이지 {page}: {len(rows)}건 (신규 {len(new_rows)}건)")

            soup_pg = BeautifulSoup(html, "html.parser")
            next_link = soup_pg.find("a", href=lambda h: h and f"page={page+1}" in h)
            if not next_link:
                break
            next_href = next_link["href"]
            if not next_href.startswith("http"):
                next_href = urljoin(f"{BASE_URL}/shop_order/", next_href)
            driver.get(next_href)
            time.sleep(0.8)
            page += 1

        log(f"약속처방 배송준비 {len(all_sp_orders)}건 수집")

        # 날짜 필터
        if filter_start_dt or filter_end_dt:
            filtered = []
            for r in all_sp_orders:
                order_dt = parse_order_datetime_obj(r.get("주문일", ""))
                if order_dt:
                    order_date_only = order_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                    if filter_start_dt:
                        start_date_only = filter_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                        if order_date_only < start_date_only:
                            continue
                    if filter_end_dt:
                        end_date_only = filter_end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                        if order_date_only > end_date_only:
                            continue
                filtered.append(r)
            all_sp_orders = filtered
            log(f"날짜 필터 후: {len(all_sp_orders)}건")

        if not all_sp_orders:
            raise NoWorkFound("배송준비 상태의 약속처방 주문이 없어요.")

        # 3. 매칭 + 송장번호 입력 + 배송중 전환
        success_count = 0
        fail_count = 0
        skip_count = 0

        for i, order in enumerate(all_sp_orders):
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                update_progress(0, "취소됨")
                break

            pct = 20 + int(75 * (i + 1) / len(all_sp_orders))
            update_progress(pct, f"송장 입력 중... ({i+1}/{len(all_sp_orders)})")

            order_no = order["주문번호"]
            tracking = detail_by_ordercode.get(order_no)
            if not tracking:
                skip_count += 1
                log(f"- 매칭 없음: {order_no} ({order.get('한의원명', '')})")
                continue

            try:
                detail_href = order["detail_href"]
                detail_url = detail_href if detail_href.startswith("http") else urljoin(f"{BASE_URL}/shop_order/", detail_href)
                driver.get(detail_url)
                time.sleep(1.0)

                # 송장번호 필드 입력
                inp = None
                for field_name in ("deliveryno", "delivery_no"):
                    try:
                        inp = WebDriverWait(driver, 8).until(
                            EC.presence_of_element_located((By.NAME, field_name))
                        )
                        break
                    except Exception:
                        pass
                if not inp:
                    raise Exception("송장번호 입력 필드를 찾을 수 없어요")
                inp.clear()
                inp.send_keys(tracking)

                # 진행상태 → 배송중(3)
                try:
                    sel_el = driver.find_element(By.NAME, "order_ing")
                    Select(sel_el).select_by_value("3")
                    time.sleep(0.3)
                except Exception:
                    pass

                # 폼 제출
                driver.execute_script(
                    "if(typeof order_change==='function'){ order_change(); }"
                    "else if(typeof delivery_insert==='function'){ delivery_insert(); }"
                    "else{ document.forms[0].submit(); }"
                )
                time.sleep(1.5)

                success_count += 1
                log(f"✓ 발송처리: {order_no} {order.get('한의원명', '')} → 송장 {tracking}")

            except Exception as e:
                fail_count += 1
                log(f"✗ 실패: {order_no} {order.get('한의원명', '')} → {e}")

        update_progress(100, "완료")
        log(f"\n완료: 발송처리 {success_count}건 / 매칭 없음 {skip_count}건 / 실패 {fail_count}건")

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def run_sp_complete_job(start_date: str = "", end_date: str = "",
                        log_callback=None, progress_callback=None, cancel_check=None):
    """약속처방 배송완료 전환: 배송중(3) 주문 CJ 추적 확인 → 완료(4) 전환"""
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    def update_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)

    driver = None
    try:
        filter_start_dt = parse_filter_datetime(start_date)
        filter_end_dt   = parse_filter_datetime(end_date)
        if filter_start_dt:
            log(f"⏰ 시작 필터: {start_date}")
        if filter_end_dt:
            log(f"⏰ 종료 필터: {end_date}")

        update_progress(5, "로그인 중...")
        _opts = Options()
        _opts.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=_opts)
        login_driver(driver, ADMIN_ID, ADMIN_PW)

        # 1. 배송중(3) 약속처방 목록 수집
        update_progress(10, "약속처방 배송중 목록 수집 중...")

        driver.get(SHOP_ORDER_LIST_URL)
        time.sleep(0.8)
        driver.execute_script(f"""
            var f = document.form2 || document.forms[0];
            if (!f) return;
            f.s_date.value     = '{start_date}';
            f.e_date.value     = '{end_date}';
            f.search.value     = 'name';
            f.s_string.value   = '';
            f.order_ings.value = '3';
        """)
        driver.execute_script("(document.form2 || document.forms[0]).submit();")
        time.sleep(1.2)

        shipped_orders = []
        seen_order_nos = set()
        page = 1
        while page <= MAX_PAGE_SAFETY_LIMIT:
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                break
            html = driver.page_source
            rows = _parse_shop_order_list_page(html)
            if not rows:
                log(f"페이지 {page}: 데이터 없음 → 수집 완료")
                break

            for r in rows:
                if r["주문번호"] in seen_order_nos:
                    continue
                seen_order_nos.add(r["주문번호"])

                # 날짜 필터
                if filter_start_dt or filter_end_dt:
                    order_dt = parse_order_datetime_obj(r.get("주문일", ""))
                    if order_dt:
                        order_date_only = order_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                        if filter_start_dt:
                            start_date_only = filter_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if order_date_only < start_date_only:
                                continue
                        if filter_end_dt:
                            end_date_only = filter_end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                            if order_date_only > end_date_only:
                                continue

                tracking = r.get("송장번호", "").strip()
                if not tracking:
                    continue
                shipped_orders.append({
                    "주문번호": r["주문번호"],
                    "한의원명": r.get("한의원명", ""),
                    "detail_href": r["detail_href"],
                    "송장번호": tracking,
                })

            log(f"페이지 {page}: {len(rows)}건")

            soup_pg = BeautifulSoup(html, "html.parser")
            next_link = soup_pg.find("a", href=lambda h: h and f"page={page+1}" in h)
            if not next_link:
                break
            next_href = next_link["href"]
            if not next_href.startswith("http"):
                next_href = urljoin(f"{BASE_URL}/shop_order/", next_href)
            driver.get(next_href)
            time.sleep(0.8)
            page += 1

        log(f"\n약속처방 배송중 {len(shipped_orders)}건 수집")

        if not shipped_orders:
            raise NoWorkFound("배송완료 전환 대상 약속처방 주문이 없어요.")

        # 2. CJ 배송현황 확인 → 완료(4) 전환
        success_count = 0
        fail_count = 0
        skip_count = 0

        for i, order in enumerate(shipped_orders):
            if cancel_check and cancel_check():
                log("⛔ 취소됨")
                update_progress(0, "취소됨")
                break

            pct = 20 + int(75 * (i + 1) / len(shipped_orders))
            update_progress(pct, f"배송현황 확인 중... ({i+1}/{len(shipped_orders)})")

            try:
                tracking_url = f"https://trace.cjlogistics.com/next/tracking.html?wblNo={order['송장번호']}"
                driver.get(tracking_url)
                try:
                    WebDriverWait(driver, 15).until(
                        lambda d: len(d.find_element(By.ID, "statusDetail").text.strip()) > 0
                    )
                except Exception:
                    time.sleep(3)

                if "배송완료" in driver.page_source:
                    detail_href = order["detail_href"]
                    detail_url = detail_href if detail_href.startswith("http") else urljoin(f"{BASE_URL}/shop_order/", detail_href)
                    driver.get(detail_url)
                    time.sleep(1.0)

                    try:
                        sel_el = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.NAME, "order_ing"))
                        )
                        Select(sel_el).select_by_value("4")
                        time.sleep(0.3)
                        driver.execute_script(
                            "if(typeof order_change==='function'){ order_change(); }"
                            "else{ document.forms[0].submit(); }"
                        )
                        time.sleep(1.5)
                    except Exception as se:
                        log(f"✗ 상태 전환 실패 ({order['주문번호']}): {se}")
                        fail_count += 1
                        continue

                    success_count += 1
                    log(f"✓ 완료 전환: {order['주문번호']} {order['한의원명']} (송장: {order['송장번호']})")
                else:
                    skip_count += 1
                    log(f"- 배송중: {order['주문번호']} {order['한의원명']} (송장: {order['송장번호']})")

            except Exception as e:
                fail_count += 1
                log(f"✗ 실패: {order['주문번호']} {order['한의원명']} → {e}")

        update_progress(100, "완료")
        log(f"\n완료: 전환 {success_count}건 / 배송중 {skip_count}건 / 실패 {fail_count}건")

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def run_shop_order_job(settings: dict, progress_callback=None, log_callback=None):
    """약속처방 주문 목록 수집 + 상세 조회 + 파일 저장"""
    driver = None

    def update_progress(pct: int, msg: str):
        if progress_callback:
            progress_callback(pct, msg)

    _sp_log_lines: list = []
    _sp_output_root = [""]   # finally에서 참조하기 위한 mutable 컨테이너

    def log(msg: str):
        _sp_log_lines.append(str(msg))
        print(msg)
        if log_callback:
            log_callback(msg)

    try:
        from selenium.webdriver.common.by import By
        import time as _time

        start_date    = settings.get("start_date", "")
        end_date      = settings.get("end_date", "")
        status_codes  = settings.get("status_codes", [""])   # 리스트, [""] = 전체
        route_filter  = settings.get("route_filter", set())  # {"H","C"} 또는 비어있으면 전체
        search_type   = settings.get("search_type", "name")
        search_str    = settings.get("search_str", "")
        max_pages     = settings.get("max_pages", 50)
        save_excel    = settings.get("save_excel", True)
        save_cj       = settings.get("save_cj", False)
        cancel_event  = settings.get("cancel_event", None)
        custom_output_dir = settings.get("output_dir", "").strip()

        def _cancelled():
            return cancel_event is not None and cancel_event.is_set()

        if custom_output_dir and os.path.isdir(custom_output_dir):
            base_dir = custom_output_dir
        elif getattr(sys, "frozen", False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))

        run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_root = os.path.join(base_dir, f"약속처방_{run_ts}")
        _sp_output_root[0] = output_root

        update_progress(5, "드라이버 시작 중...")
        from selenium.webdriver.chrome.options import Options as _COptions
        _opts = _COptions()
        _opts.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=_opts)
        login_driver(driver, ADMIN_ID, ADMIN_PW)
        update_progress(10, "로그인 완료")

        # --- 상품 면세/과세 맵 빌드 ---
        # 캐시 있으면 항상 자동 사용. 없을 때만 사용자 선택(download_goods_first)으로 다운로드.
        _cached_goods, _cached_meta = _load_goods_tax_cache()
        if _cached_goods:
            product_tax_map = _cached_goods
            log(f"상품 캐시 사용: {len(product_tax_map)}개 "
                f"({_cached_meta.get('updated_at', '?')} 기준)")
        elif settings.get("download_goods_first", False):
            log("상품 캐시 없음 → 다운로드 시작...")
            update_progress(12, "상품 면세/과세 정보 수집 중...")
            product_tax_map = _fetch_all_product_tax_map(driver, log=log)
            if product_tax_map:
                _save_goods_tax_cache(product_tax_map)
                log(f"상품 캐시 저장 완료: {len(product_tax_map)}개")
            else:
                log("⚠ 상품 정보 수집 실패 — 면세/과세 열 빈값으로 진행")
                product_tax_map = {}
        else:
            product_tax_map = {}
            log("⚠ 상품 캐시 없음 — 면세/과세 열 빈값으로 진행")
        update_progress(15, "상품 정보 준비 완료")

        # --- 목록 수집 ---
        # 폼이 POST 방식이므로 1페이지는 JS 폼 제출, 2페이지부터는 페이지 링크 클릭
        def _submit_page1(status_code):
            driver.get(SHOP_ORDER_LIST_URL)
            _time.sleep(0.8)
            driver.execute_script(f"""
                var f = document.form2 || document.forms[0];
                if (!f) return;
                f.s_date.value     = '{start_date}';
                f.e_date.value     = '{end_date}';
                f.search.value     = '{search_type}';
                f.s_string.value   = '{search_str}';
                f.order_ings.value = '{status_code}';
            """)
            driver.execute_script("(document.form2 || document.forms[0]).submit();")
            _time.sleep(1.2)

        from bs4 import BeautifulSoup as _BS
        all_list_rows = []
        seen_order_nos = set()
        for status_code in status_codes:
            if _cancelled():
                break
            log(f"진행상태 코드={status_code or '전체'} 수집 시작")
            _submit_page1(status_code)
            page = 1
            while page <= max_pages:
                if _cancelled():
                    log("취소 요청 → 목록 수집 중단")
                    break
                html = driver.page_source
                rows = _parse_shop_order_list_page(html)
                if not rows:
                    log(f"페이지 {page}: 데이터 없음 → 수집 완료")
                    break
                new_rows = [r for r in rows if r["주문번호"] not in seen_order_nos]
                for r in new_rows:
                    seen_order_nos.add(r["주문번호"])
                all_list_rows.extend(new_rows)
                log(f"페이지 {page}: {len(rows)}건 (신규 {len(new_rows)}건, 누계 {len(all_list_rows)}건)")
                # 다음 페이지 링크 찾아서 직접 클릭
                soup = _BS(html, "html.parser")
                next_link = soup.find("a", href=lambda h: h and f"page={page+1}" in h)
                if not next_link:
                    log("마지막 페이지 도달")
                    break
                next_href = next_link["href"]
                if not next_href.startswith("http"):
                    from urllib.parse import urljoin
                    next_href = urljoin(f"{BASE_URL}/shop_order/", next_href)
                driver.get(next_href)
                _time.sleep(0.8)
                page += 1

        total = len(all_list_rows)
        log(f"목록 총 {total}건 수집")

        # 경로 필터 적용 (H/C 중 선택된 것만)
        if route_filter and route_filter != {"H", "C"}:
            before = total
            all_list_rows = [r for r in all_list_rows if r.get("경로", "") in route_filter]
            total = len(all_list_rows)
            log(f"경로 필터({'/'.join(sorted(route_filter))}): {before}건 → {total}건")

        if total == 0:
            update_progress(100, "수집 완료 (0건)")
            return

        # --- 상세 조회 ---
        master_results = []
        _fuzzy_lookup_cache: dict = {}  # 상품명 퍼지 매칭 결과 메모이제이션 {name: cache_key or None}
        for idx, row in enumerate(all_list_rows):
            if _cancelled():
                log("취소 요청 → 상세 조회 중단")
                break
            pct = 10 + int(80 * idx / total)
            update_progress(pct, f"상세 조회 중 ({idx+1}/{total}): {row['주문번호']}")

            route = row["경로"]
            order_no = row["주문번호"]
            if route == "C":
                detail_url = f"{SHOP_ORDER_VIEW22_URL}?order_no={order_no}"
            else:
                detail_url = f"{SHOP_ORDER_VIEW_URL}?order_no={order_no}"

            try:
                driver.get(detail_url)
                _time.sleep(0.5)
                html = driver.page_source
                if route == "C":
                    detail = _parse_shop_order_detail_c(html)
                else:
                    detail = _parse_shop_order_detail_h(html)
            except Exception as e:
                log(f"  상세 조회 실패 ({order_no}): {e}")
                detail = {}

            merged = {**row, **detail}
            # 진행상태: 상세 페이지 우선
            if merged.get("진행상태_상세"):
                merged["진행상태"] = merged["진행상태_상세"]
            # 송장번호: 상세 페이지 우선
            if merged.get("송장번호_상세"):
                merged["송장번호"] = merged["송장번호_상세"]
            # 주문 링크
            merged["주문링크"] = detail_url

            # 주문상품 파싱: [(이름, 수량), ...]
            merged["_product_list"] = _parse_order_products(
                clean_text(merged.get("주문상품", "") or "")
            )

            # 상품별 면세/과세 정보 수집 (품목 수만큼 리스트로 보관)
            _prod_names = _extract_product_names_for_tax(merged)
            _tax_list = []
            for _n in _prod_names:
                _hit = product_tax_map.get(_n)
                if _hit is None:
                    # 공백 정규화 후 재시도
                    _n_norm = re.sub(r"\s+", " ", _n.strip())
                    _hit = next(
                        (v for k, v in product_tax_map.items()
                         if re.sub(r"\s+", " ", k.strip()) == _n_norm),
                        None,
                    )
                if _hit is None:
                    # ── 퍼지 매칭 시도: 공백 제거 후 부분 문자열 / SequenceMatcher ──
                    if _n in _fuzzy_lookup_cache:
                        _fkey = _fuzzy_lookup_cache[_n]
                        _hit = product_tax_map[_fkey] if _fkey else {}
                        if not _fkey:
                            log(f"  ⚠ 캐시 미매칭: '{_n}'")
                    else:
                        import difflib as _dl2
                        _n_c = re.sub(r"\s+", "", _n).lower()
                        # ── 0단계: 수동 매핑 우선 ──
                        _bk, _bs = None, 0.0
                        for _okw, _tkw in _ORDER_PRODUCT_NAME_MAP.items():
                            if re.sub(r"\s+", "", _okw).lower() in _n_c:
                                _tkw_c = re.sub(r"\s+", "", _tkw).lower()
                                for _k2 in product_tax_map:
                                    if _tkw_c in re.sub(r"\s+", "", _k2).lower():
                                        log(f"  수동 매핑: '{_n}' → '{_k2}'")
                                        _bk, _bs = _k2, 1.0
                                        break
                                if _bk:
                                    break
                        # ── 1단계: 자동 퍼지 매칭 ──
                        if not _bk:
                            for _k in product_tax_map:
                                _k_c = re.sub(r"\s+", "", _k).lower()
                                if len(_n_c) >= 2 and len(_k_c) >= 2:
                                    if _n_c in _k_c or _k_c in _n_c:
                                        _sc = 0.9
                                    else:
                                        _sc = _dl2.SequenceMatcher(None, _n_c, _k_c).ratio()
                                else:
                                    _sc = 0.0
                                if _sc > _bs:
                                    _bs, _bk = _sc, _k
                        if _bk and _bs >= 0.55:
                            if _bs < 1.0:
                                log(f"  퍼지 매칭: '{_n}' → '{_bk}' (score={_bs:.2f})")
                            _hit = product_tax_map[_bk]
                            _fuzzy_lookup_cache[_n] = _bk
                        else:
                            log(f"  ⚠ 캐시 미매칭: '{_n}'")
                            _hit = {}
                            _fuzzy_lookup_cache[_n] = None
                _tax_list.append(_hit)
            merged["_tax_list"] = _tax_list

            master_results.append(merged)

        # --- 주문상품 동적 열 전개: 주문상품1 / 주문상품1_수량 / 주문상품2 / ... ---
        _max_order_prods = max(
            (len(r.get("_product_list", [])) for r in master_results), default=0
        ) if master_results else 0
        _order_prod_col_names: list = []
        for _i in range(_max_order_prods):
            _order_prod_col_names += [f"주문상품{_i+1}", f"주문상품{_i+1}_수량"]
        for r in master_results:
            _plist = r.get("_product_list", [])
            for _i in range(_max_order_prods):
                if _i < len(_plist):
                    r[f"주문상품{_i+1}"]     = _plist[_i][0]
                    r[f"주문상품{_i+1}_수량"] = _plist[_i][1]
                else:
                    r[f"주문상품{_i+1}"]     = ""
                    r[f"주문상품{_i+1}_수량"] = ""

        # --- 면세/과세 열 동적 전개 ---
        # 수량반영 금액 군: 상품N_공급가액금액 / 부가세금액 / 면세금액금액
        # 단가 군:        상품N_공급가액 / 부가세 / 면세금액 / 과세구분
        _max_prods = max((len(r.get("_tax_list", [])) for r in master_results), default=0) if master_results else 0
        _qty_tax_col_names: list = []   # 수량 반영 금액 군 (주문링크 옆)
        _unit_tax_col_names: list = []  # 단가 군 (맨 끝)
        for _i in range(_max_prods):
            _pfx = f"상품{_i + 1}_"
            _qty_tax_col_names  += [f"{_pfx}공급가액금액", f"{_pfx}부가세금액", f"{_pfx}면세금액금액"]
            _unit_tax_col_names += [f"{_pfx}공급가액_단가", f"{_pfx}부가세_단가", f"{_pfx}면세금액_단가", f"{_pfx}과세구분"]

        for r in master_results:
            _tlist = r.get("_tax_list", [])
            # 주문 단위 합계: 공급가액합계(면세금액 포함, 부가세 제외) / 부가세합계
            _sum_supply = 0   # 공급가액 + 면세금액 (수량 반영)
            _sum_vat    = 0   # 부가세 (수량 반영)
            for _i in range(_max_prods):
                _pfx = f"상품{_i + 1}_"
                _ti  = _tlist[_i] if _i < len(_tlist) else {}
                _qty = _parse_qty_int(r.get(f"주문상품{_i+1}_수량", "")) if _i < len(_tlist) else 0

                _supply_unit = _ti.get("공급가액", 0) or 0
                _vat_unit    = _ti.get("부가세",   0) or 0
                _exempt_unit = _ti.get("면세금액", 0) or 0

                # 단가 컬럼
                r[f"{_pfx}공급가액_단가"] = _supply_unit if _ti else ""
                r[f"{_pfx}부가세_단가"]   = _vat_unit    if _ti else ""
                r[f"{_pfx}면세금액_단가"] = _exempt_unit if _ti else ""
                r[f"{_pfx}과세구분"]      = _ti.get("과세구분", "")

                # 수량 반영 금액 컬럼
                if _ti and _qty > 0:
                    _supply_amt = _supply_unit * _qty
                    _vat_amt    = _vat_unit    * _qty
                    _exempt_amt = _exempt_unit * _qty
                    r[f"{_pfx}공급가액금액"] = _supply_amt
                    r[f"{_pfx}부가세금액"]   = _vat_amt
                    r[f"{_pfx}면세금액금액"] = _exempt_amt
                    # 합계: 공급가액합계는 공급가액+면세금액(부가세 제외), 부가세합계는 부가세만
                    _sum_supply += _supply_amt + _exempt_amt
                    _sum_vat    += _vat_amt
                else:
                    r[f"{_pfx}공급가액금액"] = ""
                    r[f"{_pfx}부가세금액"]   = ""
                    r[f"{_pfx}면세금액금액"] = ""

            r["공급가액합계"] = _sum_supply if _sum_supply or _sum_vat else ""
            r["부가세합계"]   = _sum_vat    if _sum_supply or _sum_vat else ""

            # 카페 주문에서 결제금액이 미기재인 경우 계산된 합계로 보완
            if not r.get("결제금액_총액") and (_sum_supply or _sum_vat):
                _computed_total = _sum_supply + _sum_vat
                r["결제금액_총액"] = _computed_total
                if not r.get("처방비용"):
                    r["처방비용"] = _computed_total

        # 열 순서:
        #   주문번호~주문상품 → 결제금액군 → 주문상품N/수량 → 입금~주문링크 → 상품N_금액/단가
        _payment_cols = ["결제금액_총액", "처방비용", "배송료", "포인트",
                         "공급가액합계", "부가세합계", "결제방법"]
        _payment_set  = set(_payment_cols)
        _op_idx = (_SHOP_ORDER_EXCEL_COLS.index("주문상품")
                   if "주문상품" in _SHOP_ORDER_EXCEL_COLS else len(_SHOP_ORDER_EXCEL_COLS) - 1)
        _head_cols = [c for c in _SHOP_ORDER_EXCEL_COLS[:_op_idx + 1]
                      if c not in _payment_set]          # 주문번호 ~ 주문상품
        _tail_cols = [c for c in _SHOP_ORDER_EXCEL_COLS[_op_idx + 1:]
                      if c not in _payment_set]          # 입금 ~ 주문링크
        _excel_all_cols = (
            _head_cols
            + _payment_cols              # 결제금액_총액 ~ 결제방법 (주문상품 바로 뒤)
            + _order_prod_col_names      # 주문상품1/수량, 주문상품2/수량, ...
            + _tail_cols                 # 입금, 입금자명, ..., 주문링크
            + _qty_tax_col_names         # 상품1_공급가액금액, 상품1_부가세금액, ... (맨 뒤)
            + _unit_tax_col_names        # 상품1_공급가액_단가, ...
        )

        # --- 저장 ---
        update_progress(92, "파일 저장 중...")
        ensure_dir(output_root)

        if save_excel:
            # 없는 컬럼은 빈 문자열로 채우기
            for r in master_results:
                for c in _excel_all_cols:
                    if c not in r:
                        r[c] = ""

            # 순번 열 맨 앞에 추가
            _all_cols_with_seq = ["순번"] + _excel_all_cols
            for _si, r in enumerate(master_results, 1):
                r["순번"] = _si

            df_excel = pd.DataFrame(master_results, columns=_all_cols_with_seq)

            # 수량 "개" 제거 후 숫자 변환
            _qty_cols = [c for c in df_excel.columns if c.endswith("_수량")]
            for _qc in _qty_cols:
                df_excel[_qc] = (
                    df_excel[_qc].astype(str)
                    .str.replace(r"개$", "", regex=True)
                    .str.strip()
                    .replace("", pd.NA)
                    .pipe(pd.to_numeric, errors="coerce")
                    .astype("Int64")
                )

            # 금액 컬럼 숫자 변환 (텍스트→숫자, 엑셀 초록삼각형 방지)
            _numeric_str_cols = ["결제금액_총액", "처방비용", "배송료", "포인트",
                                 "공급가액합계", "부가세합계"]
            _numeric_val_cols = [c for c in df_excel.columns
                                 if any(c.endswith(sfx) for sfx in
                                        ("공급가액금액", "부가세금액", "면세금액금액",
                                         "공급가액_단가", "부가세_단가", "면세금액_단가"))]
            for _nc in _numeric_str_cols + _numeric_val_cols:
                if _nc in df_excel.columns:
                    df_excel[_nc] = pd.to_numeric(df_excel[_nc], errors="coerce").astype("Int64")

            excel_path = os.path.join(output_root, f"{run_ts}_약속처방_주문마스터.xlsx")
            df_excel.to_excel(excel_path, index=False)
            log(f"엑셀 저장: {excel_path} ({_max_prods}개 상품 열)")

        if save_cj:
            cj_df = build_shop_order_cj_df(master_results)
            cj_path = os.path.join(output_root, f"{run_ts}_약속처방_대한통운_파일업로드.xlsx")
            cj_df.to_excel(cj_path, index=False)
            log(f"CJ 파일 저장: {cj_path}")

        update_progress(100, f"완료! {total}건 처리 → {output_root}")
        return output_root

    except Exception as e:
        log(f"오류: {e}")
        import traceback
        log(traceback.format_exc())
        update_progress(100, f"오류 발생: {e}")
        return None
    finally:
        # 실행 로그 파일 저장
        try:
            _log_content = "\n".join(_sp_log_lines)
            _root = _sp_output_root[0]
            if _log_content.strip() and _root:
                ensure_dir(_root)
                _log_path = os.path.join(_root, "실행로그.txt")
                with open(_log_path, "w", encoding="utf-8") as _lf:
                    _lf.write(_log_content)
                print(f"실행 로그 저장: {_log_path}")
        except Exception:
            pass
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


# ---------- GUI ----------
def launch_gui():
    root = tk.Tk()
    root.title(f"케이진 탕전주문 관리 v{APP_VERSION}")
    root.geometry("540x840")
    root.resizable(False, True)   # 세로 리사이즈 허용 (맥에서 하단 잘림 대응)

    # 진행바 완료/오류 색상 (네이티브 테마 유지, Windows vista 테마에서 background 적용됨)
    _style = ttk.Style()
    _style.configure("Done.Horizontal.TProgressbar",  background="#2E7D32")
    _style.configure("Error.Horizontal.TProgressbar", background="#B71C1C")

    # ── Placeholder 헬퍼 ──
    _PH_COLOR = "gray"
    _ph_map: dict = {}   # entry → placeholder 텍스트

    def _ph_add(entry, text):
        """entry에 회색 placeholder 부착. 포커스 인/아웃 시 자동 처리."""
        _ph_map[entry] = text
        entry.insert(0, text)
        entry.config(fg=_PH_COLOR)

        def _in(e):
            if entry.get() == _ph_map.get(entry, ""):
                entry.delete(0, tk.END)
                entry.config(fg="black")

        def _out(e):
            if not entry.get().strip():
                entry.insert(0, _ph_map[entry])
                entry.config(fg=_PH_COLOR)

        entry.bind("<FocusIn>",  _in)
        entry.bind("<FocusOut>", _out)

    def _ph_get(entry):
        """placeholder 상태면 빈 문자열 반환, 아니면 실제 값 반환."""
        v = entry.get().strip()
        return "" if v == _ph_map.get(entry, object()) else v

    def _ph_set_date(entry, value):
        """set_today 등에서 날짜값 강제 세팅 (placeholder 덮어쓰기)."""
        entry.delete(0, tk.END)
        entry.insert(0, value)
        entry.config(fg="black")

    # 앱 아이콘 (32×32, 녹색 계열 동심 사각형)
    try:
        _icon_img = tk.PhotoImage(width=32, height=32)
        _icon_rows = []
        for _y in range(32):
            _row = []
            for _x in range(32):
                if _x < 3 or _x > 28 or _y < 3 or _y > 28:
                    _row.append("#1a5c2a")
                elif _x < 7 or _x > 24 or _y < 7 or _y > 24:
                    _row.append("#2e8b47")
                else:
                    _row.append("#52c068")
            _icon_rows.append("{" + " ".join(_row) + "}")
        _icon_img.put(" ".join(_icon_rows))
        root.iconphoto(True, _icon_img)
    except Exception:
        pass

    # 전역 실행 잠금 — 탭 하나가 돌아가는 동안 다른 탭 실행 방지
    _global_running = threading.Event()

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=8, pady=8)

    # ===== 탭 1: 탕전주문 조회 =====
    tab1 = ttk.Frame(notebook, padding=12)
    notebook.add(tab1, text="탕전주문 조회")

    # ── 공통 Entry 스타일 ──
    _ENTRY_KW = dict(
        relief="flat", bd=0,
        highlightthickness=1,
        highlightbackground="#7A7A7A",
        highlightcolor="#0078D4",
    )

    # ── 진행상태 (LabelFrame) ──
    status_lf = ttk.LabelFrame(tab1, text="진행상태", padding=(8, 4))
    status_lf.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 4))
    # 체크박스 4열 균등 분배
    for _c in range(4):
        status_lf.columnconfigure(_c, weight=1, uniform="scb")

    status_vars = {}
    for i, status in enumerate(ALL_STATUSES):
        var = tk.BooleanVar(value=(status == "접수대기"))
        status_vars[status] = var
        ttk.Checkbutton(status_lf, text=status, variable=var).grid(
            row=i // 4, column=i % 4, sticky="w", padx=(0, 4), pady=2
        )

    def select_all_status():
        for var in status_vars.values():
            var.set(True)

    def clear_all_status():
        for var in status_vars.values():
            var.set(False)

    # 버튼 — 오른쪽 끝 정렬 (weight=0으로 고정폭 유지)
    status_lf.columnconfigure(4, weight=0)
    status_btn_frame = ttk.Frame(status_lf)
    status_btn_frame.grid(row=0, column=4, rowspan=2, sticky="e")
    status_btn_frame.columnconfigure(0, weight=1)
    ttk.Button(status_btn_frame, text="전체선택", command=select_all_status, width=7).grid(row=0, column=0, sticky="ew", pady=(0, 3))
    ttk.Button(status_btn_frame, text="전체해제", command=clear_all_status, width=7).grid(row=1, column=0, sticky="ew")

    # ── 조회 기간 (LabelFrame) ──
    date_lf = ttk.LabelFrame(tab1, text="조회 기간", padding=(8, 4))
    date_lf.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 4))
    date_lf.columnconfigure(1, weight=1)   # 날짜칸 늘어남
    date_lf.columnconfigure(4, weight=0)   # 버튼 프레임 고정

    # 시작/종료 레이블 + 날짜 + 시간 (placeholder로 구분)
    ttk.Label(date_lf, text="시작").grid(row=0, column=0, sticky="w", padx=(0, 6))
    start_date_entry = tk.Entry(date_lf, width=13, justify="center", **_ENTRY_KW)
    start_date_entry.grid(row=0, column=1, sticky="ew", ipady=2)
    start_time_entry = tk.Entry(date_lf, width=9, justify="center", **_ENTRY_KW)
    start_time_entry.grid(row=0, column=2, padx=(6, 0), ipady=2)

    ttk.Label(date_lf, text="종료").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(5, 0))
    end_date_entry = tk.Entry(date_lf, width=13, justify="center", **_ENTRY_KW)
    end_date_entry.grid(row=1, column=1, sticky="ew", pady=(5, 0), ipady=2)
    end_time_entry = tk.Entry(date_lf, width=9, justify="center", **_ENTRY_KW)
    end_time_entry.grid(row=1, column=2, padx=(6, 0), pady=(5, 0), ipady=2)

    # placeholder 부착
    _ph_add(start_date_entry, "YYYY-MM-DD")
    _ph_add(start_time_entry, "시간 (선택)")
    _ph_add(end_date_entry,   "YYYY-MM-DD")
    _ph_add(end_time_entry,   "시간 (선택)")

    def set_today():
        today = datetime.today().strftime("%Y-%m-%d")
        _ph_set_date(start_date_entry, today); _ph_set_date(end_date_entry, today)

    def set_yesterday():
        day = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        _ph_set_date(start_date_entry, day); _ph_set_date(end_date_entry, day)

    def set_this_month():
        today = datetime.today()
        first = today.replace(day=1).strftime("%Y-%m-%d")
        if today.month == 12:
            last = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        _ph_set_date(start_date_entry, first)
        _ph_set_date(end_date_entry,   last.strftime("%Y-%m-%d"))

    def set_last_month():
        today = datetime.today()
        last = today.replace(day=1) - timedelta(days=1)
        first = last.replace(day=1)
        _ph_set_date(start_date_entry, first.strftime("%Y-%m-%d"))
        _ph_set_date(end_date_entry,   last.strftime("%Y-%m-%d"))

    def set_this_year():
        today = datetime.today()
        _ph_set_date(start_date_entry, f"{today.year}-01-01")
        _ph_set_date(end_date_entry,   f"{today.year}-12-31")

    def set_last_year():
        y = datetime.today().year - 1
        _ph_set_date(start_date_entry, f"{y}-01-01")
        _ph_set_date(end_date_entry,   f"{y}-12-31")

    def clear_dates():
        for e in (start_date_entry, start_time_entry, end_date_entry, end_time_entry):
            e.delete(0, tk.END)
            e.insert(0, _ph_map[e])
            e.config(fg=_PH_COLOR)

    # 바로가기 버튼 — 2행 4열 + 초기화 rowspan=2
    date_btn_frame = ttk.Frame(date_lf)
    date_btn_frame.grid(row=0, column=4, rowspan=2, sticky="nsew", padx=(10, 0))
    for _c in range(4):
        date_btn_frame.columnconfigure(_c, weight=1, uniform="db")

    _bw = 7  # 버튼 width
    _px = (0, 2)  # 열 사이 간격
    ttk.Button(date_btn_frame, text="오늘",   command=set_today,      width=_bw).grid(row=0, column=0, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(date_btn_frame, text="이번달", command=set_this_month, width=_bw).grid(row=0, column=1, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(date_btn_frame, text="올해",   command=set_this_year,  width=_bw).grid(row=0, column=2, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(date_btn_frame, text="어제",   command=set_yesterday,  width=_bw).grid(row=1, column=0, sticky="ew", padx=_px)
    ttk.Button(date_btn_frame, text="저번달", command=set_last_month, width=_bw).grid(row=1, column=1, sticky="ew", padx=_px)
    ttk.Button(date_btn_frame, text="작년",   command=set_last_year,  width=_bw).grid(row=1, column=2, sticky="ew", padx=_px)
    ttk.Button(date_btn_frame, text="초기화", command=clear_dates,    width=_bw).grid(row=0, column=3, rowspan=2, sticky="nsew")

    # ── 필터 (LabelFrame) ──
    hospital_filter_var = tk.StringVar(value="")
    search_target_var   = tk.StringVar(value="주문자명")
    search_filter_var   = tk.StringVar(value="")
    filter_lf = ttk.LabelFrame(tab1, text="필터", padding=(8, 4))
    filter_lf.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 4))
    filter_lf.columnconfigure(1, weight=1)

    ttk.Label(filter_lf, text="한의원").grid(row=0, column=0, sticky="w", padx=(0, 8))
    hospital_combo = ttk.Combobox(filter_lf, textvariable=hospital_filter_var,
                                  values=HOSPITAL_PRESETS, state="normal")
    hospital_combo.grid(row=0, column=1, sticky="ew")
    ttk.Button(filter_lf, text="지우기",
               command=lambda: hospital_filter_var.set("")).grid(row=0, column=2, padx=(4, 0))

    ttk.Label(filter_lf, text="검색대상").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(6, 0))
    search_row = ttk.Frame(filter_lf)
    search_row.grid(row=1, column=1, sticky="ew", pady=(6, 0))
    search_row.columnconfigure(1, weight=1)
    search_target_combo = ttk.Combobox(
        search_row, textvariable=search_target_var,
        values=["주문자명", "환자명(복용자)", "주문자명+환자명"],
        state="readonly", width=16
    )
    search_target_combo.grid(row=0, column=0, sticky="w", padx=(0, 6))
    search_entry = tk.Entry(search_row, textvariable=search_filter_var, **_ENTRY_KW)
    search_entry.grid(row=0, column=1, sticky="ew", ipady=2)
    ttk.Button(filter_lf, text="지우기",
               command=lambda: search_filter_var.set("")).grid(row=1, column=2, padx=(4, 0), pady=(6, 0))

    ttk.Label(filter_lf, text="결제방법").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=(6, 0))
    payment_filter_var = tk.StringVar(value="")
    payment_combo = ttk.Combobox(
        filter_lf, textvariable=payment_filter_var,
        values=["", "카드", "무통장", "실시간계좌이체", "가상계좌", "네이버페이", "카카오페이"],
        state="normal", width=20
    )
    payment_combo.grid(row=2, column=1, sticky="w", pady=(6, 0))
    ttk.Button(filter_lf, text="지우기",
               command=lambda: payment_filter_var.set("")).grid(row=2, column=2, padx=(4, 0), pady=(6, 0))

    ttk.Label(filter_lf, text="목록 선택 or 직접 입력 · 검색어는 콤마로 여러 개 입력 가능 · 비우면 전체",
              foreground="gray", font=("Malgun Gothic", 8)).grid(
        row=3, column=0, columnspan=3, sticky="w", pady=(3, 0)
    )

    save_excel_var = tk.BooleanVar(value=True)
    save_excel_exclude_inpatient_var = tk.BooleanVar(value=False)
    save_bulk_excel_var = tk.BooleanVar(value=False)
    sort_oldest_first_var = tk.BooleanVar(value=False)
    save_cj_excel_var = tk.BooleanVar(value=True)
    save_label_excel_var = tk.BooleanVar(value=True)
    save_decoction_pdf_var = tk.BooleanVar(value=True)
    save_dispense_pdf_var = tk.BooleanVar(value=False)
    save_dosage_text_pdf_var = tk.BooleanVar(value=False)
    save_dosage_attachment_var = tk.BooleanVar(value=False)
    print_decoction_var = tk.BooleanVar(value=False)
    print_dispense_var = tk.BooleanVar(value=False)
    print_by_hospital_var = tk.BooleanVar(value=False)
    print_inpatient_last_var = tk.BooleanVar(value=False)
    print_printer_var = tk.StringVar(value="")
    auto_change_status_var = tk.BooleanVar(value=True)
    auto_cancel_status_var = tk.BooleanVar(value=True)
    auto_dispensing_var = tk.BooleanVar(value=True)

    # ── 2단 레이아웃: col0=파일저장/자동전환, col1=자동인쇄/PDF저장
    tab1.columnconfigure(0, weight=0)

    # ── 그룹 1: 파일 저장 (좌측 상단)
    group1 = ttk.LabelFrame(tab1, text="파일 저장", padding=(8, 4))
    group1.grid(row=3, column=0, sticky="nsew", pady=(4, 2), padx=(0, 4))
    ttk.Checkbutton(group1, text="주문 엑셀 저장 (통합 시트)", variable=save_excel_var).grid(row=0, column=0, sticky="w")
    ttk.Checkbutton(group1, text="오래된 순서로 정렬", variable=sort_oldest_first_var).grid(row=1, column=0, sticky="w", padx=(16, 0))
    cb_save_excel_exclude_inpatient = ttk.Checkbutton(
        group1, text="입원 케이스 제외 후 저장", variable=save_excel_exclude_inpatient_var
    )
    cb_save_excel_exclude_inpatient.grid(row=2, column=0, sticky="w", padx=(16, 0))
    cb_save_bulk_excel = ttk.Checkbutton(
        group1, text="벌크 내역만 추출 저장", variable=save_bulk_excel_var
    )
    cb_save_bulk_excel.grid(row=3, column=0, sticky="w", padx=(16, 0))
    ttk.Checkbutton(group1, text="대한통운 파일 업로드 양식 생성", variable=save_cj_excel_var).grid(row=4, column=0, sticky="w")
    ttk.Checkbutton(group1, text="라벨 인쇄용 엑셀 생성", variable=save_label_excel_var).grid(row=5, column=0, sticky="w")

    def _update_save_excel_sub(*_):
        state = "normal" if save_excel_var.get() else "disabled"
        cb_save_excel_exclude_inpatient.config(state=state)
        cb_save_bulk_excel.config(state=state)
        if state == "disabled":
            save_excel_exclude_inpatient_var.set(False)
            save_bulk_excel_var.set(False)

    save_excel_var.trace_add("write", _update_save_excel_sub)
    _update_save_excel_sub()

    # ── 그룹 auto: 자동 전환 (좌측 하단, 기존 PDF 저장 자리)
    group_auto = ttk.LabelFrame(tab1, text="자동 전환", padding=(8, 4))
    group_auto.grid(row=4, column=0, sticky="nsew", pady=(2, 2), padx=(0, 4))
    ttk.Checkbutton(group_auto, text="입원건 입금대기 자동 전환", variable=auto_change_status_var).grid(row=0, column=0, sticky="w")
    ttk.Checkbutton(group_auto, text="취소요청건 환불취소 자동 전환", variable=auto_cancel_status_var).grid(row=1, column=0, sticky="w")
    ttk.Checkbutton(group_auto, text="정상 주문 조제중 자동 전환", variable=auto_dispensing_var).grid(row=2, column=0, sticky="w")

    # ── 그룹 3: 자동 인쇄 (우측 상단)
    group3 = ttk.LabelFrame(tab1, text="자동 인쇄", padding=(8, 4))
    group3.grid(row=3, column=1, rowspan=1, sticky="nsew", pady=(4, 2), padx=(4, 0))
    group3.columnconfigure(1, weight=1)
    ttk.Checkbutton(group3, text="탕전주문내역서 자동 인쇄", variable=print_decoction_var).grid(row=0, column=0, columnspan=3, sticky="w")
    ttk.Checkbutton(group3, text="조제지시서 자동 인쇄", variable=print_dispense_var).grid(row=1, column=0, columnspan=3, sticky="w")
    ttk.Separator(group3, orient="horizontal").grid(row=2, column=0, columnspan=3, sticky="ew", pady=(6, 4))
    ttk.Label(group3, text="인쇄 순서", foreground="#555", font=("Malgun Gothic", 9, "bold")).grid(
        row=3, column=0, columnspan=3, sticky="w", pady=(0, 2)
    )
    cb_by_hospital = ttk.Checkbutton(group3, text="한의원별로 묶어서 출력",
                                      variable=print_by_hospital_var)
    cb_by_hospital.grid(row=4, column=0, columnspan=3, sticky="w")
    cb_inpatient_last = ttk.Checkbutton(group3, text="입원 건 맨 마지막에 별도 출력",
                                         variable=print_inpatient_last_var)
    cb_inpatient_last.grid(row=5, column=0, columnspan=3, sticky="w")

    # 프린터 선택 행 (자동 인쇄 하위)
    ttk.Separator(group3, orient="horizontal").grid(row=6, column=0, columnspan=3, sticky="ew", pady=(6, 4))
    lbl_printer = ttk.Label(group3, text="출력 프린터:")
    lbl_printer.grid(row=7, column=0, sticky="w", padx=(0, 4))
    printer_combo = ttk.Combobox(group3, textvariable=print_printer_var,
                                  state="disabled", width=18)
    printer_combo.grid(row=7, column=1, sticky="ew")

    def _refresh_printers():
        names = get_printer_list()
        printer_combo["values"] = names
        if names and not print_printer_var.get():
            print_printer_var.set(names[0])

    btn_refresh_printer = ttk.Button(group3, text="↻", width=3,
                                      command=_refresh_printers)
    btn_refresh_printer.grid(row=7, column=2, padx=(4, 0))
    btn_refresh_printer.config(state="disabled")

    def _update_print_sub(*_):
        state = "normal" if (print_decoction_var.get() or print_dispense_var.get()) else "disabled"
        cb_by_hospital.config(state=state)
        cb_inpatient_last.config(state=state)
        lbl_printer.config(state=state)
        printer_combo.config(state="readonly" if state == "normal" else "disabled")
        btn_refresh_printer.config(state=state)
        if state == "normal" and not printer_combo["values"]:
            _refresh_printers()

    print_decoction_var.trace_add("write", _update_print_sub)
    print_dispense_var.trace_add("write", _update_print_sub)
    _update_print_sub()  # 초기 상태 반영

    # ── 그룹 2: PDF 저장 (우측 하단 — 기존 빈공간)
    group2 = ttk.LabelFrame(tab1, text="PDF 저장", padding=(8, 4))
    group2.grid(row=4, column=1, sticky="nsew", pady=(2, 2), padx=(4, 0))
    ttk.Checkbutton(group2, text="주문내역서 PDF 저장", variable=save_decoction_pdf_var).grid(row=0, column=0, sticky="w")
    ttk.Checkbutton(group2, text="조제지시서 PDF 저장", variable=save_dispense_pdf_var).grid(row=1, column=0, sticky="w")
    ttk.Checkbutton(group2, text="복용법 PDF 저장", variable=save_dosage_text_pdf_var).grid(row=2, column=0, sticky="w")
    ttk.Checkbutton(group2, text="복용첨부파일 다운로드", variable=save_dosage_attachment_var).grid(row=3, column=0, sticky="w")

    test_limit_var = tk.StringVar(value="")
    all_period_max_page_var = tk.StringVar(value=str(DEFAULT_ALL_PERIOD_MAX_PAGE))
    save_decoction_sheet_var = tk.BooleanVar(value=False)

    def open_test_options():
        top = tk.Toplevel(root)
        top.title("테스트 옵션")
        top.geometry("330x210")
        top.resizable(False, False)
        inner = ttk.Frame(top, padding=12)
        inner.pack(fill="both", expand=True)
        ttk.Label(inner, text="테스트 건수 제한").grid(row=0, column=0, sticky="w", pady=4)
        entry1 = ttk.Entry(inner, width=18, textvariable=test_limit_var)
        entry1.grid(row=0, column=1, sticky="w", pady=4)
        ttk.Label(inner, text="전체조회 시 최대 페이지 수").grid(row=1, column=0, sticky="w", pady=4)
        entry2 = ttk.Entry(inner, width=18, textvariable=all_period_max_page_var)
        entry2.grid(row=1, column=1, sticky="w", pady=4)
        ttk.Label(inner, text="테스트 건수는 비우면 전체").grid(row=2, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Label(inner, text="페이지 수 기본값은 7").grid(row=3, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Separator(inner, orient="horizontal").grid(row=4, column=0, columnspan=2, sticky="ew", pady=(8, 4))
        ttk.Checkbutton(
            inner, text="탕전주문내역 시트 별도 저장", variable=save_decoction_sheet_var
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=2)

        def save_and_close():
            v1 = entry1.get().strip()
            v2 = entry2.get().strip()
            if v1:
                try:
                    int(v1)
                except ValueError:
                    messagebox.showerror("입력 오류", "테스트 건수 제한은 숫자만 입력해줘.")
                    return
            if v2:
                try:
                    page_val = int(v2)
                    if page_val <= 0:
                        raise ValueError
                except ValueError:
                    messagebox.showerror("입력 오류", "전체조회 최대 페이지 수는 1 이상의 숫자여야 해.")
                    return
            else:
                all_period_max_page_var.set(str(DEFAULT_ALL_PERIOD_MAX_PAGE))
            top.destroy()

        ttk.Button(inner, text="적용", command=save_and_close).grid(row=6, column=0, sticky="ew", pady=(10, 0))
        ttk.Button(inner, text="닫기", command=top.destroy).grid(row=6, column=1, sticky="ew", pady=(10, 0))

    _all_group_vars = [
        save_excel_var, save_excel_exclude_inpatient_var, save_bulk_excel_var, sort_oldest_first_var,
        save_cj_excel_var, save_label_excel_var,
        save_decoction_pdf_var, save_dispense_pdf_var, save_dosage_text_pdf_var, save_dosage_attachment_var,
        print_decoction_var, print_dispense_var, auto_change_status_var, auto_cancel_status_var,
        auto_dispensing_var,
    ]

    def select_all_groups():
        for v in _all_group_vars:
            v.set(True)

    def clear_all_groups():
        for v in _all_group_vars:
            v.set(False)

    test_row_frame = ttk.Frame(tab1)
    # columnspan=2 (좌/우 컬럼 전체 폭) + sticky="ew" → 오른쪽 버튼이 자동 인쇄 박스 오른쪽 끝과 정렬
    test_row_frame.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    ttk.Button(test_row_frame, text="전체선택", command=select_all_groups).pack(side="left", padx=(0, 4))
    ttk.Button(test_row_frame, text="전체해제", command=clear_all_groups).pack(side="left")
    ttk.Button(test_row_frame, text="테스트 옵션", command=open_test_options).pack(side="right")

    # ---------- 저장 위치 선택 ----------
    _cfg = load_app_config()
    output_dir_var = tk.StringVar(value=_cfg.get("output_dir", ""))
    output_dir_frame = ttk.Frame(tab1)
    output_dir_frame.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    output_dir_frame.columnconfigure(1, weight=1)
    ttk.Label(output_dir_frame, text="저장 위치:").grid(row=0, column=0, sticky="w", padx=(0, 6))
    ttk.Entry(output_dir_frame, textvariable=output_dir_var).grid(row=0, column=1, sticky="ew", padx=(0, 4))

    def _browse_output_dir():
        d = filedialog.askdirectory(title="결과 폴더를 저장할 위치 선택")
        if d:
            output_dir_var.set(d)
            save_app_config({**load_app_config(), "output_dir": d})

    ttk.Button(output_dir_frame, text="찾아보기", command=_browse_output_dir, width=8).grid(row=0, column=2, sticky="e")

    status_label = ttk.Label(tab1, text="대기 중", foreground="gray")
    status_label.grid(row=7, column=0, columnspan=3, sticky="w", pady=(12, 4))

    progress_var = tk.IntVar(value=0)
    progress_bar = ttk.Progressbar(
        tab1, orient="horizontal", length=300,
        mode="determinate", maximum=100, variable=progress_var
    )
    progress_bar.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(0, 8))

    cancel_event = threading.Event()

    btn_frame1 = ttk.Frame(tab1)
    # columnspan=2 (좌/우 컬럼 전체 폭) + sticky="ew" → 오른쪽 버튼이 자동 인쇄 박스 오른쪽 끝과 정렬
    btn_frame1.grid(row=9, column=0, columnspan=2, sticky="ew", pady=(4, 0))

    run_button = ttk.Button(btn_frame1, text="실행")
    run_button.pack(side="left", padx=(0, 4))

    cancel_button = ttk.Button(btn_frame1, text="취소", state="disabled",
                               command=lambda: cancel_event.set())
    cancel_button.pack(side="left", padx=(0, 4))

    close_button = ttk.Button(btn_frame1, text="종료", command=root.destroy)
    close_button.pack(side="left", padx=(0, 4))

    # 오른쪽 정렬: 출력 폴더 열기(맨 오른쪽) + 결과 다시 보기(그 왼쪽)
    open_folder_btn = ttk.Button(btn_frame1, text="출력 폴더 열기", state="disabled")
    open_folder_btn.pack(side="right", padx=(4, 0))

    # 마지막 작업 결과 요약 저장 (다시 보기용)
    last_result_msg = {"text": ""}
    show_result_btn = ttk.Button(btn_frame1, text="결과 다시 보기", state="disabled")
    show_result_btn.pack(side="right", padx=(4, 0))

    def gui_progress(percent: int, message: str):
        root.after(0, lambda p=percent: progress_var.set(p))
        if percent == 100:
            fg, bar_style = "#2E7D32", "Done.Horizontal.TProgressbar"
        elif percent == 0:
            fg, bar_style = "#B71C1C", "Error.Horizontal.TProgressbar"
        else:
            fg, bar_style = "#E65100", None
        root.after(0, lambda m=message, c=fg: status_label.config(text=m, foreground=c))
        if bar_style:
            root.after(0, lambda s=bar_style: progress_bar.config(style=s))

    def on_run():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭이 현재 실행 중입니다.\n완료 후 다시 시도해주세요.")
            return
        if not ensure_admin_credentials(root):
            return
        start_date = ""
        end_date = ""
        start_raw = " ".join(filter(None, [_ph_get(start_date_entry), _ph_get(start_time_entry)]))
        end_raw = " ".join(filter(None, [_ph_get(end_date_entry), _ph_get(end_time_entry)]))
        try:
            start_date = normalize_date_input(start_raw)
            end_date = normalize_date_input(end_raw)
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return

        today_str = datetime.today().strftime("%Y-%m-%d")
        is_all_period = False

        if start_date and not end_date:
            end_date = today_str
        elif end_date and not start_date:
            messagebox.showwarning("입력 확인", "종료일만 입력했습니다. 시작일도 입력해줘.")
            return
        elif not start_date and not end_date:
            go_all = messagebox.askyesno(
                "전체조회 확인",
                "시작일과 종료일이 비어 있습니다. 전체기간으로 조회합니다. 계속할까요?"
            )
            if not go_all:
                return
            is_all_period = True

        selected_statuses = [status for status, var in status_vars.items() if var.get()]
        if not selected_statuses:
            go_all_status = messagebox.askyesno(
                "진행상태 확인",
                "진행상태를 선택하지 않았습니다. 전체 상태로 조회합니다. 계속할까요?"
            )
            if not go_all_status:
                return
            selected_statuses = ALL_STATUSES[:]

        limit_raw = test_limit_var.get().strip()
        test_limit = None
        if limit_raw:
            try:
                test_limit = int(limit_raw)
            except ValueError:
                messagebox.showerror("입력 오류", "테스트 건수 제한은 숫자만 입력해줘.")
                return

        max_page_raw = all_period_max_page_var.get().strip()
        try:
            all_period_max_page = int(max_page_raw) if max_page_raw else DEFAULT_ALL_PERIOD_MAX_PAGE
            if all_period_max_page <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("입력 오류", "전체조회 최대 페이지 수는 1 이상의 숫자여야 해.")
            return

        max_page_limit = all_period_max_page if is_all_period else MAX_PAGE_SAFETY_LIMIT

        if save_label_excel_var.get() and not save_excel_var.get():
            save_excel_var.set(True)
            messagebox.showinfo("옵션 조정", "라벨 인쇄용 엑셀 생성을 위해 주문 엑셀 저장을 함께 켰어요.")

        settings = {
            "start_date": start_date,
            "end_date": end_date,
            "test_limit": test_limit,
            "target_statuses": selected_statuses,
            "max_page_limit": max_page_limit,
            "save_excel": save_excel_var.get(),
            "save_excel_exclude_inpatient": save_excel_exclude_inpatient_var.get(),
            "save_bulk_excel": save_bulk_excel_var.get(),
            "sort_oldest_first": sort_oldest_first_var.get(),
            "save_cj_excel": save_cj_excel_var.get(),
            "save_label_excel": save_label_excel_var.get(),
            "save_decoction_pdf": save_decoction_pdf_var.get(),
            "save_dispense_pdf": save_dispense_pdf_var.get(),
            "save_dosage_text_pdf": save_dosage_text_pdf_var.get(),
            "save_dosage_attachment": save_dosage_attachment_var.get(),
            "print_decoction": print_decoction_var.get(),
            "print_dispense": print_dispense_var.get(),
            "print_printer_name": print_printer_var.get(),
            "print_by_hospital": print_by_hospital_var.get(),
            "print_inpatient_last": print_inpatient_last_var.get(),
            "auto_change_status": auto_change_status_var.get(),
            "auto_cancel_status": auto_cancel_status_var.get(),
            "auto_dispensing_status": auto_dispensing_var.get(),
            "save_decoction_sheet": save_decoction_sheet_var.get(),
            "hospital_filter": hospital_filter_var.get().strip(),
            "payment_filter": payment_filter_var.get().strip(),
            "search_target": search_target_var.get(),
            "search_filter": search_filter_var.get().strip(),
            "cancel_event": cancel_event,
            "output_dir": output_dir_var.get().strip(),
        }

        cancel_event.clear()
        _global_running.set()
        run_button.config(state="disabled")
        cancel_button.config(state="normal")
        close_button.config(state="disabled")
        open_folder_btn.config(state="disabled", command=None, text="출력 폴더 열기")
        show_result_btn.config(state="disabled", command=None)
        progress_var.set(0)
        progress_bar.config(style="Horizontal.TProgressbar")
        status_label.config(text="실행 준비 중...", foreground="#E65100")

        def worker():
            try:
                result = run_job(settings, progress_callback=gui_progress)
                master_results = result["master_results"] if result else []
                auto_change_count = result["auto_change_count"] if result else 0
                cancel_change_count = result["cancel_change_count"] if result else 0
                dispensing_change_count = result.get("dispensing_change_count", 0) if result else 0

                if cancel_event.is_set():
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소되었습니다."))
                else:
                    total_cnt = len(master_results)
                    root.after(0, lambda: progress_var.set(100))
                    _out = result.get("output_root", "") if result else ""
                    no_data = bool(result and result.get("no_data"))
                    _folder_name = os.path.basename(_out) if _out else ""
                    _status_text = "조회 완료 - 주문 없음" if no_data else f"완료 ✓  {total_cnt}건 처리"
                    root.after(0, lambda t=_status_text: status_label.config(
                        text=t, foreground="#2E7D32"
                    ))
                    status_summary = "\n".join([
                        f"{s}: {sum(1 for r in master_results if r.get('진행상태') == s)}건"
                        for s in set(r.get('진행상태', '') for r in master_results)
                        if s
                    ])
                    from collections import Counter
                    hosp_counts = Counter(
                        r.get('_hospital_folder') or clean_text(r.get('한의원명', ''))
                        for r in master_results
                        if r.get('_hospital_folder') or clean_text(r.get('한의원명', ''))
                    )
                    hosp_summary = "\n".join(
                        f"  {h}: {c}건"
                        for h, c in sorted(hosp_counts.items(), key=lambda x: -x[1])
                    )
                    if no_data:
                        msg = (
                            "조회는 정상적으로 완료됐지만 조건에 맞는 주문이 없어요.\n\n"
                            "조회 기간, 진행상태, 한의원/이름 필터를 한 번 확인해주세요."
                        )
                        _result_title = "조회 결과 없음"
                    else:
                        msg = (
                            f"작업 완료.  총 {total_cnt}건\n\n"
                            f"{status_summary}\n\n"
                            f"[한의원별]\n{hosp_summary}\n\n"
                            f"입금대기 자동전환: {auto_change_count}건\n"
                            f"환불취소 자동전환: {cancel_change_count}건\n"
                            f"조제중 자동전환: {dispensing_change_count}건"
                        )
                        _result_title = "완료"
                    root.after(0, lambda m=msg, t=_result_title: messagebox.showinfo(t, m))
                    # 결과 다시 보기 버튼 활성화
                    last_result_msg["text"] = msg
                    def _show_result(m=msg, t=_result_title):
                        messagebox.showinfo(t, m)
                    root.after(0, lambda f=_show_result: show_result_btn.config(
                        state="normal", command=f
                    ))
                    # 폴더 열기 버튼 — 폴더명을 버튼 텍스트에 표시
                    if _out and os.path.exists(_out):
                        def _open_folder(p=_out):
                            try:
                                if sys.platform == "win32":
                                    os.startfile(p)
                                else:
                                    subprocess.run(["open", p])
                            except Exception:
                                pass
                        _btn_label = f"📂 {_folder_name}" if _folder_name else "출력 폴더 열기"
                        root.after(0, lambda f=_open_folder, t=_btn_label: open_folder_btn.config(
                            state="normal", command=f, text=t
                        ))
            except Exception as e:
                err_msg = str(e)
                if is_user_cancel_message(err_msg):
                    root.after(0, lambda: status_label.config(text="취소됨", foreground="gray"))
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                else:
                    root.after(0, lambda: status_label.config(text="오류 발생", foreground="#B71C1C"))
                    root.after(0, lambda msg=err_msg: messagebox.showerror("오류", msg))
            finally:
                _global_running.clear()
                root.after(0, lambda: run_button.config(state="normal"))
                root.after(0, lambda: cancel_button.config(state="disabled"))
                root.after(0, lambda: close_button.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    run_button.config(command=on_run)
    tab1.columnconfigure(1, weight=1)

    # ===== 탭 2: 송장번호 입력 =====
    tab2 = ttk.Frame(notebook, padding=12)
    notebook.add(tab2, text="송장번호 입력")
    tab2.columnconfigure(0, weight=1)

    # ── 공유 조회 기간 (대한통운 발송처리 · 벌크 발송처리 공통) ──
    d_date_lf = ttk.LabelFrame(tab2, text="조회 기간", padding=(8, 4))
    d_date_lf.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 8))
    d_date_lf.columnconfigure(1, weight=1)
    d_date_lf.columnconfigure(4, weight=0)

    ttk.Label(d_date_lf, text="시작").grid(row=0, column=0, sticky="w", padx=(0, 6))
    d_start_date_entry = tk.Entry(d_date_lf, width=13, justify="center", **_ENTRY_KW)
    d_start_date_entry.grid(row=0, column=1, sticky="ew", ipady=2)
    d_start_time_entry = tk.Entry(d_date_lf, width=9, justify="center", **_ENTRY_KW)
    d_start_time_entry.grid(row=0, column=2, padx=(6, 0), ipady=2)

    ttk.Label(d_date_lf, text="종료").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(5, 0))
    d_end_date_entry = tk.Entry(d_date_lf, width=13, justify="center", **_ENTRY_KW)
    d_end_date_entry.grid(row=1, column=1, sticky="ew", pady=(5, 0), ipady=2)
    d_end_time_entry = tk.Entry(d_date_lf, width=9, justify="center", **_ENTRY_KW)
    d_end_time_entry.grid(row=1, column=2, padx=(6, 0), pady=(5, 0), ipady=2)

    _ph_add(d_start_date_entry, "YYYY-MM-DD")
    _ph_add(d_start_time_entry, "시간 (선택)")
    _ph_add(d_end_date_entry,   "YYYY-MM-DD")
    _ph_add(d_end_time_entry,   "시간 (선택)")

    def d_set_today():
        today = datetime.today().strftime("%Y-%m-%d")
        _ph_set_date(d_start_date_entry, today); _ph_set_date(d_end_date_entry, today)

    def d_set_yesterday():
        day = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        _ph_set_date(d_start_date_entry, day); _ph_set_date(d_end_date_entry, day)

    def d_set_this_month():
        today = datetime.today()
        first = today.replace(day=1).strftime("%Y-%m-%d")
        if today.month == 12:
            last = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        _ph_set_date(d_start_date_entry, first)
        _ph_set_date(d_end_date_entry,   last.strftime("%Y-%m-%d"))

    def d_set_last_month():
        today = datetime.today()
        last = today.replace(day=1) - timedelta(days=1)
        first = last.replace(day=1)
        _ph_set_date(d_start_date_entry, first.strftime("%Y-%m-%d"))
        _ph_set_date(d_end_date_entry,   last.strftime("%Y-%m-%d"))

    def d_set_this_year():
        today = datetime.today()
        _ph_set_date(d_start_date_entry, f"{today.year}-01-01")
        _ph_set_date(d_end_date_entry,   f"{today.year}-12-31")

    def d_set_last_year():
        y = datetime.today().year - 1
        _ph_set_date(d_start_date_entry, f"{y}-01-01")
        _ph_set_date(d_end_date_entry,   f"{y}-12-31")

    def d_clear_dates():
        for e in (d_start_date_entry, d_start_time_entry, d_end_date_entry, d_end_time_entry):
            e.delete(0, tk.END)
            e.insert(0, _ph_map[e])
            e.config(fg=_PH_COLOR)

    d_date_btn_frame = ttk.Frame(d_date_lf)
    d_date_btn_frame.grid(row=0, column=4, rowspan=2, sticky="nsew", padx=(10, 0))
    for _c in range(4):
        d_date_btn_frame.columnconfigure(_c, weight=1, uniform="ddb")

    ttk.Button(d_date_btn_frame, text="오늘",   command=d_set_today,      width=_bw).grid(row=0, column=0, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(d_date_btn_frame, text="이번달", command=d_set_this_month, width=_bw).grid(row=0, column=1, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(d_date_btn_frame, text="올해",   command=d_set_this_year,  width=_bw).grid(row=0, column=2, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(d_date_btn_frame, text="어제",   command=d_set_yesterday,  width=_bw).grid(row=1, column=0, sticky="ew", padx=_px)
    ttk.Button(d_date_btn_frame, text="저번달", command=d_set_last_month, width=_bw).grid(row=1, column=1, sticky="ew", padx=_px)
    ttk.Button(d_date_btn_frame, text="작년",   command=d_set_last_year,  width=_bw).grid(row=1, column=2, sticky="ew", padx=_px)
    ttk.Button(d_date_btn_frame, text="초기화", command=d_clear_dates,    width=_bw).grid(row=0, column=3, rowspan=2, sticky="nsew")

    def append_log(msg: str):
        def _do():
            log_text.config(state="normal")
            log_text.insert(tk.END, msg + "\n")
            log_text.see(tk.END)
            log_text.config(state="disabled")
        root.after(0, _do)

    def delivery_gui_progress(pct: int, msg: str):
        root.after(0, lambda p=pct: delivery_progress_var.set(p))
        root.after(0, lambda m=msg: delivery_status_label.config(
            text=m, foreground="green" if pct == 100 else "red"
        ))

    # ── 대한통운 파일접수내역 발송처리 ──
    cj_delivery_lf = ttk.LabelFrame(tab2, text="대한통운 파일접수내역 발송처리", padding=(8, 4))
    cj_delivery_lf.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 6))
    cj_delivery_lf.columnconfigure(0, weight=1)

    detail_path_var = tk.StringVar()
    _cj_file_row = ttk.Frame(cj_delivery_lf)
    _cj_file_row.grid(row=0, column=0, sticky="ew")
    _cj_file_row.columnconfigure(0, weight=1)
    detail_entry = ttk.Entry(_cj_file_row, textvariable=detail_path_var)
    detail_entry.grid(row=0, column=0, sticky="ew")

    def browse_detail():
        paths = filedialog.askopenfilenames(
            title="파일접수 내역 발송처리 파일 선택 (여러 파일 선택 가능)",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if paths:
            detail_path_var.set(";".join(paths))

    ttk.Button(_cj_file_row, text="찾아보기", command=browse_detail).grid(row=0, column=1, sticky="w", padx=(4, 0))
    ttk.Label(cj_delivery_lf,
              text="※ 대한통운 사이트에서 운송장 출력 후 엑셀 파일로 다운받아 첨부해 주세요.",
              foreground="gray", font=("Malgun Gothic", 8)
              ).grid(row=1, column=0, sticky="w", pady=(2, 4))

    d_action_frame = ttk.Frame(cj_delivery_lf)
    d_action_frame.grid(row=2, column=0, sticky="ew", pady=(0, 2))
    delivery_cancel_event = threading.Event()
    delivery_cancel_btn = ttk.Button(d_action_frame, text="취소", state="disabled",
                                     command=lambda: delivery_cancel_event.set())
    delivery_cancel_btn.pack(side="right", padx=(4, 0))
    delivery_run_btn = ttk.Button(d_action_frame, text="실행")
    delivery_run_btn.pack(side="right")

    def on_delivery_run():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭이 현재 실행 중입니다.\n완료 후 다시 시도해주세요.")
            return
        if not ensure_admin_credentials(root):
            return
        detail_paths = [p.strip() for p in detail_path_var.get().split(";") if p.strip()]
        if not detail_paths:
            messagebox.showerror("입력 오류", "파일접수 내역 발송처리 파일을 선택해주세요.")
            return
        for path in detail_paths:
            if not os.path.exists(path):
                messagebox.showerror("파일 오류", f"파일을 찾을 수 없어요:\n{path}")
                return

        try:
            start_date = normalize_date_input(" ".join(filter(None, [_ph_get(d_start_date_entry), _ph_get(d_start_time_entry)])))
            end_date = normalize_date_input(" ".join(filter(None, [_ph_get(d_end_date_entry), _ph_get(d_end_time_entry)])))
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return

        _global_running.set()
        delivery_cancel_event.clear()
        delivery_run_btn.config(state="disabled")
        delivery_cancel_btn.config(state="normal")
        delivery_progress_var.set(0)
        delivery_status_label.config(text="실행 중...", foreground="red")
        log_text.config(state="normal")
        log_text.delete("1.0", tk.END)
        log_text.config(state="disabled")

        def worker():
            _log_lines = []

            def _log_capture(msg: str):
                _log_lines.append(msg)
                append_log(msg)

            def _save_delivery_log():
                try:
                    _save_dir = output_dir_var.get().strip()
                    if not (_save_dir and os.path.isdir(_save_dir)):
                        _save_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
                                    else os.path.dirname(os.path.abspath(__file__))
                    _ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    _log_path = os.path.join(_save_dir, f"{_ts}_발송처리_로그.txt")
                    with open(_log_path, "w", encoding="utf-8") as _lf:
                        _lf.write("\n".join(_log_lines))
                    append_log(f"[로그 저장] {_log_path}")
                except Exception as _e:
                    append_log(f"[로그 저장 실패] {_e}")

            try:
                for detail_path in detail_paths:
                    if delivery_cancel_event.is_set():
                        break
                    _log_capture(f"\n=== 파일 처리 중: {os.path.basename(detail_path)} ===")
                    run_delivery_job(
                        detail_path,
                        start_date=start_date,
                        end_date=end_date,
                        log_callback=_log_capture,
                        progress_callback=delivery_gui_progress,
                        cancel_check=delivery_cancel_event.is_set,
                    )
                _save_delivery_log()
                if delivery_cancel_event.is_set():
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소됐어요."))
                else:
                    root.after(0, lambda: messagebox.showinfo("완료", "모든 파일 처리가 완료됐어요."))
            except Exception as e:
                _save_delivery_log()
                err_msg = str(e)
                if isinstance(e, NoWorkFound):
                    root.after(0, lambda m=err_msg: messagebox.showinfo("조회 결과 없음", m))
                    root.after(0, lambda: delivery_status_label.config(text="처리할 대상 없음", foreground="gray"))
                    root.after(0, lambda: delivery_progress_var.set(100))
                elif is_user_cancel_message(err_msg):
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                    root.after(0, lambda: delivery_status_label.config(text="취소됨", foreground="gray"))
                else:
                    root.after(0, lambda m=err_msg: messagebox.showerror("오류", m))
                    root.after(0, lambda: delivery_status_label.config(text="오류 발생", foreground="red"))
            finally:
                _global_running.clear()
                root.after(0, lambda: delivery_run_btn.config(state="normal"))
                root.after(0, lambda: delivery_cancel_btn.config(state="disabled"))

        threading.Thread(target=worker, daemon=True).start()

    delivery_run_btn.config(command=on_delivery_run)

    # ── 벌크 발송처리 ──
    bulk_frame = ttk.LabelFrame(tab2, text="벌크 발송처리 (고래한방 · 스캔 기반)", padding=(8, 4))
    bulk_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 6))
    bulk_frame.columnconfigure(1, weight=1)

    _bulk_scan_result = {}

    # CJ 파일 업로드
    ttk.Label(bulk_frame, text="CJ 파일:").grid(row=0, column=0, sticky="w")
    bk_file_var = tk.StringVar()
    ttk.Entry(bulk_frame, textvariable=bk_file_var, width=28).grid(row=0, column=1, sticky="ew")
    ttk.Button(bulk_frame, text="찾아보기", command=lambda: bk_file_var.set(
        filedialog.askopenfilename(title="CJ 파일 선택", filetypes=[("Excel", "*.xlsx *.xls")]) or bk_file_var.get()
    )).grid(row=0, column=2, sticky="w", padx=(4, 0))

    # 지점 선택 + 운송장 직접 입력 (한 줄)
    ttk.Label(bulk_frame, text="또는 운송장번호 직접 입력 (/ 구분):",
              foreground="gray").grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 2))

    # 지점별 운송장 문자열 저장 (콤보 전환 시 교체)
    _bk_branch_values = {b: "" for b in GORAE_BRANCHES}
    _bk_branch_counts = {b: 0 for b in GORAE_BRANCHES}
    _bk_current_branch = {"name": GORAE_BRANCHES[0]}

    bk_input_frame = ttk.Frame(bulk_frame)
    bk_input_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=2)
    bk_input_frame.columnconfigure(5, weight=1)  # 운송장 입력칸이 가로 확장

    bk_branch_var = tk.StringVar(value=GORAE_BRANCHES[0])
    bk_branch_combo = ttk.Combobox(
        bk_input_frame, textvariable=bk_branch_var,
        values=list(GORAE_BRANCHES), state="readonly", width=8
    )
    bk_branch_combo.grid(row=0, column=0, sticky="w")
    bk_branch_count_label = ttk.Label(bk_input_frame, text="(-)", foreground="gray")
    bk_branch_count_label.grid(row=0, column=1, sticky="w", padx=(6, 4))

    # 박스당 주문 수 — 스캔 결과 그룹핑 단위 (CJ 양식 생성 · 발송처리 공통 적용)
    bk_box_size_var = tk.StringVar(value=str(BULK_BOX_SIZE))
    ttk.Label(bk_input_frame, text="박스당").grid(row=0, column=2, sticky="w", padx=(4, 2))
    bk_box_size_spin = ttk.Spinbox(
        bk_input_frame, from_=1, to=50, width=4, textvariable=bk_box_size_var
    )
    bk_box_size_spin.grid(row=0, column=3, sticky="w")
    ttk.Label(bk_input_frame, text="개").grid(row=0, column=4, sticky="w", padx=(1, 8))

    bk_tracking_entry = ttk.Entry(bk_input_frame)
    bk_tracking_entry.grid(row=0, column=5, sticky="ew")

    def _get_box_size() -> int:
        """bk_box_size_var 값을 읽어 int 반환 (오류 시 BULK_BOX_SIZE 기본값)"""
        try:
            return max(1, int(bk_box_size_var.get() or BULK_BOX_SIZE))
        except ValueError:
            return BULK_BOX_SIZE

    def _make_branch_groups(scan_result: dict, box_size: int) -> dict:
        """flat 스캔 결과를 박스 크기 기준으로 그룹핑"""
        return {
            b: [orders[i:i + box_size] for i in range(0, len(orders), box_size)]
            for b, orders in scan_result.items() if orders
        }

    def _bk_refresh_branch_label():
        b = bk_branch_var.get()
        n = _bk_branch_counts.get(b, 0)
        if n:
            box_size = _get_box_size()
            import math
            num_tracking = math.ceil(n / box_size)
            bk_branch_count_label.config(
                text=f"({n}개 · 운송장 {num_tracking}개 필요)", foreground="gray"
            )
        else:
            bk_branch_count_label.config(text="(없음)", foreground="gray")

    def _bk_on_branch_change(*_):
        # 이전 지점 입력 저장
        old = _bk_current_branch["name"]
        _bk_branch_values[old] = bk_tracking_entry.get()
        # 새 지점 입력 로드
        new = bk_branch_var.get()
        _bk_current_branch["name"] = new
        bk_tracking_entry.delete(0, tk.END)
        bk_tracking_entry.insert(0, _bk_branch_values.get(new, ""))
        _bk_refresh_branch_label()

    bk_branch_combo.bind("<<ComboboxSelected>>", _bk_on_branch_change)
    # 박스당 개수 바뀌면 라벨도 즉시 갱신
    bk_box_size_var.trace_add("write", lambda *_: _bk_refresh_branch_label())

    # 호환성을 위한 dict (기존 API 유지) — 레이블은 모두 동일 위젯, 엔트리는 가상
    bk_branch_labels = {b: bk_branch_count_label for b in GORAE_BRANCHES}
    bk_branch_entries = {}  # on_bk_ship 에서 _bk_branch_values 로 직접 조회

    bk_status_label = ttk.Label(bulk_frame, text="", foreground="blue")
    bk_status_label.grid(row=3, column=0, columnspan=3, sticky="w", pady=(4, 0))

    # 발송처리 실행 + 취소 버튼 (우측 정렬)
    bk_action_frame = ttk.Frame(bulk_frame)
    bk_action_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(4, 2))
    bk_cancel_event = threading.Event()
    bk_cancel_btn = ttk.Button(bk_action_frame, text="취소", state="disabled",
                                command=lambda: bk_cancel_event.set())
    bk_cancel_btn.pack(side="right", padx=(4, 0))
    bk_ship_btn = ttk.Button(bk_action_frame, text="발송처리 실행", state="disabled")
    bk_ship_btn.pack(side="right")

    def bk_prog(pct, msg):
        fg = "green" if pct == 100 else "#E65100"
        root.after(0, lambda p=pct: delivery_progress_var.set(p))
        root.after(0, lambda m=msg, c=fg: delivery_status_label.config(text=m, foreground=c))
        root.after(0, lambda m=msg, c=fg: bk_status_label.config(text=m, foreground=c))

    def _update_scan_ui(result: dict):
        """스캔 완료 후 UI 반영"""
        _bulk_scan_result.clear()
        _bulk_scan_result.update(result)
        for b in GORAE_BRANCHES:
            _bk_branch_counts[b] = len(result.get(b, []))
        _bk_refresh_branch_label()
        parts = [f"{b} {len(result.get(b, []))}개" for b in GORAE_BRANCHES if result.get(b)]
        bk_status_label.config(
            text=f"스캔 결과: {' / '.join(parts) if parts else '없음'}",
            foreground="blue" if parts else "gray"
        )
        bk_ship_btn.config(state="normal" if result else "disabled")
        bk_cj_btn.config(state="normal" if result else "disabled")

    def _reset_bulk_scan_state(show_message: bool = True):
        _bulk_scan_result.clear()
        for b in GORAE_BRANCHES:
            _bk_branch_values[b] = ""
            _bk_branch_counts[b] = 0
        _bk_current_branch["name"] = GORAE_BRANCHES[0]
        bk_branch_var.set(GORAE_BRANCHES[0])
        bk_tracking_entry.delete(0, tk.END)
        bk_file_var.set("")
        _bk_refresh_branch_label()
        bk_ship_btn.config(state="disabled")
        bk_cj_btn.config(state="disabled")
        if show_message:
            bk_status_label.config(text="스캔 결과 초기화됨", foreground="gray")

    def do_bk_scan():
        """상단 조회기간을 읽어 벌크 스캔 실행"""
        if not ensure_admin_credentials(root):
            return
        try:
            sd = normalize_date_input(" ".join(filter(None, [_ph_get(d_start_date_entry), _ph_get(d_start_time_entry)])))
            ed = normalize_date_input(" ".join(filter(None, [_ph_get(d_end_date_entry), _ph_get(d_end_time_entry)])))
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭 실행 중입니다.")
            return
        _global_running.set()
        bk_scan_btn.config(state="disabled")
        bk_status_label.config(text="스캔 중...", foreground="#E65100")
        delivery_progress_var.set(0)
        log_text.config(state="normal")
        log_text.delete("1.0", tk.END)
        log_text.config(state="disabled")

        def worker():
            try:
                result = run_bulk_scan(sd, ed, log_callback=append_log, progress_callback=bk_prog)
                root.after(0, lambda: _update_scan_ui(result))
            except Exception as e:
                err_msg = str(e)
                if isinstance(e, NoWorkFound):
                    root.after(0, lambda m=err_msg: messagebox.showinfo("조회 결과 없음", m))
                    root.after(0, lambda: bk_status_label.config(text="벌크 주문 없음", foreground="gray"))
                    root.after(0, lambda: delivery_progress_var.set(100))
                elif is_user_cancel_message(err_msg):
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                    root.after(0, lambda: bk_status_label.config(text="취소됨", foreground="gray"))
                else:
                    root.after(0, lambda m=err_msg: messagebox.showerror("스캔 오류", m))
                    root.after(0, lambda: bk_status_label.config(text="스캔 오류", foreground="red"))
            finally:
                _global_running.clear()
                root.after(0, lambda: bk_scan_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def do_bk_cj():
        if not _bulk_scan_result:
            messagebox.showwarning("스캔 필요", "먼저 스캔을 실행해주세요.")
            return
        box_size = _get_box_size()
        branch_groups = _make_branch_groups(_bulk_scan_result, box_size)
        save_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
            else os.path.dirname(os.path.abspath(__file__))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(save_dir, f"{ts}_벌크_대한통운_업로드양식.xlsx")
        try:
            build_bulk_cj_upload_df(branch_groups).to_excel(out_path, index=False)
            messagebox.showinfo("저장 완료", f"저장:\n{out_path}\n(박스당 {box_size}개 기준)")
            if sys.platform == "win32":
                os.startfile(save_dir)
            elif sys.platform == "darwin":
                os.system(f"open '{save_dir}'")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    # 왼쪽: 스캔 / CJ양식 / 초기화 | 오른쪽: 발송처리 실행 / 취소
    bk_scan_btn = ttk.Button(bk_action_frame, text="스캔", command=do_bk_scan)
    bk_scan_btn.pack(side="left")
    bk_cj_btn = ttk.Button(bk_action_frame, text="CJ 양식 생성", state="disabled", command=do_bk_cj)
    bk_cj_btn.pack(side="left", padx=(4, 0))
    ttk.Button(bk_action_frame, text="초기화", command=lambda: _reset_bulk_scan_state()).pack(side="left", padx=(4, 0))

    def on_bk_ship():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭 실행 중입니다.")
            return
        if not ensure_admin_credentials(root):
            return
        if not _bulk_scan_result:
            messagebox.showwarning("스캔 필요", "먼저 스캔을 실행해주세요.")
            return

        # 박스 크기 기준으로 flat 스캔 결과를 그룹핑
        box_size = _get_box_size()
        branch_groups_for_ship = _make_branch_groups(_bulk_scan_result, box_size)

        tracking_by_group = {}
        cj_file = bk_file_var.get().strip()

        if cj_file and os.path.exists(cj_file):
            try:
                df_cj = pd.read_excel(cj_file, header=0, dtype=str)
                tracking_col = next((c for c in df_cj.columns if "운송장번호" in clean_text(str(c))), None)
                ordercode_col = next((c for c in df_cj.columns if "고객주문번호" in clean_text(str(c))), None)
                if not tracking_col:
                    messagebox.showerror("오류", "CJ 파일에서 운송장번호 컬럼을 찾을 수 없어요.")
                    return
                for _, row in df_cj.iterrows():
                    t = normalize_tracking_no(str(row.get(tracking_col, "")))
                    gid = clean_text(str(row.get(ordercode_col, ""))) if ordercode_col else ""
                    if t and gid:
                        tracking_by_group[gid] = t
                bk_status_label.config(text=f"CJ 파일에서 {len(tracking_by_group)}개 운송장 읽음")
            except Exception as e:
                messagebox.showerror("파일 오류", str(e))
                return
        else:
            # 현재 엔트리에 보이는 값도 저장 (사용자가 콤보를 바꾸지 않고 발송처리 누를 수 있음)
            _bk_branch_values[_bk_current_branch["name"]] = bk_tracking_entry.get()
            for branch in GORAE_BRANCHES:
                groups = branch_groups_for_ship.get(branch, [])
                raw = _bk_branch_values.get(branch, "").strip()
                if not raw:
                    continue
                nums = [normalize_tracking_no(t) for t in raw.split("/") if t.strip()]
                for gi, num in enumerate(nums, start=1):
                    if num and gi <= len(groups):
                        tracking_by_group[f"고래{branch}_{gi}"] = num

        if not tracking_by_group:
            messagebox.showwarning("입력 필요", "CJ 파일 또는 운송장 번호를 입력해주세요.")
            return

        _global_running.set()
        bk_cancel_event.clear()
        bk_ship_btn.config(state="disabled")
        bk_cancel_btn.config(state="normal")

        def worker():
            try:
                run_bulk_ship(branch_groups_for_ship, tracking_by_group,
                              progress_callback=bk_prog,
                              cancel_check=bk_cancel_event.is_set)
                if bk_cancel_event.is_set():
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소됐어요."))
                else:
                    root.after(0, lambda: messagebox.showinfo("완료", "벌크 발송처리가 완료됐어요."))
            except Exception as e:
                err_msg = str(e)
                if isinstance(e, NoWorkFound):
                    root.after(0, lambda m=err_msg: messagebox.showinfo("처리 대상 없음", m))
                    root.after(0, lambda: bk_status_label.config(text="발송 처리할 대상 없음", foreground="gray"))
                elif is_user_cancel_message(err_msg):
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                    root.after(0, lambda: bk_status_label.config(text="취소됨", foreground="gray"))
                else:
                    root.after(0, lambda m=err_msg: messagebox.showerror("오류", m))
            finally:
                _global_running.clear()
                root.after(0, lambda: bk_ship_btn.config(state="normal"))
                root.after(0, lambda: bk_cancel_btn.config(state="disabled"))

        threading.Thread(target=worker, daemon=True).start()

    bk_ship_btn.config(command=on_bk_ship)

    # ── 약속처방 송장번호 입력 ──
    sp_delivery_lf = ttk.LabelFrame(tab2, text="약속처방 송장번호 입력", padding=(8, 4))
    sp_delivery_lf.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 6))
    sp_delivery_lf.columnconfigure(0, weight=1)

    sp_detail_path_var = tk.StringVar()
    _sp_file_row = ttk.Frame(sp_delivery_lf)
    _sp_file_row.grid(row=0, column=0, sticky="ew")
    _sp_file_row.columnconfigure(0, weight=1)
    sp_detail_entry = ttk.Entry(_sp_file_row, textvariable=sp_detail_path_var)
    sp_detail_entry.grid(row=0, column=0, sticky="ew")

    def sp_browse_detail():
        paths = filedialog.askopenfilenames(
            title="파일접수 내역 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if paths:
            sp_detail_path_var.set(";".join(paths))

    ttk.Button(_sp_file_row, text="찾아보기", command=sp_browse_detail).grid(row=0, column=1, sticky="w", padx=(4, 0))
    ttk.Label(sp_delivery_lf,
              text="※ 대한통운 파일접수 엑셀의 주문번호로 약속처방 주문을 매칭해 송장을 입력합니다.",
              foreground="gray", font=("Malgun Gothic", 8)
              ).grid(row=1, column=0, sticky="w", pady=(2, 4))

    sp_d_action_frame = ttk.Frame(sp_delivery_lf)
    sp_d_action_frame.grid(row=2, column=0, sticky="ew", pady=(0, 2))
    sp_delivery_cancel_event = threading.Event()
    sp_delivery_cancel_btn = ttk.Button(sp_d_action_frame, text="취소", state="disabled",
                                        command=lambda: sp_delivery_cancel_event.set())
    sp_delivery_cancel_btn.pack(side="right", padx=(4, 0))
    sp_delivery_run_btn = ttk.Button(sp_d_action_frame, text="실행")
    sp_delivery_run_btn.pack(side="right")

    def on_sp_delivery_run():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭이 현재 실행 중입니다.\n완료 후 다시 시도해주세요.")
            return
        if not ensure_admin_credentials(root):
            return
        sp_paths = [p.strip() for p in sp_detail_path_var.get().split(";") if p.strip()]
        if not sp_paths:
            messagebox.showerror("입력 오류", "파일접수 내역 파일을 선택해주세요.")
            return
        for path in sp_paths:
            if not os.path.exists(path):
                messagebox.showerror("파일 오류", f"파일을 찾을 수 없어요:\n{path}")
                return
        try:
            start_date = normalize_date_input(" ".join(filter(None, [_ph_get(d_start_date_entry), _ph_get(d_start_time_entry)])))
            end_date   = normalize_date_input(" ".join(filter(None, [_ph_get(d_end_date_entry), _ph_get(d_end_time_entry)])))
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return

        _global_running.set()
        sp_delivery_cancel_event.clear()
        sp_delivery_run_btn.config(state="disabled")
        sp_delivery_cancel_btn.config(state="normal")
        delivery_progress_var.set(0)
        delivery_status_label.config(text="실행 중...", foreground="red")
        log_text.config(state="normal")
        log_text.delete("1.0", tk.END)
        log_text.config(state="disabled")

        def worker():
            try:
                for sp_path in sp_paths:
                    if sp_delivery_cancel_event.is_set():
                        break
                    append_log(f"\n=== 파일 처리 중: {os.path.basename(sp_path)} ===")
                    run_sp_delivery_job(
                        sp_path,
                        start_date=start_date,
                        end_date=end_date,
                        log_callback=append_log,
                        progress_callback=delivery_gui_progress,
                        cancel_check=sp_delivery_cancel_event.is_set,
                    )
                if sp_delivery_cancel_event.is_set():
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소됐어요."))
                else:
                    root.after(0, lambda: messagebox.showinfo("완료", "약속처방 송장 입력이 완료됐어요."))
            except Exception as e:
                err_msg = str(e)
                if isinstance(e, NoWorkFound):
                    root.after(0, lambda m=err_msg: messagebox.showinfo("조회 결과 없음", m))
                    root.after(0, lambda: delivery_status_label.config(text="처리할 대상 없음", foreground="gray"))
                    root.after(0, lambda: delivery_progress_var.set(100))
                elif is_user_cancel_message(err_msg):
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                    root.after(0, lambda: delivery_status_label.config(text="취소됨", foreground="gray"))
                else:
                    root.after(0, lambda m=err_msg: messagebox.showerror("오류", m))
                    root.after(0, lambda: delivery_status_label.config(text="오류 발생", foreground="red"))
            finally:
                _global_running.clear()
                root.after(0, lambda: sp_delivery_run_btn.config(state="normal"))
                root.after(0, lambda: sp_delivery_cancel_btn.config(state="disabled"))

        threading.Thread(target=worker, daemon=True).start()

    sp_delivery_run_btn.config(command=on_sp_delivery_run)

    # ── 공유 상태창 + 진행바 (세 섹션 모두 여기 출력) ──
    delivery_status_label = ttk.Label(tab2, text="대기 중", foreground="blue")
    delivery_status_label.grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 2))

    delivery_progress_var = tk.IntVar(value=0)
    delivery_progress_bar = ttk.Progressbar(
        tab2, orient="horizontal", length=300,
        mode="determinate", maximum=100, variable=delivery_progress_var
    )
    delivery_progress_bar.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(0, 4))

    # ── 로그창 ──
    log_text = tk.Text(tab2, height=7, width=50, state="disabled", font=("Malgun Gothic", 9))
    log_text.grid(row=6, column=0, columnspan=2, sticky="nsew", pady=(4, 4))

    log_scroll = ttk.Scrollbar(tab2, orient="vertical", command=log_text.yview)
    log_scroll.grid(row=6, column=2, sticky="ns")
    log_text.config(yscrollcommand=log_scroll.set)

    def _log_scroll_wheel(event):
        log_text.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"
    log_text.bind("<MouseWheel>", _log_scroll_wheel)
    tab2.rowconfigure(6, weight=1)

    # ===== 탭 3: 배송완료 전환 =====
    tab3 = ttk.Frame(notebook, padding=12)
    notebook.add(tab3, text="배송완료 전환")

    c_date_lf = ttk.LabelFrame(tab3, text="조회 기간", padding=(8, 4))
    c_date_lf.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 8))
    c_date_lf.columnconfigure(1, weight=1)
    c_date_lf.columnconfigure(4, weight=0)

    ttk.Label(c_date_lf, text="시작").grid(row=0, column=0, sticky="w", padx=(0, 6))
    c_start_date_entry = tk.Entry(c_date_lf, width=13, justify="center", **_ENTRY_KW)
    c_start_date_entry.grid(row=0, column=1, sticky="ew", ipady=2)
    c_start_time_entry = tk.Entry(c_date_lf, width=9, justify="center", **_ENTRY_KW)
    c_start_time_entry.grid(row=0, column=2, padx=(6, 0), ipady=2)

    ttk.Label(c_date_lf, text="종료").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(5, 0))
    c_end_date_entry = tk.Entry(c_date_lf, width=13, justify="center", **_ENTRY_KW)
    c_end_date_entry.grid(row=1, column=1, sticky="ew", pady=(5, 0), ipady=2)
    c_end_time_entry = tk.Entry(c_date_lf, width=9, justify="center", **_ENTRY_KW)
    c_end_time_entry.grid(row=1, column=2, padx=(6, 0), pady=(5, 0), ipady=2)

    _ph_add(c_start_date_entry, "YYYY-MM-DD")
    _ph_add(c_start_time_entry, "시간 (선택)")
    _ph_add(c_end_date_entry,   "YYYY-MM-DD")
    _ph_add(c_end_time_entry,   "시간 (선택)")

    def c_set_today():
        today = datetime.today().strftime("%Y-%m-%d")
        _ph_set_date(c_start_date_entry, today); _ph_set_date(c_end_date_entry, today)

    def c_set_yesterday():
        day = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        _ph_set_date(c_start_date_entry, day); _ph_set_date(c_end_date_entry, day)

    def c_set_this_month():
        today = datetime.today()
        first = today.replace(day=1).strftime("%Y-%m-%d")
        if today.month == 12:
            last = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        _ph_set_date(c_start_date_entry, first)
        _ph_set_date(c_end_date_entry,   last.strftime("%Y-%m-%d"))

    def c_set_last_month():
        today = datetime.today()
        last = today.replace(day=1) - timedelta(days=1)
        first = last.replace(day=1)
        _ph_set_date(c_start_date_entry, first.strftime("%Y-%m-%d"))
        _ph_set_date(c_end_date_entry,   last.strftime("%Y-%m-%d"))

    def c_set_this_year():
        today = datetime.today()
        _ph_set_date(c_start_date_entry, f"{today.year}-01-01")
        _ph_set_date(c_end_date_entry,   f"{today.year}-12-31")

    def c_set_last_year():
        y = datetime.today().year - 1
        _ph_set_date(c_start_date_entry, f"{y}-01-01")
        _ph_set_date(c_end_date_entry,   f"{y}-12-31")

    def c_clear_dates():
        for e in (c_start_date_entry, c_start_time_entry, c_end_date_entry, c_end_time_entry):
            e.delete(0, tk.END)
            e.insert(0, _ph_map[e])
            e.config(fg=_PH_COLOR)

    c_date_btn_frame = ttk.Frame(c_date_lf)
    c_date_btn_frame.grid(row=0, column=4, rowspan=2, sticky="nsew", padx=(10, 0))
    for _c in range(4):
        c_date_btn_frame.columnconfigure(_c, weight=1, uniform="cdb")

    ttk.Button(c_date_btn_frame, text="오늘",   command=c_set_today,      width=_bw).grid(row=0, column=0, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(c_date_btn_frame, text="이번달", command=c_set_this_month, width=_bw).grid(row=0, column=1, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(c_date_btn_frame, text="올해",   command=c_set_this_year,  width=_bw).grid(row=0, column=2, sticky="ew", padx=_px, pady=(0, 3))
    ttk.Button(c_date_btn_frame, text="어제",   command=c_set_yesterday,  width=_bw).grid(row=1, column=0, sticky="ew", padx=_px)
    ttk.Button(c_date_btn_frame, text="저번달", command=c_set_last_month, width=_bw).grid(row=1, column=1, sticky="ew", padx=_px)
    ttk.Button(c_date_btn_frame, text="작년",   command=c_set_last_year,  width=_bw).grid(row=1, column=2, sticky="ew", padx=_px)
    ttk.Button(c_date_btn_frame, text="초기화", command=c_clear_dates,    width=_bw).grid(row=0, column=3, rowspan=2, sticky="nsew")

    c_tangjeon_lf = ttk.LabelFrame(tab3, text="탕전주문 배송완료 전환", padding=(8, 4))
    c_tangjeon_lf.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 6))
    c_ignore_status_var = tk.BooleanVar(value=False)
    ttk.Checkbutton(c_tangjeon_lf, text="상태 무관 조회 (배송중 외 상태도 포함, 버그 복구용)",
                    variable=c_ignore_status_var).pack(anchor="w")
    c_action_frame = ttk.Frame(c_tangjeon_lf)
    c_action_frame.pack(fill="x")
    c_cancel_event = threading.Event()
    c_cancel_btn = ttk.Button(c_action_frame, text="취소", state="disabled",
                              command=lambda: c_cancel_event.set())
    c_cancel_btn.pack(side="right", padx=(4, 0))
    c_run_btn_outer = ttk.Button(c_action_frame, text="실행")
    c_run_btn_outer.pack(side="right")

    # ── 약속처방 배송완료 전환 ──
    sp_complete_lf = ttk.LabelFrame(tab3, text="약속처방 배송완료 전환", padding=(8, 4))
    sp_complete_lf.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 6))
    sp_c_action_frame = ttk.Frame(sp_complete_lf)
    sp_c_action_frame.pack(fill="x")
    sp_complete_cancel_event = threading.Event()
    sp_complete_cancel_btn = ttk.Button(sp_c_action_frame, text="취소", state="disabled",
                                        command=lambda: sp_complete_cancel_event.set())
    sp_complete_cancel_btn.pack(side="right", padx=(4, 0))
    sp_complete_run_btn = ttk.Button(sp_c_action_frame, text="실행")
    sp_complete_run_btn.pack(side="right")

    c_status_label = ttk.Label(tab3, text="대기 중", foreground="blue")
    c_status_label.grid(row=3, column=0, columnspan=2, sticky="w", pady=(4, 4))

    c_progress_var = tk.IntVar(value=0)
    c_progress_bar = ttk.Progressbar(
        tab3, orient="horizontal", length=380,
        mode="determinate", maximum=100, variable=c_progress_var
    )
    c_progress_bar.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(0, 8))

    c_log_text = tk.Text(tab3, height=14, width=50, state="disabled", font=("Malgun Gothic", 9))
    c_log_text.grid(row=5, column=0, columnspan=2, sticky="nsew", pady=(0, 8))

    c_log_scroll = ttk.Scrollbar(tab3, orient="vertical", command=c_log_text.yview)
    c_log_scroll.grid(row=5, column=2, sticky="ns")
    c_log_text.config(yscrollcommand=c_log_scroll.set)

    # disabled 상태에서도 마우스 휠 스크롤 가능하도록 바인딩
    def _c_log_scroll_wheel(event):
        c_log_text.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"
    c_log_text.bind("<MouseWheel>", _c_log_scroll_wheel)   # Windows / macOS

    def c_append_log(msg: str):
        def _do():
            c_log_text.config(state="normal")
            c_log_text.insert(tk.END, msg + "\n")
            c_log_text.see(tk.END)
            c_log_text.config(state="disabled")
        root.after(0, _do)

    def c_gui_progress(pct: int, msg: str):
        root.after(0, lambda p=pct: c_progress_var.set(p))
        root.after(0, lambda m=msg: c_status_label.config(
            text=m, foreground="green" if pct == 100 else "red"
        ))

    def on_complete_run():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭이 현재 실행 중입니다.\n완료 후 다시 시도해주세요.")
            return
        if not ensure_admin_credentials(root):
            return
        start_date = ""
        end_date = ""
        try:
            start_date = normalize_date_input(" ".join(filter(None, [_ph_get(c_start_date_entry), _ph_get(c_start_time_entry)])))
            end_date = normalize_date_input(" ".join(filter(None, [_ph_get(c_end_date_entry), _ph_get(c_end_time_entry)])))
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return

        _global_running.set()
        c_cancel_event.clear()
        c_run_btn_outer.config(state="disabled")
        c_cancel_btn.config(state="normal")
        c_progress_var.set(0)
        c_status_label.config(text="실행 중...", foreground="red")
        c_log_text.config(state="normal")
        c_log_text.delete("1.0", tk.END)
        c_log_text.config(state="disabled")

        def worker():
            try:
                run_complete_job(
                    start_date=start_date,
                    end_date=end_date,
                    ignore_status=c_ignore_status_var.get(),
                    log_callback=c_append_log,
                    progress_callback=c_gui_progress,
                    cancel_check=c_cancel_event.is_set,
                )
                if c_cancel_event.is_set():
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소됐어요."))
                else:
                    root.after(0, lambda: messagebox.showinfo("완료", "배송완료 전환이 완료됐어요."))
            except Exception as e:
                err_msg = str(e)
                if isinstance(e, NoWorkFound):
                    root.after(0, lambda m=err_msg: messagebox.showinfo("조회 결과 없음", m))
                    root.after(0, lambda: c_status_label.config(text="전환 대상 없음", foreground="gray"))
                    root.after(0, lambda: c_progress_var.set(100))
                elif is_user_cancel_message(err_msg):
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                    root.after(0, lambda: c_status_label.config(text="취소됨", foreground="gray"))
                else:
                    root.after(0, lambda m=err_msg: messagebox.showerror("오류", m))
                    root.after(0, lambda: c_status_label.config(text="오류 발생", foreground="red"))
            finally:
                _global_running.clear()
                root.after(0, lambda: c_run_btn_outer.config(state="normal"))
                root.after(0, lambda: c_cancel_btn.config(state="disabled"))

        threading.Thread(target=worker, daemon=True).start()

    c_run_btn_outer.config(command=on_complete_run)

    def on_sp_complete_run():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭이 현재 실행 중입니다.\n완료 후 다시 시도해주세요.")
            return
        if not ensure_admin_credentials(root):
            return
        try:
            start_date = normalize_date_input(" ".join(filter(None, [_ph_get(c_start_date_entry), _ph_get(c_start_time_entry)])))
            end_date   = normalize_date_input(" ".join(filter(None, [_ph_get(c_end_date_entry), _ph_get(c_end_time_entry)])))
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return

        _global_running.set()
        sp_complete_cancel_event.clear()
        sp_complete_run_btn.config(state="disabled")
        sp_complete_cancel_btn.config(state="normal")
        c_progress_var.set(0)
        c_status_label.config(text="실행 중...", foreground="red")
        c_log_text.config(state="normal")
        c_log_text.delete("1.0", tk.END)
        c_log_text.config(state="disabled")

        def worker():
            try:
                run_sp_complete_job(
                    start_date=start_date,
                    end_date=end_date,
                    log_callback=c_append_log,
                    progress_callback=c_gui_progress,
                    cancel_check=sp_complete_cancel_event.is_set,
                )
                if sp_complete_cancel_event.is_set():
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소됐어요."))
                else:
                    root.after(0, lambda: messagebox.showinfo("완료", "약속처방 배송완료 전환이 완료됐어요."))
            except Exception as e:
                err_msg = str(e)
                if isinstance(e, NoWorkFound):
                    root.after(0, lambda m=err_msg: messagebox.showinfo("조회 결과 없음", m))
                    root.after(0, lambda: c_status_label.config(text="전환 대상 없음", foreground="gray"))
                    root.after(0, lambda: c_progress_var.set(100))
                elif is_user_cancel_message(err_msg):
                    root.after(0, lambda: messagebox.showinfo("작업 취소", "요청하신 작업을 중단했어요."))
                    root.after(0, lambda: c_status_label.config(text="취소됨", foreground="gray"))
                else:
                    root.after(0, lambda m=err_msg: messagebox.showerror("오류", m))
                    root.after(0, lambda: c_status_label.config(text="오류 발생", foreground="red"))
            finally:
                _global_running.clear()
                root.after(0, lambda: sp_complete_run_btn.config(state="normal"))
                root.after(0, lambda: sp_complete_cancel_btn.config(state="disabled"))

        threading.Thread(target=worker, daemon=True).start()

    sp_complete_run_btn.config(command=on_sp_complete_run)
    tab3.rowconfigure(5, weight=1)
    tab3.columnconfigure(0, weight=1)

    # ===== 탭 4: 약속처방 주문관리 =====
    tab4 = ttk.Frame(notebook, padding=12)
    notebook.add(tab4, text="약속처방 주문관리")
    tab4.columnconfigure(0, weight=1)

    # ── 진행상태 ──
    _SP_STATUSES = ["주문처리중", "배송준비", "배송중", "완료", "예약발송", "환불/취소"]
    sp_status_lf = ttk.LabelFrame(tab4, text="진행상태", padding=(8, 4))
    sp_status_lf.grid(row=0, column=0, sticky="ew", pady=(0, 4))
    for _c in range(4):
        sp_status_lf.columnconfigure(_c, weight=1, uniform="spcb")

    sp_status_vars = {}
    for i, s in enumerate(_SP_STATUSES):
        var = tk.BooleanVar(value=True)
        sp_status_vars[s] = var
        ttk.Checkbutton(sp_status_lf, text=s, variable=var).grid(
            row=i // 4, column=i % 4, sticky="w", padx=(0, 4), pady=2
        )
    sp_status_lf.columnconfigure(4, weight=0)
    sp_sbtn_frame = ttk.Frame(sp_status_lf)
    sp_sbtn_frame.grid(row=0, column=4, rowspan=2, sticky="e")
    ttk.Button(sp_sbtn_frame, text="전체선택", width=7,
               command=lambda: [v.set(True) for v in sp_status_vars.values()]).grid(row=0, column=0, sticky="ew", pady=(0, 3))
    ttk.Button(sp_sbtn_frame, text="전체해제", width=7,
               command=lambda: [v.set(False) for v in sp_status_vars.values()]).grid(row=1, column=0, sticky="ew")

    # ── 주문경로 ──
    sp_route_lf = ttk.LabelFrame(tab4, text="주문경로", padding=(8, 4))
    sp_route_lf.grid(row=1, column=0, sticky="ew", pady=(0, 4))
    sp_route_vars = {}
    for _ri, _rt in enumerate(["H", "C"]):
        _rv = tk.BooleanVar(value=True)
        sp_route_vars[_rt] = _rv
        _label = "H (사이트 직접주문)" if _rt == "H" else "C (카페 주문)"
        ttk.Checkbutton(sp_route_lf, text=_label, variable=_rv).grid(
            row=0, column=_ri, sticky="w", padx=(0, 16)
        )

    # ── 조회 기간 ──
    sp_date_lf = ttk.LabelFrame(tab4, text="조회 기간", padding=(8, 4))
    sp_date_lf.grid(row=2, column=0, sticky="ew", pady=(0, 4))
    sp_date_lf.columnconfigure(1, weight=1)

    ttk.Label(sp_date_lf, text="시작").grid(row=0, column=0, sticky="w", padx=(0, 6))
    sp_start_entry = tk.Entry(sp_date_lf, width=13, justify="center", **_ENTRY_KW)
    sp_start_entry.grid(row=0, column=1, sticky="ew", ipady=2)
    ttk.Label(sp_date_lf, text="종료").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(5, 0))
    sp_end_entry = tk.Entry(sp_date_lf, width=13, justify="center", **_ENTRY_KW)
    sp_end_entry.grid(row=1, column=1, sticky="ew", pady=(5, 0), ipady=2)
    _ph_add(sp_start_entry, "YYYY-MM-DD")
    _ph_add(sp_end_entry,   "YYYY-MM-DD")

    sp_date_btn_frame = ttk.Frame(sp_date_lf)
    sp_date_btn_frame.grid(row=0, column=2, rowspan=2, sticky="nsew", padx=(10, 0))
    for _c in range(4):
        sp_date_btn_frame.columnconfigure(_c, weight=1, uniform="spdb")

    def sp_set_today():
        t = datetime.today().strftime("%Y-%m-%d")
        _ph_set_date(sp_start_entry, t); _ph_set_date(sp_end_entry, t)
    def sp_set_yesterday():
        d = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        _ph_set_date(sp_start_entry, d); _ph_set_date(sp_end_entry, d)
    def sp_set_this_month():
        today = datetime.today()
        first = today.replace(day=1).strftime("%Y-%m-%d")
        if today.month == 12:
            last = today.replace(year=today.year+1, month=1, day=1) - timedelta(days=1)
        else:
            last = today.replace(month=today.month+1, day=1) - timedelta(days=1)
        _ph_set_date(sp_start_entry, first); _ph_set_date(sp_end_entry, last.strftime("%Y-%m-%d"))
    def sp_set_last_month():
        last = datetime.today().replace(day=1) - timedelta(days=1)
        _ph_set_date(sp_start_entry, last.replace(day=1).strftime("%Y-%m-%d"))
        _ph_set_date(sp_end_entry, last.strftime("%Y-%m-%d"))
    def sp_set_this_year():
        y = datetime.today().year
        _ph_set_date(sp_start_entry, f"{y}-01-01"); _ph_set_date(sp_end_entry, f"{y}-12-31")
    def sp_set_last_year():
        y = datetime.today().year - 1
        _ph_set_date(sp_start_entry, f"{y}-01-01"); _ph_set_date(sp_end_entry, f"{y}-12-31")
    def sp_clear_dates():
        for e in (sp_start_entry, sp_end_entry):
            e.delete(0, tk.END); e.insert(0, _ph_map[e]); e.config(fg=_PH_COLOR)

    _bw2 = 7
    ttk.Button(sp_date_btn_frame, text="오늘",   command=sp_set_today,      width=_bw2).grid(row=0, column=0, sticky="ew", padx=(0,2), pady=(0,3))
    ttk.Button(sp_date_btn_frame, text="이번달", command=sp_set_this_month, width=_bw2).grid(row=0, column=1, sticky="ew", padx=(0,2), pady=(0,3))
    ttk.Button(sp_date_btn_frame, text="올해",   command=sp_set_this_year,  width=_bw2).grid(row=0, column=2, sticky="ew", padx=(0,2), pady=(0,3))
    ttk.Button(sp_date_btn_frame, text="어제",   command=sp_set_yesterday,  width=_bw2).grid(row=1, column=0, sticky="ew", padx=(0,2))
    ttk.Button(sp_date_btn_frame, text="저번달", command=sp_set_last_month, width=_bw2).grid(row=1, column=1, sticky="ew", padx=(0,2))
    ttk.Button(sp_date_btn_frame, text="작년",   command=sp_set_last_year,  width=_bw2).grid(row=1, column=2, sticky="ew", padx=(0,2))
    ttk.Button(sp_date_btn_frame, text="초기화", command=sp_clear_dates,    width=_bw2).grid(row=0, column=3, rowspan=2, sticky="nsew")

    # ── 필터 ──
    sp_filter_lf = ttk.LabelFrame(tab4, text="필터", padding=(8, 4))
    sp_filter_lf.grid(row=3, column=0, sticky="ew", pady=(0, 4))
    sp_filter_lf.columnconfigure(1, weight=1)

    sp_search_target_var = tk.StringVar(value="주문자")
    sp_search_filter_var = tk.StringVar(value="")
    ttk.Label(sp_filter_lf, text="검색대상").grid(row=0, column=0, sticky="w", padx=(0, 8))
    sp_search_row = ttk.Frame(sp_filter_lf)
    sp_search_row.grid(row=0, column=1, sticky="ew")
    sp_search_row.columnconfigure(1, weight=1)
    ttk.Combobox(sp_search_row, textvariable=sp_search_target_var,
                 values=["주문자", "수취인", "한의원명"], state="readonly", width=10
                 ).grid(row=0, column=0, sticky="w", padx=(0, 6))
    sp_search_entry = tk.Entry(sp_search_row, textvariable=sp_search_filter_var, **_ENTRY_KW)
    sp_search_entry.grid(row=0, column=1, sticky="ew", ipady=2)
    ttk.Button(sp_filter_lf, text="지우기",
               command=lambda: sp_search_filter_var.set("")).grid(row=0, column=2, padx=(4, 0))

    ttk.Label(sp_filter_lf, text="비우면 전체 조회",
              foreground="gray", font=("Malgun Gothic", 8)).grid(
        row=1, column=0, columnspan=3, sticky="w", pady=(3, 0))

    # ── 파일 저장 ──
    sp_save_lf = ttk.LabelFrame(tab4, text="파일 저장", padding=(8, 4))
    sp_save_lf.grid(row=4, column=0, sticky="ew", pady=(0, 4))
    sp_save_excel_var = tk.BooleanVar(value=True)
    sp_save_cj_var    = tk.BooleanVar(value=False)
    ttk.Checkbutton(sp_save_lf, text="주문마스터 엑셀", variable=sp_save_excel_var).grid(row=0, column=0, sticky="w")
    ttk.Checkbutton(sp_save_lf, text="대한통운 파일 업로드 양식", variable=sp_save_cj_var).grid(row=0, column=1, sticky="w", padx=(16, 0))

    # ── 상품 내역 ──
    sp_goods_lf = ttk.LabelFrame(tab4, text="상품 내역", padding=(8, 4))
    sp_goods_lf.grid(row=5, column=0, sticky="ew", pady=(0, 4))
    sp_goods_lf.columnconfigure(0, weight=1)

    sp_goods_status_var = tk.StringVar(value="캐시 없음")
    ttk.Label(sp_goods_lf, textvariable=sp_goods_status_var, foreground="gray",
              font=("Malgun Gothic", 9)).grid(row=0, column=0, sticky="w")

    # 상품 캐시는 항상 자동 사용. 업데이트는 [상품 내역 관리] 다이얼로그에서 수동으로.

    def _refresh_goods_status_label():
        _c, _m = _load_goods_tax_cache()
        if _c:
            sp_goods_status_var.set(
                f"캐시: {_m.get('count', len(_c))}개 상품 ({_m.get('updated_at', '?')} 기준)"
            )
        else:
            sp_goods_status_var.set("캐시 없음")

    def _open_goods_dialog():
        dlg = tk.Toplevel(root)
        dlg.title("상품 내역 관리")
        dlg.geometry("640x520")
        dlg.resizable(False, True)
        dlg.transient(root)
        dlg.grab_set()

        # 캐시 상태
        info_lf = ttk.LabelFrame(dlg, text="캐시 상태", padding=(8, 4))
        info_lf.pack(fill="x", padx=8, pady=(8, 4))
        dlg_cache_var = tk.StringVar()
        ttk.Label(info_lf, textvariable=dlg_cache_var).pack(anchor="w")

        # 상품 목록
        list_lf = ttk.LabelFrame(dlg, text="상품 목록", padding=(8, 4))
        list_lf.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        dlg_goods_text = tk.Text(list_lf, state="disabled",
                                  font=("Malgun Gothic", 9), wrap="none")
        dlg_goods_text.pack(side="left", fill="both", expand=True)
        dlg_goods_scroll = ttk.Scrollbar(list_lf, orient="vertical", command=dlg_goods_text.yview)
        dlg_goods_scroll.pack(side="right", fill="y")
        dlg_goods_text.configure(yscrollcommand=dlg_goods_scroll.set)

        # 로그 (다운로드 진행 시)
        dlg_log_var = tk.StringVar(value="")
        ttk.Label(dlg, textvariable=dlg_log_var, foreground="gray",
                  font=("Malgun Gothic", 8)).pack(anchor="w", padx=8)

        # 버튼 행
        btn_f = ttk.Frame(dlg)
        btn_f.pack(fill="x", padx=8, pady=(0, 8))
        dlg_update_btn = ttk.Button(btn_f, text="사이트 다운로드")
        dlg_update_btn.pack(side="left", padx=(0, 4))
        dlg_export_btn = ttk.Button(btn_f, text="전체 내역 엑셀")
        dlg_export_btn.pack(side="left", padx=(0, 4))
        ttk.Button(btn_f, text="닫기", command=dlg.destroy).pack(side="right")

        def _dlg_refresh():
            _c, _m = _load_goods_tax_cache()
            if _c:
                updated_at = _m.get("updated_at", "?")
                count = _m.get("count", len(_c))
                dlg_cache_var.set(f"마지막 업데이트: {updated_at}  |  상품 수: {count}개")
                dlg_goods_text.configure(state="normal")
                dlg_goods_text.delete("1.0", "end")
                for name, info in _c.items():
                    공급가액 = info.get("공급가액", "?")
                    면세금액 = info.get("면세금액", "?")
                    과세구분 = info.get("과세구분", "?")
                    dlg_goods_text.insert("end",
                        f"{name}  |  공급가액={공급가액}  면세금액={면세금액}  →  {과세구분}\n")
                dlg_goods_text.configure(state="disabled")
            else:
                dlg_cache_var.set("캐시 없음 — 다운로드 버튼을 눌러 가져오세요.")
                dlg_goods_text.configure(state="normal")
                dlg_goods_text.delete("1.0", "end")
                dlg_goods_text.configure(state="disabled")
            _refresh_goods_status_label()

        _dlg_refresh()

        def _do_download():
            if _global_running.is_set():
                messagebox.showwarning("실행 중", "다른 작업이 실행 중입니다.\n완료 후 다시 시도해주세요.", parent=dlg)
                return
            if not ensure_admin_credentials(root):
                return
            dlg_update_btn.config(state="disabled", text="다운로드 중...")
            dlg_log_var.set("Chrome 드라이버 시작 중...")

            def _thread():
                _drv = None
                try:
                    from selenium.webdriver.chrome.options import Options as _COptions
                    _opts = _COptions()
                    _opts.add_argument("--start-maximized")
                    _drv = webdriver.Chrome(options=_opts)
                    login_driver(_drv, ADMIN_ID, ADMIN_PW)
                    root.after(0, lambda: dlg_log_var.set("로그인 완료, 상품 목록 수집 중..."))

                    def _log_cb(msg):
                        root.after(0, lambda _m=msg: dlg_log_var.set(_m))

                    result = _fetch_all_product_tax_map(_drv, log=_log_cb)
                    if result:
                        _save_goods_tax_cache(result)
                        root.after(0, lambda: dlg_log_var.set(f"완료: {len(result)}개 상품 저장"))

                        # GitHub 배포 여부 자동 제안
                        def _after_save(_cnt=len(result)):
                            ans = messagebox.askyesno(
                                "다운로드 완료",
                                f"{_cnt}개 상품 정보 저장 완료!\n\nGitHub에 배포해서 다른 직원들에게 공유할까요?",
                                parent=dlg,
                            )
                            if ans:
                                _do_github_deploy_silent(dlg, dlg_log_var)

                        root.after(0, _after_save)
                    else:
                        root.after(0, lambda: dlg_log_var.set("상품 정보를 가져오지 못했습니다."))
                        root.after(0, lambda: messagebox.showwarning(
                            "주의", "상품 정보를 가져오지 못했습니다.", parent=dlg))
                except Exception as _e:
                    _em = str(_e)
                    root.after(0, lambda: dlg_log_var.set(f"오류: {_em}"))
                    root.after(0, lambda: messagebox.showerror("오류", _em, parent=dlg))
                finally:
                    if _drv:
                        try:
                            _drv.quit()
                        except Exception:
                            pass
                    root.after(0, lambda: dlg_update_btn.config(
                        state="normal", text="사이트 다운로드"))
                    root.after(0, _dlg_refresh)

            threading.Thread(target=_thread, daemon=True).start()

        dlg_update_btn.config(command=_do_download)

        def _do_export_excel():
            if _global_running.is_set():
                messagebox.showwarning("실행 중", "다른 작업이 실행 중입니다.\n완료 후 다시 시도해주세요.", parent=dlg)
                return
            if not ensure_admin_credentials(root):
                return
            # 저장 경로 먼저 선택
            save_path = filedialog.asksaveasfilename(
                parent=dlg,
                title="상품 전체 내역 저장",
                defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
                initialfile=f"상품전체내역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
            if not save_path:
                return

            dlg_export_btn.config(state="disabled", text="수집 중...")
            dlg_log_var.set("Chrome 드라이버 시작 중...")

            def _thread():
                _drv = None
                try:
                    from selenium.webdriver.chrome.options import Options as _COptions
                    _opts = _COptions()
                    _opts.add_argument("--start-maximized")
                    _drv = webdriver.Chrome(options=_opts)
                    login_driver(_drv, ADMIN_ID, ADMIN_PW)
                    root.after(0, lambda: dlg_log_var.set("로그인 완료, 약속처방 상품 수집 중..."))

                    def _log_cb(msg):
                        root.after(0, lambda _m=msg: dlg_log_var.set(_m))

                    shop_rows = _fetch_shop_product_all_details(_drv, log=_log_cb)
                    root.after(0, lambda: dlg_log_var.set(f"약속처방 {len(shop_rows)}개 완료 → 카페 상품 수집 중..."))
                    # 약속처방 결과로 shop_map 구성 (카페 퍼지 매칭용)
                    _shop_map_for_cafe = {
                        r["상품명"]: {
                            "공급가액": r.get("처방비용_공급가액") or 0,
                            "부가세":   r.get("처방비용_부가세")   or 0,
                            "면세금액": r.get("처방비용_면세금액") or 0,
                            "과세구분": r.get("처방비용_과세구분") or "",
                        }
                        for r in shop_rows if r.get("상품명") and r.get("처방비용_과세구분")
                    }
                    cafe_rows = _fetch_cafe_product_all_details(_drv, shop_map=_shop_map_for_cafe, log=_log_cb)

                    rows = shop_rows + cafe_rows
                    if rows:
                        import pandas as _pd
                        df = _pd.DataFrame(rows)

                        # ── 열 통합/정리 ──
                        # 상품ID: p_seq(약속처방) + seqno(카페) → 단일 열
                        df["상품ID"] = df.get("p_seq", _pd.Series(dtype=str)).fillna("").astype(str).replace("nan", "")
                        if "seqno" in df.columns:
                            mask = df["상품ID"].eq("")
                            df.loc[mask, "상품ID"] = df.loc[mask, "seqno"].fillna("").astype(str).replace("nan", "")
                        df["상품ID"] = df["상품ID"].replace("", _pd.NA)

                        # 분류: 상품분류(약속처방) + 분류(카페) → 단일 열
                        df["_분류merged"] = df.get("상품분류", _pd.Series(dtype=str)).fillna("").astype(str).replace("nan", "")
                        if "분류" in df.columns:
                            mask2 = df["_분류merged"].eq("")
                            df.loc[mask2, "_분류merged"] = df.loc[mask2, "분류"].fillna("").astype(str).replace("nan", "")
                        df["분류"] = df["_분류merged"].replace("", _pd.NA)
                        df.drop(columns=["_분류merged"], inplace=True)

                        # 불필요 열 제거
                        _drop = {"p_seq", "seqno", "번호", "판매가", "처방명", "처방비용", "상품분류"}
                        df.drop(columns=[c for c in _drop if c in df.columns], inplace=True)

                        # ── 목표 열 순서 ──
                        _fixed = ["구분", "상품명", "상품ID", "분류",
                                  "처방상태", "활성",
                                  "처방비용_총액", "처방비용_공급가액", "처방비용_부가세",
                                  "처방비용_면세금액", "처방비용_과세구분",
                                  "할인_N개이상", "할인_금액",
                                  "등록일"]
                        _err_cols  = [c for c in df.columns if c.startswith("_")]
                        _front     = [c for c in _fixed if c in df.columns]
                        _rest      = [c for c in df.columns if c not in _front and c not in _err_cols]
                        df = df[_front + _rest + _err_cols]
                        df.to_excel(save_path, index=False)
                        _total = len(rows)
                        root.after(0, lambda: dlg_log_var.set(
                            f"저장 완료: 약속처방 {len(shop_rows)}개 + 카페 {len(cafe_rows)}개 = {_total}개 → {os.path.basename(save_path)}"))
                        root.after(0, lambda: messagebox.showinfo(
                            "저장 완료",
                            f"약속처방 {len(shop_rows)}개 + 카페 {len(cafe_rows)}개\n총 {_total}개 상품 전체 내역을 저장했습니다.\n\n{save_path}",
                            parent=dlg,
                        ))
                    else:
                        root.after(0, lambda: dlg_log_var.set("상품 정보를 가져오지 못했습니다."))
                        root.after(0, lambda: messagebox.showwarning("주의", "상품 정보를 가져오지 못했습니다.", parent=dlg))
                except Exception as _e:
                    _em = str(_e)
                    root.after(0, lambda: dlg_log_var.set(f"오류: {_em}"))
                    root.after(0, lambda: messagebox.showerror("오류", _em, parent=dlg))
                finally:
                    if _drv:
                        try:
                            _drv.quit()
                        except Exception:
                            pass
                    root.after(0, lambda: dlg_export_btn.config(state="normal", text="전체 내역 엑셀"))

            threading.Thread(target=_thread, daemon=True).start()

        dlg_export_btn.config(command=_do_export_excel)

        # ── GitHub 배포 (다운로드 완료 후 자동 제안용) ──
        def _do_github_deploy_silent(parent_dlg, log_var):
            """다운로드 완료 후 자동으로 호출 — PAT 확인 후 GitHub 업로드."""
            from tkinter import simpledialog as _sd
            token = _read_github_token()
            if not token:
                token = _sd.askstring(
                    "GitHub Personal Access Token",
                    "GitHub PAT를 입력하세요. (repo 권한 필요)\n한 번 저장하면 이후엔 자동으로 사용됩니다.",
                    parent=parent_dlg,
                )
                if not token:
                    return
                token = token.strip()
                _write_github_token(token)

            log_var.set("GitHub 배포 중...")

            def _t(_tok=token):
                msg = _upload_cache_to_github(
                    _tok,
                    log=lambda m: root.after(0, lambda _m=m: log_var.set(_m))
                )
                if "완료" in msg:
                    root.after(0, lambda: messagebox.showinfo("배포 완료", msg, parent=parent_dlg))
                else:
                    root.after(0, lambda: messagebox.showerror("배포 실패", msg, parent=parent_dlg))

            threading.Thread(target=_t, daemon=True).start()

    ttk.Button(sp_goods_lf, text="상품 내역 관리", command=_open_goods_dialog).grid(
        row=0, column=2, sticky="e", padx=(0, 0))

    # 시작 시 캐시 상태 표시
    _refresh_goods_status_label()

    # ── 저장 위치 (탭1과 공유) ──
    sp_out_frame = ttk.Frame(tab4)
    sp_out_frame.grid(row=6, column=0, sticky="ew", pady=(0, 4))
    sp_out_frame.columnconfigure(1, weight=1)
    ttk.Label(sp_out_frame, text="저장 위치:").grid(row=0, column=0, sticky="w", padx=(0, 6))
    ttk.Entry(sp_out_frame, textvariable=output_dir_var).grid(row=0, column=1, sticky="ew", padx=(0, 4))
    ttk.Button(sp_out_frame, text="찾아보기", command=_browse_output_dir, width=8).grid(row=0, column=2)

    # ── 상태 + 진행바 ──
    sp_status_label = ttk.Label(tab4, text="대기 중", foreground="gray")
    sp_status_label.grid(row=7, column=0, sticky="w", pady=(4, 2))

    sp_progress_var = tk.IntVar(value=0)
    sp_progress = ttk.Progressbar(tab4, orient="horizontal", mode="determinate",
                                   maximum=100, variable=sp_progress_var)
    sp_progress.grid(row=8, column=0, sticky="ew", pady=(0, 6))

    # ── 버튼 행 ──
    sp_btn_frame = ttk.Frame(tab4)
    sp_btn_frame.grid(row=9, column=0, sticky="ew", pady=(4, 0))

    sp_cancel_event = threading.Event()
    sp_run_btn = ttk.Button(sp_btn_frame, text="실행")
    sp_run_btn.pack(side="left", padx=(0, 4))
    sp_cancel_btn = ttk.Button(sp_btn_frame, text="취소", state="disabled",
                               command=lambda: sp_cancel_event.set())
    sp_cancel_btn.pack(side="left", padx=(0, 4))

    sp_open_folder_btn = ttk.Button(sp_btn_frame, text="출력 폴더 열기", state="disabled")
    sp_open_folder_btn.pack(side="right", padx=(4, 0))
    sp_last_result = {"text": ""}
    sp_show_result_btn = ttk.Button(sp_btn_frame, text="결과 다시 보기", state="disabled",
                                    command=lambda: messagebox.showinfo("마지막 결과", sp_last_result["text"]))
    sp_show_result_btn.pack(side="right", padx=(4, 0))

    # ── 로그 ──
    sp_log_text = tk.Text(tab4, height=12, width=70, state="disabled", font=("Malgun Gothic", 9))
    sp_log_text.grid(row=10, column=0, sticky="nsew", pady=(8, 0))
    sp_log_scroll = ttk.Scrollbar(tab4, orient="vertical", command=sp_log_text.yview)
    sp_log_scroll.grid(row=10, column=1, sticky="ns")
    sp_log_text.configure(yscrollcommand=sp_log_scroll.set)
    tab4.rowconfigure(10, weight=1)

    def _sp_log_wheel(event):
        sp_log_text.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"
    sp_log_text.bind("<MouseWheel>", _sp_log_wheel)

    def sp_log(msg: str):
        def _do():
            sp_log_text.configure(state="normal")
            sp_log_text.insert(tk.END, msg + "\n")
            sp_log_text.see(tk.END)
            sp_log_text.configure(state="disabled")
        root.after(0, _do)

    def sp_update_progress(pct: int, msg: str):
        sp_progress_var.set(pct)
        if pct == 100:
            fg, bar_style = "#2E7D32", "Done.Horizontal.TProgressbar"
        elif pct == 0:
            fg, bar_style = "#B71C1C", "Error.Horizontal.TProgressbar"
        else:
            fg, bar_style = "#E65100", None
        sp_status_label.config(text=msg, foreground=fg)
        if bar_style:
            sp_progress.config(style=bar_style)

    def on_sp_run():
        if _global_running.is_set():
            messagebox.showwarning("실행 중", "다른 탭이 현재 실행 중입니다.\n완료 후 다시 시도해주세요.")
            return
        if not ensure_admin_credentials(root):
            return

        # 상품 캐시 확인 — 없으면 미리 물어봄
        _cached_check, _ = _load_goods_tax_cache()
        _download_goods_first = False
        if not _cached_check:
            ans = messagebox.askyesnocancel(
                "상품 캐시 없음",
                "저장된 상품 정보가 없습니다.\n\n"
                "[예]  지금 상품 정보 다운로드 후 주문 조회 진행\n"
                "[아니오]  면세/과세 열 비운 채로 진행\n"
                "[취소]  중단",
            )
            if ans is None:   # 취소
                return
            _download_goods_first = bool(ans)  # True=예, False=아니오

        # 날짜 값 읽기: YYYY-MM-DD 그대로 (사이트 입력 필드 포맷)
        start_raw = _ph_get(sp_start_entry).strip()
        end_raw   = _ph_get(sp_end_entry).strip()

        # 진행상태: 체크된 것들 → 각각 별도 요청 (사이트는 단일 상태 필터만 지원)
        # 체크된 상태코드 목록 수집
        _sc_map = {"주문처리중":"1","배송준비":"2","배송중":"3","완료":"4","예약발송":"6","환불/취소":"5"}
        checked_codes = [_sc_map[s] for s, v in sp_status_vars.items() if v.get()]
        # 전부 체크 or 전부 미체크 → 전체("")로 처리
        if len(checked_codes) == len(_SP_STATUSES) or len(checked_codes) == 0:
            checked_codes = [""]

        # 검색 필터
        _st_map = {"주문자":"name","수취인":"r_name","한의원명":"han_name"}
        search_type = _st_map.get(sp_search_target_var.get(), "name")
        search_str  = sp_search_filter_var.get().strip()

        # 주문경로 필터: 둘 다 체크되거나 둘 다 미체크면 전체 (빈 set)
        _route_checked = {r for r, v in sp_route_vars.items() if v.get()}
        _route_filter = _route_checked if 0 < len(_route_checked) < 2 else set()

        sp_settings = {
            "start_date": start_raw,
            "end_date": end_raw,
            "status_codes": checked_codes,
            "route_filter": _route_filter,
            "search_type": search_type,
            "search_str": search_str,
            "save_excel": sp_save_excel_var.get(),
            "save_cj": sp_save_cj_var.get(),
            "output_dir": output_dir_var.get().strip(),
            "max_pages": 100,
            "cancel_event": sp_cancel_event,
            "download_goods_first": _download_goods_first,
        }

        sp_log_text.configure(state="normal")
        sp_log_text.delete("1.0", "end")
        sp_log_text.configure(state="disabled")
        sp_progress["value"] = 0

        _global_running.set()
        sp_cancel_event.clear()
        sp_run_btn.config(state="disabled")
        sp_cancel_btn.config(state="normal")

        def run_thread():
            try:
                result = run_shop_order_job(
                    sp_settings,
                    progress_callback=lambda p, m: root.after(0, lambda _p=p, _m=m: sp_update_progress(_p, _m)),
                    log_callback=lambda m: root.after(0, lambda _m=m: sp_log(_m)),
                )
                if sp_cancel_event.is_set():
                    root.after(0, lambda: sp_update_progress(0, "취소됨"))
                    root.after(0, lambda: messagebox.showinfo("취소", "작업이 취소됐어요."))
                elif result:
                    sp_last_result["text"] = f"저장 위치:\n{result}"
                    root.after(0, lambda: sp_update_progress(100, "완료"))
                    root.after(0, lambda: sp_show_result_btn.config(state="normal"))
                    if os.path.isdir(result):
                        root.after(0, lambda d=result: sp_open_folder_btn.config(
                            state="normal",
                            command=lambda _d=d: (
                                os.startfile(_d) if sys.platform == "win32"
                                else os.system(f"open '{_d}'")
                            )
                        ))
                        # 완료 시 자동으로 폴더 열기
                        def _auto_open(d=result):
                            if sys.platform == "win32":
                                os.startfile(d)
                            elif sys.platform == "darwin":
                                os.system(f"open '{d}'")
                        root.after(0, _auto_open)
                    root.after(0, lambda r=result: messagebox.showinfo("완료", f"저장 완료!\n{r}"))
            except Exception as e:
                err = str(e)
                root.after(0, lambda: sp_update_progress(0, "오류 발생"))
                root.after(0, lambda: messagebox.showerror("오류", err))
            finally:
                _global_running.clear()
                root.after(0, lambda: sp_run_btn.config(state="normal"))
                root.after(0, lambda: sp_cancel_btn.config(state="disabled"))
                root.after(0, _refresh_goods_status_label)  # 캐시 갱신 후 상태 라벨 업데이트

        threading.Thread(target=run_thread, daemon=True).start()

    sp_run_btn.config(command=on_sp_run)

    # 앱 시작 2초 후 업데이트 체크 (UI 로딩 완료 후 실행)
    root.after(2000, lambda: check_and_prompt_update(root))

    # 앱 시작 4초 후 GitHub 상품 캐시 자동 동기화 (백그라운드, 조용히)
    def _startup_gh_cache_sync():
        result = _sync_cache_from_github()
        if "업데이트 완료" in result:
            root.after(0, _refresh_goods_status_label)

    root.after(4000, lambda: threading.Thread(target=_startup_gh_cache_sync, daemon=True).start())

    root.mainloop()


if __name__ == "__main__":
    try:
        launch_gui()
    except Exception as _startup_exc:
        _write_startup_error(_startup_exc)
        try:
            _root = tk.Tk()
            _root.withdraw()
            messagebox.showerror(
                "실행 오류",
                f"프로그램 시작 중 오류가 발생했어요.\n\n"
                f"{_startup_exc}\n\n"
                f"자세한 내용은 startup_error.log에 저장했어요.",
            )
            _root.destroy()
        except Exception:
            pass
        raise
